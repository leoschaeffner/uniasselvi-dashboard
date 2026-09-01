"""
UNIASSELVI - Dashboard de Portfolios
Le as planilhas da pasta planilhas/ e gera saida/dashboard.html

PATCHES v2 aplicados:
  1. CH SEMANAL lida de 01_CONTROLE_TUTORIA.xlsx (não depende de LOTACAO)
  2. tem_lotacao = True somente quando há ch_semanal > 0 nos tutores
  3. status_ordem usa PERIODOS_ORDENS para datas de início corretas
  4. Situação do tutor corrigida quando nenhuma ordem está vencida ainda
  5. Campo 'sit' sincronizado após deduplicação por email
  6. Prints duplicados removidos de verificar_e_localizar()
  7. datas_por_tutor adicionado no formato antigo de agendas
"""

import pandas as pd
import re
import json, os, sys, math, webbrowser, time, threading, glob, hashlib, base64, unicodedata
from pathlib import Path
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

# ── Configuração multi-semestre ─────────────────────────────────────────────
# Lida de config_semestre.json. Fallback hardcoded para 2026/1.
_SEMESTRES_DEFAULT = {
    '2026/1': {
        'prazos': {
            'Ordem 1': '14/03/2026', 'Ordem 2': '11/04/2026',
            'Ordem 3': '09/05/2026', 'Ordem 4': '06/06/2026', 'Ordem 5': '04/07/2026',
        },
        'periodos': {
            'Ordem 1': {'inicio': '16/02/2026', 'fim': '14/03/2026', 'semanas': 4},
            'Ordem 2': {'inicio': '16/03/2026', 'fim': '11/04/2026', 'semanas': 4},
            'Ordem 3': {'inicio': '13/04/2026', 'fim': '09/05/2026', 'semanas': 4},
            'Ordem 4': {'inicio': '11/05/2026', 'fim': '06/06/2026', 'semanas': 4},
            'Ordem 5': {'inicio': '08/06/2026', 'fim': '04/07/2026', 'semanas': 4},
        },
    },
    '2026/2': {
        'prazos': {
            'Ordem 1': '22/08/2026', 'Ordem 2': '19/09/2026',
            'Ordem 3': '17/10/2026', 'Ordem 4': '14/11/2026', 'Ordem 5': '12/12/2026',
        },
        'periodos': {
            'Ordem 1': {'inicio': '27/07/2026', 'fim': '22/08/2026', 'semanas': 4},
            'Ordem 2': {'inicio': '24/08/2026', 'fim': '19/09/2026', 'semanas': 4},
            'Ordem 3': {'inicio': '21/09/2026', 'fim': '17/10/2026', 'semanas': 4},
            'Ordem 4': {'inicio': '19/10/2026', 'fim': '14/11/2026', 'semanas': 4},
            'Ordem 5': {'inicio': '16/11/2026', 'fim': '12/12/2026', 'semanas': 4},
        },
    },
}

_DISCIPLINAS_POR_ORDEM_GLOBAL = {}
def _carregar_semestres():
    import os as _os, json as _json
    _cfg_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'config_semestre.json')
    _sems = dict(_SEMESTRES_DEFAULT)
    if _os.path.isfile(_cfg_path):
        try:
            with open(_cfg_path, encoding='utf-8') as _f: _cfg = _json.load(_f)
            # config pode ter 'semestres' (dict) ou o formato antigo (semestre único)
            if 'semestres' in _cfg:
                _sems.update(_cfg['semestres'])
            elif 'semestre' in _cfg:
                _sem_key = _cfg['semestre']
                _sems[_sem_key] = {'prazos': _cfg.get('prazos', {}), 'periodos': _cfg.get('periodos', {})}
            print(f"  [CONFIG] Semestres carregados: {sorted(_sems.keys())}")
            # Carregar mapeamento de disciplinas por ordem (opcional)
            _disc_file = _cfg.get('disciplinas_por_ordem')
            if _disc_file:
                _disc_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), _disc_file)
                if _os.path.isfile(_disc_path):
                    with open(_disc_path, encoding='utf-8') as _fd:
                        _DISCIPLINAS_POR_ORDEM_GLOBAL.update(_json.load(_fd))
                    _total_disc = sum(len(v) for ordens in _DISCIPLINAS_POR_ORDEM_GLOBAL.values() for v in ordens.values())
                    print(f"  [CONFIG] Disciplinas por ordem: {_total_disc} total")
        except Exception as _e:
            print(f"  [AVISO] Erro ao ler config_semestre.json: {_e} — usando padrão")
    return _sems

ALL_SEMESTRES = _carregar_semestres()

def _data_para_semestre(data_str):
    """Dado '2026-03-15', retorna '2026/1' ou '2026/2' ou None."""
    if not data_str or data_str == 'None': return None
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(str(data_str)[:10], '%Y-%m-%d').date()
        for sem, cfg in sorted(ALL_SEMESTRES.items()):
            for ord_cfg in cfg.get('periodos', {}).values():
                try:
                    ini = _dt.strptime(ord_cfg['inicio'], '%d/%m/%Y').date()
                    fim = _dt.strptime(ord_cfg['fim'],    '%d/%m/%Y').date()
                    if ini <= d <= fim: return sem
                except: pass
    except: pass
    return None

def _ordem_relativa(ordem_forms, semestre):
    """Ordem 1 no Forms sempre = Ordem 1 do semestre em questão."""
    return ordem_forms  # As ordens são relativas dentro de cada semestre

# Para compatibilidade com o resto do código — usa o semestre mais recente como padrão
_sem_atual = sorted(ALL_SEMESTRES.keys())[-1]
PRAZOS_ORDENS   = ALL_SEMESTRES[_sem_atual]['prazos']
PERIODOS_ORDENS = ALL_SEMESTRES[_sem_atual]['periodos']
SEMESTRE_ATUAL  = _sem_atual
print(f"  [CONFIG] Semestre ativo (dashboard): {SEMESTRE_ATUAL}")
# ── Fim configuração multi-semestre ──────────────────────────────────────────
CH_ADMIN_FATOR   = 0.25
CH_PRATICA_DURAC = 1.5


def _parse_ch(v):
    """PATCH 1 helper — converte CH SEMANAL (HH:MM ou decimal) para float horas."""
    if v is None: return None
    sv = str(v).strip()
    if sv in ('', 'nan', 'NaN', 'None', '0', '0.0'): return None
    try:
        if ':' in sv:
            parts = sv.split(':')
            result = float(parts[0]) + float(parts[1]) / 60
        else:
            result = float(sv.replace(',', '.'))
        return result if result > 0 else None
    except (ValueError, TypeError):
        return None


# PATCH 30: helpers pra item 4 (Análise de Agendas) — nenhum arquivo novo é
# necessário pra isso; dia da semana e turno são derivados de DT_GERENCIADA/
# HR_GERENCIADA, que já estão no CSV que alimenta o pipeline hoje.
_DIAS_SEMANA_PT = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']  # Python weekday(): Segunda=0

def _dia_semana_pt(iso_str):
    """Dado '2026-03-15', retorna 'Domingo' (nome do dia em português) ou '' se vazio/inválido."""
    if not iso_str:
        return ''
    try:
        d = datetime.strptime(str(iso_str)[:10], '%Y-%m-%d')
        return _DIAS_SEMANA_PT[d.weekday()]
    except Exception:
        return ''

def _turno_de_horario(hr_str):
    """Extrai a hora de início de uma string tipo '19:00 - 20:30' e classifica em
    Madrugada (00h-05h59) / Manhã (06h-11h59) / Tarde (12h-17h59) / Noite (18h-23h59).
    Madrugada é o sinal-chave pro alerta de 'horário incomum' (junto com Domingo)."""
    import re as _re_turno
    if not hr_str:
        return ''
    m = _re_turno.match(r'\s*(\d{1,2}):(\d{2})', str(hr_str))
    if not m:
        return ''
    hora = int(m.group(1))
    if 0 <= hora < 6: return 'Madrugada'
    if 6 <= hora < 12: return 'Manhã'
    if 12 <= hora < 18: return 'Tarde'
    return 'Noite'


def achar_pasta_script():
    candidatos = []
    try:
        p = os.path.dirname(os.path.abspath(sys.argv[0]))
        if os.path.isdir(p): candidatos.append(p)
    except: pass
    try:
        p = os.path.dirname(os.path.abspath(__file__))
        if os.path.isdir(p): candidatos.append(p)
    except: pass
    try:
        p = os.getcwd()
        if os.path.isdir(p): candidatos.append(p)
    except: pass
    for p in candidatos:
        if os.path.isdir(os.path.join(p, "planilhas")): return p
        if os.path.isfile(os.path.join(p, "processar.py")): return p
    return candidatos[0] if candidatos else os.getcwd()


SCRIPT_DIR = achar_pasta_script()


def ler_url_file(path_url):
    try:
        with open(path_url, encoding='utf-8', errors='replace') as f:
            for line in f:
                if line.upper().startswith('URL='): return line[4:].strip()
    except: pass
    return None


def forcar_download_onedrive(path_url_file, destino, label):
    import subprocess, shutil, time
    path_xlsx = path_url_file.replace('.url', '').replace('.URL', '')
    try:
        subprocess.run(['attrib', '-P', '+U', path_url_file], capture_output=True, timeout=10)
    except Exception: pass
    for _ in range(6):
        if os.path.isfile(path_xlsx):
            with open(path_xlsx, 'rb') as f:
                header = f.read(4)
            if header == b'PK\x03\x04':
                shutil.copy2(path_xlsx, destino)
                print(f"  [OneDrive] Sincronizado: {label}")
                return destino
        time.sleep(5)
        print(f"  [OneDrive] Aguardando sync para {label}...")
    print(f"  [OneDrive] Timeout aguardando {label}.")
    return None


_KEYWORDS = {
    '01_CONTROLE_TUTORIA.xlsx': ['CONTROLE'],
    'PORTIFOLIO_TUTOR.xlsx':     ['PORTFOLIO', 'PORTIFOLIO', 'PORTF'],
    'REL_GERAL_DE_GERENCIAMENTO.xlsx': ['GERENCIAMENTO', 'REL_GERAL'],
}
_ONEDRIVE_NAMES = {
    '01_CONTROLE_TUTORIA.xlsx': ['CONTROLE'],
    'PORTIFOLIO_TUTOR.xlsx':     ['PORTF', 'PORTFOLIO'],
}


def _bate(caminho_arq, padrao):
    bn  = os.path.basename(caminho_arq).upper()
    kws = _KEYWORDS.get(padrao, [os.path.splitext(padrao)[0].upper()])
    return any(kw in bn for kw in kws)


def achar_arquivo(pasta, padrao):
    pasta_planilhas = os.path.join(pasta, "planilhas")
    direto = os.path.join(pasta_planilhas, padrao)
    if os.path.isfile(direto): return direto
    for arq in glob.glob(os.path.join(pasta_planilhas, "*.xls")) + glob.glob(os.path.join(pasta_planilhas, "*.xlsx")):
        if _bate(arq, padrao): return arq
    for arq in glob.glob(os.path.join(pasta_planilhas, "*.url")) + glob.glob(os.path.join(pasta_planilhas, "*.xlsx.url")) + glob.glob(os.path.join(pasta_planilhas, "*.xls.url")):
        if _bate(arq, padrao):
            url = ler_url_file(arq)
            if url:
                destino = os.path.join(pasta_planilhas, padrao)
                resultado = forcar_download_onedrive(arq, destino, padrao)
                if resultado: return resultado
    usuario = os.environ.get('USERNAME', os.environ.get('USER', 'leona'))
    for base in [
        f"C:\\Users\\{usuario}\\OneDrive - Uniasselvi",
        f"C:\\Users\\{usuario}\\OneDrive - UNIASSELVI",
        f"C:\\Users\\{usuario}\\OneDrive - Grupo Uniasselvi",
        f"C:\\Users\\{usuario}\\OneDrive",
    ]:
        if not os.path.isdir(base): continue
        for arq in glob.glob(os.path.join(base, "*.url")):
            if _bate(arq, padrao):
                url = ler_url_file(arq)
                if url:
                    destino = os.path.join(pasta_planilhas, padrao)
                    resultado = forcar_download_onedrive(arq, destino, padrao)
                    if resultado: return resultado
        for arq in glob.glob(os.path.join(base, "**", "*.xls"), recursive=True) + glob.glob(os.path.join(base, "**", "*.xlsx"), recursive=True):
            if _bate(arq, padrao): return arq
    return None


WATCH_MODE = len(sys.argv) > 1 and sys.argv[1].lower() == "watch"

CAT_MAP = {
    'ENF-INS (Multidisciplinar II)':
        'Multidisciplinar II - Enfermagem e Instrumentação Cirúrgica',
    'BIO-FISIO-EST-TO (Multidisciplinar III)':
        'Multidisciplinar III - Biomedicina Estética, Fisioterapia, Terapia Ocupacional e Estética e Cosmética',
    'BIO-FAR (Multidisciplinar I)':
        'Multidisciplinar I - Biomedicina e Farmácia',
    'NUTRI (Multidisciplinar IV)':
        'Multidisciplinar IV - Nutrição',
    'QUÍMICA E FÍSICA':
        'Química e Física - Agronomia',
    'ENGMAKER':
        'EngeMaker | Química e Física - Engenharias e Licenciaturas',
    'ENGMAKER+QUÍMICA E FÍSICA':
        'EngeMaker | Química e Física - Engenharias e Licenciaturas',
}

# PATCH 82: mais de uma vez já apareceu uma variante do rótulo de categoria com
# prefixo "BIO-" duplicado (ex: "BIO-BIO-FISIO-EST-TO"), vinda de uma exportação
# diferente do CONTROLE/GIOCONDA — isso virava uma "categoria fantasma" separada
# em qualquer lugar que agrupa pelo valor CRU da categoria (o filtro de
# Portfólios, por exemplo), mesmo já existindo uma entrada no CAT_MAP pra dar o
# nome de exibição certo a essa variante (band-aid que só cobria o texto
# mostrado, não a contagem/agrupamento em si). Normaliza a categoria na
# FONTE — assim ela nunca mais aparece como algo diferente da categoria certa
# em nenhum lugar do sistema, e essa proteção vale pra qualquer futura
# duplicação de prefixo "BIO-", não só esse caso específico.
def _normaliza_categoria_bio_duplicado(s):
    s2 = str(s or '').strip()
    while s2.upper().startswith('BIO-BIO-'):
        s2 = s2[4:]
    return s2




def ts():
    BRT = timezone(timedelta(hours=-3))
    return datetime.now(BRT).strftime('%H:%M:%S')


def limpar(obj):
    if isinstance(obj, dict):   return {k: limpar(v) for k, v in obj.items()}
    if isinstance(obj, list):   return [limpar(v) for v in obj]
    if isinstance(obj, float) and math.isnan(obj): return None
    return obj


# PATCH 25c: snapshot de colunas detectadas + contagens-chave, comparado contra
# a rodada anterior. Só imprime avisos no log (GitHub Actions) — nada disso
# aparece na UI do dashboard (decisão do Leo: um KPI de "match rate" visível
# pros usuários geraria mais confusão do que ajuda). Objetivo: pegar cedo o
# caso "alguém renomeou uma coluna na planilha-fonte" ou "a próxima rodada
# perdeu um monte de submissões sem ninguém notar" — exatamente o tipo de
# problema que causou o sumiço de portfólios da tutora Cleya Da Silva Santana.
_SNAPSHOT_QUEDA_LIMIAR = 0.15  # aviso se uma contagem cair mais de 15% sem explicação

def _verificar_snapshot_regressao(colunas_detectadas, contagens):
    snap_path = os.path.join(SCRIPT_DIR, 'snapshot_manifest.json')
    anterior = None
    if os.path.isfile(snap_path):
        try:
            with open(snap_path, encoding='utf-8') as f:
                anterior = json.load(f)
        except Exception as e:
            print(f"[{ts()}] [SNAPSHOT] Aviso: não consegui ler snapshot anterior ({e}) — seguindo sem comparação")

    if anterior:
        # 1) Mudança nas colunas detectadas (PATCH 25 / guardrail #4)
        cols_ant = anterior.get('colunas_detectadas', {})
        for chave_col, valor_atual in colunas_detectadas.items():
            valor_anterior = cols_ant.get(chave_col)
            if valor_anterior is not None and valor_anterior != valor_atual:
                print(f"[{ts()}] ⚠️  [SNAPSHOT] Coluna '{chave_col}' mudou de nome entre rodadas: "
                      f"{valor_anterior!r} -> {valor_atual!r}. Se isso não foi intencional, "
                      f"confira se a planilha-fonte teve o cabeçalho renomeado.")
        # 2) Queda anormal de contagem (guardrail #2)
        cont_ant = anterior.get('contagens', {})
        for chave_cont, valor_atual in contagens.items():
            valor_anterior = cont_ant.get(chave_cont)
            if isinstance(valor_anterior, (int, float)) and valor_anterior > 0:
                queda = (valor_anterior - valor_atual) / valor_anterior
                if queda > _SNAPSHOT_QUEDA_LIMIAR:
                    print(f"[{ts()}] ⚠️  [SNAPSHOT] Queda de {queda*100:.1f}% em '{chave_cont}': "
                          f"{valor_anterior} -> {valor_atual}. Pode ser problema real de dados "
                          f"(ex: matching quebrado) — vale checar antes de considerar normal.")
    else:
        print(f"[{ts()}] [SNAPSHOT] Nenhum snapshot anterior encontrado — esta rodada vira a baseline.")

    try:
        with open(snap_path, 'w', encoding='utf-8') as f:
            json.dump({'gerado_em': ts(), 'colunas_detectadas': colunas_detectadas, 'contagens': contagens}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[{ts()}] [SNAPSHOT] Aviso: não consegui salvar snapshot desta rodada ({e})")


def verificar_e_localizar():
    pasta_planilhas = os.path.join(SCRIPT_DIR, "planilhas")
    os.makedirs(pasta_planilhas, exist_ok=True)
    print(f"  Script em : {SCRIPT_DIR}")
    print(f"  Planilhas : {pasta_planilhas}")
    print()
    cfg = {}
    cfg_file = os.path.join(SCRIPT_DIR, "config_links.json")
    if os.path.isfile(cfg_file):
        try:
            with open(cfg_file, encoding="utf-8") as f: cfg = json.load(f)
        except: pass
    cam_t = cfg.get("caminho_planilha_tutores", "").strip().strip('"')
    cam_p = cfg.get("caminho_planilha_portfolio", "").strip().strip('"')
    if cam_t and os.path.isfile(cam_t):
        p1 = cam_t; print(f"  [OK] {os.path.basename(p1)}")
    else:
        p1 = achar_arquivo(SCRIPT_DIR, "01_CONTROLE_TUTORIA.xlsx")
        if p1: print(f"  [OK] {os.path.basename(p1)}")
        else:  print(f"  [FALTA] 01_CONTROLE_TUTORIA.xlsx")
    if cam_p and os.path.isfile(cam_p):
        p2 = cam_p; print(f"  [OK] {os.path.basename(p2)}")
    else:
        # PATCH 102: nome do arquivo real é "PORTIFOLIO" (com I), não
        # "PORTFOLIO" — sem isso, o match exato nunca acontece e cai sempre
        # no fallback aproximado, que pode escolher o arquivo errado quando
        # existe mais de um "PORTIFOLIO_TUTOR*" na pasta.
        p2 = achar_arquivo(SCRIPT_DIR, "PORTIFOLIO_TUTOR.xlsx")
        if p2: print(f"  [OK] {os.path.basename(p2)}")
        else:  print(f"  [FALTA] PORTIFOLIO_TUTOR.xlsx")
    # PATCH 10: planilha nova de portfólios 2026/2 (formulário customizado, schema próprio)
    p2b = achar_arquivo(SCRIPT_DIR, "PORTIFOLIO_TUTOR_2026_2.xlsx")
    if p2b: print(f"  [OK] {os.path.basename(p2b)}")
    else:   print(f"  [INFO] PORTIFOLIO_TUTOR_2026_2.xlsx não encontrada (ainda sem envios 2026/2?)")
    # PATCH 6: prints duplicados de p1/p2 removidos aqui
    tmpl = os.path.join(SCRIPT_DIR, "template_dashboard.html")
    if os.path.isfile(tmpl): print(f"  [OK] template_dashboard.html")
    else:                    print(f"  [FALTA] template_dashboard.html")
    p3 = achar_arquivo(SCRIPT_DIR, "REL_GERAL_DE_GERENCIAMENTO.xlsx")
    if p3: print(f"  [OK] {os.path.basename(p3)}")
    else:  print(f"  [INFO] REL_GERAL_DE_GERENCIAMENTO.xlsx não encontrada (módulo desativado)")
    # PATCH 18: planilha de gerenciamento específica de 2026/2 (export novo, CSV)
    p3b = achar_arquivo(SCRIPT_DIR, "REL_GERAL_DE_GERENCIAMENTO_26_02.csv")
    if p3b: print(f"  [OK] {os.path.basename(p3b)}")
    else:   print(f"  [INFO] REL_GERAL_DE_GERENCIAMENTO_26_02.csv não encontrada")
    p4 = achar_arquivo(SCRIPT_DIR, "LOTACAO_TUTORES.xlsm") or achar_arquivo(SCRIPT_DIR, "LOTACAO_TUTORES.xlsx")
    if p4: print(f"  [OK] {os.path.basename(p4)}")
    else:  print(f"  [INFO] LOTACAO_TUTORES não encontrada (.xlsx/.xlsm)")
    # ── CSV de alunos por hub (igual aos outros arquivos: URL no Secret/env) ──
    p5 = None
    # 1. Tentar achar na pasta planilhas/ (já baixado anteriormente)
    p5 = achar_arquivo(SCRIPT_DIR, "Relatorio_alunos_por_hub.csv")
    if p5:
        print(f"  [OK] {os.path.basename(p5)}")
    else:
        # 2. Tentar baixar via variável de ambiente URL_ALUNOS_HUB (Secret GitHub)
        import re
        url_hub = os.environ.get("URL_ALUNOS_HUB", "").strip()
        if url_hub:
            print(f"  [Baixando] Relatorio_alunos_por_hub.csv via URL_ALUNOS_HUB...")
            try:
                import urllib.request
                # Converter link SharePoint/OneDrive para download direto
                # Tentar múltiplos formatos de URL
                def _build_dl_urls(url):
                    urls = []
                    if 'sharepoint.com' in url:
                        # Formato 1: adicionar &download=1
                        sep = '&' if '?' in url else '?'
                        urls.append(url + sep + 'download=1')
                        # Formato 2: download.aspx com token
                        m = re.search(r'/([A-Za-z0-9_-]{20,})[?]', url)
                        if m:
                            base = re.match(r'(https://[^/]+)', url).group(1)
                            user = re.search(r'/personal/([^/]+)/', url)
                            if user:
                                urls.append(f"{base}/personal/{user.group(1)}/_layouts/15/download.aspx?share={m.group(1)}")
                    elif '1drv.ms' in url:
                        sep = '&' if '?' in url else '?'
                        urls.append(url + sep + 'download=1')
                    urls.append(url)  # URL original como último recurso
                    return urls

                dest = os.path.join(pasta_planilhas, "Relatorio_alunos_por_hub.csv")
                downloaded = False
                for url_dl in _build_dl_urls(url_hub):
                    try:
                        req = urllib.request.Request(url_dl, headers={
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
                        with urllib.request.urlopen(req, timeout=120) as r:
                            data = r.read()
                        if len(data) > 10000 and b'<!DOCTYPE' not in data[:500]:
                            with open(dest, 'wb') as f_out: f_out.write(data)
                            p5 = dest
                            print(f"  [OK] Relatorio_alunos_por_hub.csv ({len(data):,} bytes)")
                            downloaded = True
                            break
                        else:
                            print(f"  [AVISO] URL retornou conteúdo inválido ({len(data)} bytes): {url_dl[:80]}")
                    except Exception as ex:
                        print(f"  [AVISO] Erro ao baixar: {ex} | URL: {url_dl[:80]}")
                if not downloaded:
                    print(f"  [ERRO] Não foi possível baixar o CSV de alunos — verifique URL_ALUNOS_HUB")
            except Exception as e:
                print(f"  [ERRO] Não foi possível baixar CSV de alunos: {e}")
        else:
            print(f"  [INFO] Relatorio_alunos_por_hub.csv não encontrado (defina URL_ALUNOS_HUB)")

    # ── PATCH 88: planilha de acompanhamento de onboarding (preenchida pelos
    # tutores/equipe) — mesmo padrão de secret+URL do Relatorio_alunos_por_hub,
    # só que aqui é um .xlsx, não .csv. Precisa da variável de ambiente
    # URL_ONBOARDING_TUTORES (secret novo no GitHub) apontando pro link de
    # download direto do OneDrive/SharePoint desse arquivo.
    p6 = achar_arquivo(SCRIPT_DIR, "Acompanhamento_Onboarding.xlsx")
    if p6:
        print(f"  [OK] {os.path.basename(p6)}")
    else:
        url_onb = os.environ.get("URL_ONBOARDING_TUTORES", "").strip()
        if url_onb:
            print(f"  [Baixando] Acompanhamento_Onboarding.xlsx via URL_ONBOARDING_TUTORES...")
            try:
                import urllib.request as _urlreq
                def _build_dl_urls_onb(url):
                    urls = []
                    if 'sharepoint.com' in url:
                        sep = '&' if '?' in url else '?'
                        urls.append(url + sep + 'download=1')
                        m = re.search(r'/([A-Za-z0-9_-]{20,})[?]', url)
                        if m:
                            base = re.match(r'(https://[^/]+)', url).group(1)
                            user = re.search(r'/personal/([^/]+)/', url)
                            if user:
                                urls.append(f"{base}/personal/{user.group(1)}/_layouts/15/download.aspx?share={m.group(1)}")
                    elif '1drv.ms' in url:
                        sep = '&' if '?' in url else '?'
                        urls.append(url + sep + 'download=1')
                    urls.append(url)
                    return urls
                dest_onb = os.path.join(pasta_planilhas, "Acompanhamento_Onboarding.xlsx")
                downloaded_onb = False
                for url_dl in _build_dl_urls_onb(url_onb):
                    try:
                        req = _urlreq.Request(url_dl, headers={
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
                        with _urlreq.urlopen(req, timeout=120) as r:
                            data = r.read()
                        if len(data) > 2000 and b'<!DOCTYPE' not in data[:500]:
                            with open(dest_onb, 'wb') as f_out: f_out.write(data)
                            p6 = dest_onb
                            print(f"  [OK] Acompanhamento_Onboarding.xlsx ({len(data):,} bytes)")
                            downloaded_onb = True
                            break
                    except Exception as ex:
                        print(f"  [AVISO] Erro ao baixar onboarding: {ex} | URL: {url_dl[:80]}")
                if not downloaded_onb:
                    print(f"  [ERRO] Não foi possível baixar Acompanhamento_Onboarding.xlsx — verifique URL_ONBOARDING_TUTORES")
            except Exception as e:
                print(f"  [ERRO] Não foi possível baixar onboarding: {e}")
        else:
            print(f"  [INFO] Acompanhamento_Onboarding.xlsx não encontrada (secret URL_ONBOARDING_TUTORES ainda não configurado)")
    return p1, p2, tmpl, p3, p3b, p4, p5, p6


def ler_excel(path, **kwargs):
    for engine in ('openpyxl', 'xlrd', None):
        try:
            kw = dict(kwargs)
            if engine: kw['engine'] = engine
            return pd.read_excel(path, **kw)
        except Exception: continue
    raise ValueError(f"Não foi possível ler {path} com nenhum engine disponível")


def _ler_arquivo_gerenciamento(path):
    # PATCH 18: REL_GERAL_DE_GERENCIAMENTO pode vir como .xlsx (export antigo) ou
    # .csv (export novo, ISO-8859-1, separador ';', tudo entre aspas)
    if str(path).lower().endswith('.csv'):
        last_err = None
        for enc in ('latin-1', 'utf-8', 'cp1252'):
            try:
                df = pd.read_csv(path, sep=';', encoding=enc, dtype=str)
                if len(df.columns) > 1: return df
            except Exception as e:
                last_err = e
        raise ValueError(f"Não foi possível ler CSV de gerenciamento ({path}): {last_err}")
    return ler_excel(path)


def processar(p1, p2):
    print(f"[{ts()}] Lendo tutores...")
    with open(p1, 'rb') as _f:
        _magic = _f.read(8); _preview = _f.read(200)
    print(f"  [DEBUG] Magic bytes de {p1}: {_magic.hex()}")
    if _magic[:2] == b'PK': print(f"  [DEBUG] Formato ZIP/XLSX confirmado")
    elif _magic[:2] in (b'\xd0\xcf', b'\xCF\xD0'): print(f"  [DEBUG] Formato XLS (OLE2) confirmado")
    else:
        print(f"  [ERRO] Arquivo não é Excel válido. Conteúdo inicial:")
        print((_magic + _preview).decode('utf-8', errors='replace')[:300])
        raise ValueError(f"Arquivo {p1} não é um Excel válido — SharePoint pode ter retornado HTML de erro")
    df_t = ler_excel(p1, sheet_name='Base de Tutores', header=1)
    col_sit  = next((c for c in df_t.columns if 'SITUA' in str(c).upper()), None)
    col_nome = next((c for c in df_t.columns if 'NOME'  in str(c).upper() and 'TUTOR' in str(c).upper()), None)
    # Busca flexível de colunas no CONTROLE_TUTORIA
    col_polo = next((c for c in df_t.columns if str(c).strip().upper() == 'POLO'), None) or                next((c for c in df_t.columns if 'POLO' in str(c).upper() and 'HUB' not in str(c).upper()), None) or                next((c for c in df_t.columns if 'POLO' in str(c).upper()), None) or 'POLO'
    col_cur  = next((c for c in df_t.columns if str(c).strip().upper() == 'CURSOS'), None) or                next((c for c in df_t.columns if 'CURSO' in str(c).upper() and 'VINC' not in str(c).upper()), None) or                next((c for c in df_t.columns if 'CURSO' in str(c).upper()), None) or 'CURSOS'
    col_email= next((c for c in df_t.columns if 'E-MAIL' in str(c).upper() or 'EMAIL' in str(c).upper()), None)
    print(f"[{ts()}] CONTROLE colunas detectadas: polo='{col_polo}' cursos='{col_cur}' email='{col_email}'")
    print(f"[{ts()}] CONTROLE todas colunas: {list(df_t.columns[:20])}")
    # PATCH 25a: validação obrigatória de colunas críticas — se a busca flexível
    # não achou a coluna de verdade, col_polo/col_cur caem no fallback literal
    # ('POLO'/'CURSOS'), que NÃO existe no DataFrame. Sem essa checagem, o script
    # seguia rodando silenciosamente com POLO/CURSOS vazios em todo mundo (dados
    # incompletos sem nenhum aviso). Agora para a execução com um erro claro,
    # listando as colunas disponíveis, assim que uma renomeação de coluna na
    # planilha-fonte quebra a detecção — em vez de gerar um dashboard manco.
    _colunas_criticas_controle = {
        'POLO': col_polo, 'CURSOS': col_cur, 'NOME DO TUTOR': col_nome, 'E-MAIL': col_email,
    }
    _faltando_controle = [nome for nome, col in _colunas_criticas_controle.items()
                          if col is None or col not in df_t.columns]
    if _faltando_controle:
        raise ValueError(
            f"[FALHA CRÍTICA] Coluna(s) obrigatória(s) não encontrada(s) no CONTROLE_TUTORIA: "
            f"{_faltando_controle}. Colunas disponíveis na planilha: {list(df_t.columns)}. "
            f"Provável renomeação de coluna na planilha-fonte — ajuste a busca flexível "
            f"acima ou corrija o cabeçalho na planilha antes de rodar novamente."
        )
    col_cat    = next((c for c in df_t.columns if 'CATEGORIA' in str(c).upper()), None)
    col_inicio = next((c for c in df_t.columns if str(c).upper().strip() in ('INÍCIO','INICIO')), None)
    col_whats  = next((c for c in df_t.columns if 'WHATSAPP' in str(c).upper()), None)
    col_chapa  = next((c for c in df_t.columns if 'CHAPA' in str(c).upper()), None)
    # PATCH 1: Detectar coluna CH SEMANAL na planilha de controle
    col_ch = next((c for c in df_t.columns if str(c).upper().strip() == 'CH SEMANAL' or
                   ('CH' in str(c).upper() and 'SEMAL' in str(c).upper())), None)
    if col_ch:
        print(f"[{ts()}] CH SEMANAL encontrada: '{col_ch}'")
        ch_vals = df_t[col_ch].dropna()
        print(f"[{ts()}] Amostra CH SEMANAL: {list(ch_vals.head(5))}")
    else:
        print(f"[{ts()}] CH SEMANAL não encontrada — colunas CH disponíveis: {[c for c in df_t.columns if 'CH' in str(c).upper()]}")
    # Filtrar tutores ativos: incluir Ativo + afastamentos temporários (Licença Maternidade, etc.)
    # Excluir apenas Inativo e Desligado explicitamente
    _SITUACOES_EXCLUIR = {'inativo', 'desligado', 'rescindido', 'demitido', 'encerrado',
                           'admissão prox.mês', 'admissao prox.mes', 'em admissão',
                           'pendente', 'aguardando'}
    if col_sit:
        _sit_norm = df_t[col_sit].astype(str).str.strip().str.lower()
        df_at = df_t[~_sit_norm.isin(_SITUACOES_EXCLUIR)].copy()
        _contagem = df_at[col_sit].value_counts().to_dict()
        print(f"[{ts()}] Situações incluídas: {_contagem}")
        # PATCH 105: captura os tutores desligados numa lista separada — o
        # filtro acima os exclui de TUDO no Vinci (correto, pra não confundir
        # com ativos), mas o Leo pediu um relatório específico de quem saiu,
        # com a data de desligamento, filtrável por mês.
        _situacoes_desligado = {'desligado', 'rescindido', 'demitido', 'encerrado'}
        df_desligados_raw = df_t[_sit_norm.isin(_situacoes_desligado)].copy()
        col_desligamento = next((c for c in df_t.columns if 'DESLIGAMENTO' in str(c).upper()), None)
        tutores_desligados = []
        for _, _tr in df_desligados_raw.iterrows():
            # PATCH 106: a coluna DESLIGAMENTO vem como data/hora de verdade
            # (pandas Timestamp), não texto — str() direto produzia algo tipo
            # "2026-02-11 00:00:00", que nunca batia com o formato DD/MM/AAAA
            # que o filtro de mês no Vinci espera. Formata explicitamente.
            _val_desl = _tr.get(col_desligamento, '') if col_desligamento else ''
            _data_desl = ''
            if _val_desl not in ('', None) and not (isinstance(_val_desl, float) and str(_val_desl) == 'nan'):
                try:
                    _data_desl = pd.to_datetime(_val_desl).strftime('%d/%m/%Y')
                except Exception:
                    _data_desl = str(_val_desl).strip()
            tutores_desligados.append({
                'n': str(_tr.get(col_nome, '') or ''),
                'p': str(_tr.get(col_polo, '') or ''),
                'c': str(_tr.get(col_cat, '') or '') if col_cat else '',
                'situacao': str(_tr.get(col_sit, '') or ''),
                'data_desligamento': _data_desl,
            })
        print(f"[{ts()}] Tutores desligados capturados (relatório separado): {len(tutores_desligados)}")
    else:
        df_at = df_t.copy()
        tutores_desligados = []
    # Usar CHAVE LOTAÇÃO do CONTROLE diretamente (mesma chave usada pelo Forms)
    col_chave_lot = next((c for c in df_t.columns if 'CHAVE' in str(c).upper() and 'LOTA' in str(c).upper()
                          and 'CLASSIF' not in str(c).upper()), None)
    if col_chave_lot:
        df_at['_CHAVE'] = df_at[col_chave_lot].astype(str).str.strip()
        print(f"[{ts()}] CONTROLE usando coluna '{col_chave_lot}' como chave")
    else:
        df_at['_CHAVE'] = df_at[col_polo].astype(str).str.strip() + df_at[col_cur].astype(str).str.strip()
        print(f"[{ts()}] CONTROLE chave construída de POLO + CURSOS")
    _sample_chaves = df_at['_CHAVE'].dropna().head(8).tolist()
    print(f"[{ts()}] CONTROLE chaves (amostra): {_sample_chaves}")
    print(f"[{ts()}] Lendo portfolios...")
    df_p = ler_excel(p2, sheet_name='Sheet1')
    df_p['_SEMESTRE_ORIGEM'] = '2026/1'  # PATCH 127: origem do arquivo, não a data — ver nota mais abaixo


    def col(df, *partes):
        for c in df.columns:
            cu = str(c).upper()
            if all(p.upper() in cu for p in partes): return c
        return None
    c_chave = col(df_p, 'CHAVE', 'LINK')
    c_polo_p = col(df_p, 'POLO')
    c_proto = col(df_p, 'PROTOCOLOS', 'ATIVIDADES')
    for sfx in (':7', ':8', ':6', ':9'):
        proto_sfx = [c for c in df_p.columns if 'PROTOCOLOS' in str(c).upper() and str(c).endswith(sfx)]
        if proto_sfx: c_proto = proto_sfx[0]; break
    else:
        proto_any = [c for c in df_p.columns if 'PROTOCOLOS' in str(c).upper() or 'ATIVIDADE' in str(c).upper()]
        if proto_any: c_proto = proto_any[0]
    data_cols = [c for c in df_p.columns if 'DATA DA APLICA' in str(c).upper() and str(c).endswith(':7')]
    if not data_cols: data_cols = [c for c in df_p.columns if 'DATA DA APLICA' in str(c).upper()]
    if not data_cols: data_cols = [c for c in df_p.columns if 'DATA' in str(c).upper() and 'APLICA' in str(c).upper()]
    c_data = data_cols[0] if data_cols else None
    def soma_estudantes(df):
        # PATCH 122: a lógica antiga ("acha UMA coluna, a de maior sufixo")
        # só funcionava por acaso pro PORTIFOLIO_TUTOR.xlsx de 2026/1 porque
        # tinha uma coluna extra adicionada à mão ("...Experimento?72") com
        # fórmula =SOMA(base, ?1, ?2, ?3, ?4, ?5, ?6) — ou seja, o "aluno_col"
        # nunca foi "a repetição mais completa", era literalmente a SOMA de
        # todas as repetições daquela linha (confirmado reproduzindo a fórmula
        # em Python: bate 100% com a coluna original). O de 2026/2 não tem
        # essa coluna auxiliar, então a busca antiga caía no fallback errado
        # e os alunos saíam todos vazios. Substituído por somar de verdade
        # todas as colunas de "Estudantes" (exceto a "?72" auxiliar, se
        # existir, pra não contar em dobro) — funciona igual nos dois
        # arquivos, sem depender de coluna extra adicionada à mão.
        cols = [c for c in df.columns
                if 'ESTUDANTES' in str(c).upper() and 'PONTOS' not in str(c).upper()
                and 'COMENT' not in str(c).upper() and not str(c).endswith('72')]
        if not cols: return None
        return df[cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
    aluno_soma = soma_estudantes(df_p)
    cat_cols = [c for c in df_p.columns if 'CATEGORIA' in str(c).upper() and 'PONTOS' not in str(c).upper() and 'COMENT' not in str(c).upper()]
    c_cat = cat_cols[0] if cat_cols else None
    print(f"[{ts()}] Colunas: chave={c_chave}, proto={c_proto}, data={c_data}, alunos=soma de {sum(1 for c in df_p.columns if 'ESTUDANTES' in str(c).upper() and 'PONTOS' not in str(c).upper() and 'COMENT' not in str(c).upper() and not str(c).endswith('72'))} colunas, cat={c_cat}")
    # PATCH 25a: mesma validação crítica, agora pro PORTIFOLIO_TUTOR — sem chave e
    # sem protocolo não há como casar nenhuma submissão a nenhum tutor; melhor
    # parar aqui com um erro explícito do que gerar um dashboard sem portfólios.
    _faltando_portfolio = [nome for nome, col in {'CHAVE/LINK': c_chave, 'PROTOCOLOS': c_proto}.items() if not col]
    if _faltando_portfolio:
        raise ValueError(
            f"[FALHA CRÍTICA] Coluna(s) obrigatória(s) não encontrada(s) no PORTIFOLIO_TUTOR: "
            f"{_faltando_portfolio}. Colunas disponíveis na planilha: {list(df_p.columns)}. "
            f"Provável renomeação de coluna no formulário/planilha-fonte."
        )
    c_ordem_cols = [c for c in df_p.columns if 'ORDEM' in str(c).upper() and 'PONTOS' not in str(c).upper() and 'COMENT' not in str(c).upper()]
    c_ordem = c_ordem_cols[0] if c_ordem_cols else None
    print(f"[{ts()}] Coluna ordem: {c_ordem}")
    df_p['_CHAVE']  = df_p[c_chave].astype(str).str.strip() if c_chave else ''
    df_p['_POLO']   = df_p[c_polo_p].astype(str).str.strip() if c_polo_p else ''
    df_p['_PROTO']  = df_p[c_proto].astype(str).str.strip() if c_proto else ''
    df_p['_DATA']   = pd.to_datetime(df_p[c_data], errors='coerce') if c_data else pd.NaT
    df_p['_ALUNOS'] = aluno_soma.fillna(0).astype(int) if aluno_soma is not None else 0
    df_p['_CAT']    = df_p[c_cat].astype(str).str.strip() if c_cat else ''
    df_p['_ORDEM']  = df_p[c_ordem].astype(str).str.strip() if c_ordem else 'Ordem 1'
    # ── MEC Cache ────────────────────────────────────────────────────────────
    mec_cache = {}
    mec_file = os.path.join(SCRIPT_DIR, 'mec_cache.json')
    if os.path.isfile(mec_file):
        with open(mec_file, encoding='utf-8') as f: mec_cache = json.load(f)
        print(f"[{ts()}] MEC cache: {len(mec_cache)} tutores")
    # ── Fim MEC Cache ─────────────────────────────────────────────────────────

    catalogo_oficial = {}
    id_to_perfil = {}  # PATCH 10: id da prática -> código de perfil (ex: '206'->'EMF-ISN')
    cat_file = os.path.join(SCRIPT_DIR, 'catalogo_oficial.json')
    if os.path.isfile(cat_file):
        with open(cat_file, encoding='utf-8') as f: raw = json.load(f)
        for cat_nome, praticas in raw.items():
            if isinstance(praticas, list) and praticas:
                if isinstance(praticas[0], dict):
                    catalogo_oficial[cat_nome] = sorted(set(p['nome'] for p in praticas))
                    for p in praticas:
                        if p.get('id') and p.get('perfil'): id_to_perfil.setdefault(str(p['id']), p['perfil'])
                else: catalogo_oficial[cat_nome] = sorted(set(praticas))
        print(f"[{ts()}] Catalogo oficial (JSON): {len(catalogo_oficial)} categorias")
    # PATCH 10: reforça id_to_perfil com mapa verificado direto do catálogo do formulário
    # (garante a correspondência mesmo se catalogo_oficial.json não tiver id/perfil)
    idp_file = os.path.join(SCRIPT_DIR, 'id_to_perfil.json')
    if os.path.isfile(idp_file):
        with open(idp_file, encoding='utf-8') as f: idp_extra = json.load(f)
        for k, v in idp_extra.items(): id_to_perfil.setdefault(k, v)
        print(f"[{ts()}] id_to_perfil: {len(id_to_perfil)} práticas mapeadas")

    # PATCH 15: nome da prática -> perfil correto (só pra códigos ambíguos BFI/BTO/COS-TIP)
    def _norm_proto(s):
        # normaliza unicode (resolve variantes de hífen/travessão), colapsa espaços
        s = unicodedata.normalize('NFKC', str(s or ''))
        s = s.replace('–', '-').replace('—', '-')  # en-dash/em-dash -> hífen comum
        return ' '.join(s.split()).strip()
    NOME_TO_PERFIL = {}
    nomep_file = os.path.join(SCRIPT_DIR, 'nome_to_perfil.json')
    if os.path.isfile(nomep_file):
        with open(nomep_file, encoding='utf-8') as f: _ntp_raw = json.load(f)
        NOME_TO_PERFIL = {_norm_proto(k): v for k, v in _ntp_raw.items()}
        print(f"[{ts()}] nome_to_perfil: {len(NOME_TO_PERFIL)} práticas (correção BFI/BTO/COS-TIP)")
    if not catalogo_oficial:
        cat_xlsx = achar_arquivo(SCRIPT_DIR, 'CATALOGO_EXPERIMENTOS.xlsx')
        if not cat_xlsx:
            for f in os.listdir(os.path.join(SCRIPT_DIR, 'planilhas')) if os.path.isdir(os.path.join(SCRIPT_DIR, 'planilhas')) else []:
                fu = f.upper()
                if ('RELAT' in fu and 'EXPER' in fu) or ('CATALOGO' in fu and 'EXPER' in fu):
                    cat_xlsx = os.path.join(SCRIPT_DIR, 'planilhas', f); break
        if cat_xlsx and os.path.isfile(cat_xlsx):
            try:
                df_cat = pd.read_excel(cat_xlsx)
                c_cat_nome = next((c for c in df_cat.columns if 'CATEGORIA' in str(c).upper()), None)
                c_exp_nome = next((c for c in df_cat.columns if 'EXPERIMENTO' in str(c).upper() or 'NOME' in str(c).upper()), None)
                c_sit = next((c for c in df_cat.columns if 'SITUA' in str(c).upper()), None)
                if c_cat_nome and c_exp_nome:
                    if c_sit: df_cat = df_cat[df_cat[c_sit].astype(str).str.strip().str.upper() == 'ATIVO']
                    for cat_val, grp in df_cat.groupby(c_cat_nome):
                        cat_str = str(cat_val).strip()
                        if cat_str and cat_str != 'nan':
                            nomes = sorted(set(str(n).strip() for n in grp[c_exp_nome].dropna() if str(n).strip() and str(n).strip() != 'nan'))
                            if nomes: catalogo_oficial[cat_str] = nomes
                    print(f"[{ts()}] Catalogo oficial (Excel): {len(catalogo_oficial)} categorias, {sum(len(v) for v in catalogo_oficial.values())} práticas")
            except Exception as e: print(f"[{ts()}] AVISO: Erro ao ler catálogo Excel: {e}")

    # PATCH 128: schema REAL confirmado pelo log do GitHub Actions (o arquivo
    # que roda de verdade via secret URL_PORTFOLIO_2026_2, NÃO o que foi
    # enviado manualmente pro chat de apoio, que era outro arquivo/formato):
    # colunas ID, DATA_ENVIO, EMAIL_TUTOR, NOME_TUTOR, POLO, CATEGORIA_LAB,
    # ORDEM_DISCIPLINA, PROTOCOLO_ID, PROTOCOLO_NOME, DISCIPLINA,
    # DATA_APLICACAO, QTD_ESTUDANTES, COMENTARIOS, FOTO_1/2/3 — um sistema de
    # envio novo e mais simples pro 2026/2 (1 linha por submissão, não o
    # formulário antigo com perguntas repetidas). Esse é exatamente o schema
    # que o PATCH 10 original (antes do PATCH 122) já esperava — o PATCH 122
    # "corrigiu" com base num arquivo de referência que não era o real, e
    # quebrou o merge de verdade (log mostrava "sem coluna de Polo/Categoria/
    # Protocolos — pulando merge", 0 registros). Revertido pro schema real,
    # mantendo os pontos bons que vieram depois (marcação de semestre pra
    # dedup, log mais claro).
    p2b = achar_arquivo(SCRIPT_DIR, "PORTIFOLIO_TUTOR_2026_2.xlsx")
    if p2b:
        try:
            df_novo = ler_excel(p2b, sheet_name='PORTIFOLIOS')
        except Exception:
            df_novo = ler_excel(p2b, sheet_name=0)
        df_novo.columns = [str(c).strip().upper() for c in df_novo.columns]
        if len(df_novo):
            def _g(coluna): return df_novo[coluna] if coluna in df_novo.columns else pd.Series([''] * len(df_novo))
            # PATCH 129: se a coluna PROTOCOLO_ID tiver QUALQUER célula vazia
            # (comum em planilha real), o pandas promove a coluna inteira pra
            # float64 — "55" vira "55.0" — e isso nunca bate com as chaves do
            # id_to_perfil.json (strings limpas tipo "55"). Corrigido
            # convertendo via float->int quando possível, só caindo pro texto
            # cru se não for numérico de jeito nenhum.
            def _limpar_id(v):
                s = str(v).strip()
                if s in ('', 'nan', 'None'): return ''
                try:
                    return str(int(float(s)))
                except (ValueError, TypeError):
                    return s
            _protoid = _g('PROTOCOLO_ID').map(_limpar_id)
            _perfil_mapeado = _protoid.map(id_to_perfil)
            _sem_perfil = int(_perfil_mapeado.isna().sum())
            if _sem_perfil:
                _ids_exemplo = sorted(set(_protoid[_perfil_mapeado.isna()].tolist()))[:8]
                print(f"[{ts()}] AVISO: {_sem_perfil} envios em PORTIFOLIO_TUTOR_2026_2 com PROTOCOLO_ID não mapeado em id_to_perfil.json. Exemplos de ID: {_ids_exemplo}")
            df_novo['_CHAVE']  = _g('POLO').astype(str).str.strip() + _perfil_mapeado.fillna('')
            df_novo['_POLO']   = _g('POLO').astype(str).str.strip()
            df_novo['_PROTO']  = _g('PROTOCOLO_NOME').astype(str).str.strip()
            df_novo['_DATA']   = pd.to_datetime(_g('DATA_APLICACAO'), errors='coerce')
            df_novo['_ALUNOS'] = pd.to_numeric(_g('QTD_ESTUDANTES'), errors='coerce').fillna(0).astype(int)
            df_novo['_CAT']    = _g('CATEGORIA_LAB').astype(str).str.strip()
            df_novo['_ORDEM']  = _g('ORDEM_DISCIPLINA').astype(str).str.strip().replace('', 'Ordem 1')
            df_novo['EMAIL']      = _g('EMAIL_TUTOR').astype(str).str.strip()
            df_novo['NOME_TUTOR'] = _g('NOME_TUTOR').astype(str).str.strip()
            df_novo['_SEMESTRE_ORIGEM'] = '2026/2'  # PATCH 127
            print(f"[{ts()}] PORTIFOLIO_TUTOR_2026_2 colunas: polo={'POLO' in df_novo.columns}, categoria={'CATEGORIA_LAB' in df_novo.columns}, protocolo_id={'PROTOCOLO_ID' in df_novo.columns}, qtd_estudantes={'QTD_ESTUDANTES' in df_novo.columns}")
            df_p = pd.concat([df_p, df_novo], ignore_index=True)
            print(f"[{ts()}] PORTIFOLIO_TUTOR_2026_2: {len(df_novo)} envios mesclados (2026/2), {len(df_novo)-_sem_perfil} com chave completa (perfil mapeado)")
        else:
            print(f"[{ts()}] PORTIFOLIO_TUTOR_2026_2: 0 envios ainda")
    else:
        print(f"[{ts()}] PORTIFOLIO_TUTOR_2026_2.xlsx não encontrada — só 2026/1 nesta rodada")

    # PATCH 123 (revisado): "Alunos registrados no Portfólio" — pedido pelo
    # Leo pra complementar o "Alunos Agendados" (GIOCONDA) na Visão Geral.
    # MESMO PROBLEMA que já foi corrigido do lado do GIOCONDA (PATCH 38): o
    # número de alunos presentes se repete a cada prática enviada da mesma
    # turma (polo+categoria) — somar direto infla o total. Deduplicado do
    # mesmo jeito (groupby polo+categoria, pega o MAIOR valor, não soma).
    #
    # BUG REAL encontrado depois de publicado (Leo reportou "647% Agendado
    # sobre Registrado", matematicamente impossível): esse número saía SEMPRE
    # somando TODO o histórico de portfólio (2026/1 + 2026/2 juntos, sem
    # filtro), enquanto "Alunos Agendados" (GIOCONDA) é por semestre
    # selecionado — comparação de coisas de escopo diferente. Corrigido
    # calculando por semestre de verdade (mesmo classificador de data que o
    # resto do pipeline usa, _data_para_semestre), guardando os 3 recortes
    # (2026/1, 2026/2, Ambos) — o front escolhe o certo conforme a aba de
    # semestre selecionada, do mesmo jeito que já faz pros outros KPIs.
    def _calc_portfolio_dedup(df_sub):
        _geral, _por_polo, _por_cat = 0, {}, {}
        if '_ALUNOS' in df_sub.columns and '_POLO' in df_sub.columns and '_CAT' in df_sub.columns:
            _base = df_sub[(df_sub['_POLO'].astype(str).str.len() > 0) & (df_sub['_CAT'].astype(str).str.len() > 0)]
            if len(_base):
                _dedup = _base.groupby(['_POLO', '_CAT'])['_ALUNOS'].max()
                _geral = int(_dedup.sum())
                for polo, s in _dedup.groupby(level=0).sum().items(): _por_polo[str(polo)] = int(s)
                for cat, s in _dedup.groupby(level=1).sum().items(): _por_cat[str(cat)] = int(s)
        return {'geral': _geral, 'por_polo': _por_polo, 'por_categoria': _por_cat}

    # PATCH 127: classificação por semestre usa a ORIGEM DO ARQUIVO
    # (_SEMESTRE_ORIGEM, marcado na hora de ler cada planilha), não a data de
    # aplicação. Tentei usar _data_para_semestre() primeiro (mesmo
    # classificador do resto do pipeline) mas ele só reconhece datas dentro
    # das janelas estreitas de cada Ordem (~4 semanas) — 93% das datas de
    # submissão de portfólio caem FORA dessas janelas (o tutor preenche
    # atrasado, ou fora do período oficial), então quase tudo virava "sem
    # semestre". Como cada semestre já vem de um arquivo Excel separado, usar
    # a origem do arquivo é mais simples e muito mais confiável.
    _sems_disponiveis = sorted(set(s for s in df_p['_SEMESTRE_ORIGEM'].dropna().unique()) | {SEMESTRE_ATUAL})
    portfolio_alunos_dedup_por_semestre = {}
    for _sem in _sems_disponiveis:
        portfolio_alunos_dedup_por_semestre[_sem] = _calc_portfolio_dedup(df_p[df_p['_SEMESTRE_ORIGEM'] == _sem])
    portfolio_alunos_dedup_por_semestre['Ambos'] = _calc_portfolio_dedup(df_p)
    # compat: 'portfolio_alunos_dedup' no nível raiz = semestre ativo (mesmo
    # padrão do dados.update(ger_dados) mais abaixo pro gerenciamento)
    _port_dedup_atual = portfolio_alunos_dedup_por_semestre.get(SEMESTRE_ATUAL, {'geral': 0, 'por_polo': {}, 'por_categoria': {}})
    _port_alunos_geral = _port_dedup_atual['geral']
    _port_alunos_por_polo = _port_dedup_atual['por_polo']
    _port_alunos_por_cat = _port_dedup_atual['por_categoria']
    for _sem, _d in portfolio_alunos_dedup_por_semestre.items():
        print(f"[{ts()}] Alunos registrados no Portfólio DEDUPLICADOS ({_sem}): {_d['geral']:,}")
    # (guardado em variável local — entra no dict final via return limpar({...})
    # lá embaixo, não existe "dados" incremental dentro desta função)

    chave_to_cat_raw = {}; chave_to_cf = {}; chave_alias = {}
    polo_biofar_cursos = {}
    for _, t in df_at.iterrows():
        polo_   = str(t.get(col_polo, '') or '').strip()
        cursos_ = str(t.get(col_cur,  '') or '').strip()
        cat_    = _normaliza_categoria_bio_duplicado(str(t.get(col_cat,  '') or '').strip()) if col_cat else ''
        if cursos_ in ('BBI', 'BFR') and 'BIO-FAR' in cat_.upper():
            if polo_ not in polo_biofar_cursos: polo_biofar_cursos[polo_] = set()
            polo_biofar_cursos[polo_].add(cursos_)
    # Pré-calcular polos que têm tutor BFI legítimo (para não criar alias conflitante)
    _polos_com_bfi = set()
    for _, t in df_at.iterrows():
        _cur_t = str(t.get(col_cur, '') or '').strip()
        _pol_t = str(t.get(col_polo, '') or '').strip()
        if _cur_t == 'BFI':
            _polos_com_bfi.add(_pol_t)
    for _, t in df_at.iterrows():
        polo = str(t.get(col_polo, '') or '').strip()
        cursos = str(t.get(col_cur, '') or '').strip()
        cat_raw = _normaliza_categoria_bio_duplicado(str(t.get(col_cat, '') or '').strip()) if col_cat else ''
        cf = CAT_MAP.get(cat_raw, cat_raw)
        chave = polo + cursos
        if chave and cat_raw:
            chave_to_cat_raw[chave] = cat_raw
            chave_to_cf[chave] = cf
            if cursos in ('BBI', 'BFR'):
                outros = polo_biofar_cursos.get(polo, set()) - {cursos}
                # Só adicionar alias BFI se NÃO existe tutor BFI legítimo neste polo
                # (evita colisão que faz portfólios BFI serem atribuídos a BIO-FAR)
                variantes = [] if polo in _polos_com_bfi else [polo + 'BFI']
                for outro in outros:
                    variantes += [polo+cursos+'-'+outro, polo+outro+'-'+cursos, polo+cursos+outro, polo+outro+cursos]
                for v in variantes:
                    chave_to_cf.setdefault(v, cf)
                    chave_alias.setdefault(v, chave)


    # PATCH 139: a checagem "essa prática já é conhecida oficialmente sob OUTRA
    # categoria, não deixa vazar" só funcionava com correspondência EXATA de
    # texto — o Leo reportou práticas de Engenharia (ENGMAKER) aparecendo sob
    # Enfermagem. Causa: o protocolo enviado vem com um prefixo de categoria no
    # nome ("ENGMAKER - Prática de hidrossanitária"), enquanto o catálogo
    # oficial guarda só "Prática de hidrossanitária" (sem prefixo) — a
    # comparação exata nunca batia, `cat_oficial` saía None, e a prática
    # escapava da proteção, podendo ser adicionada em QUALQUER categoria que a
    # chave da submissão apontasse (mesmo errada). Normalizado (minúsculas,
    # remove prefixo tipo "PALAVRA - " ou "PALAVRA/PALAVRA - " no início,
    # espaços colapsados) pra reconhecer a mesma prática independente de como
    # o nome foi digitado/formatado na submissão real.
    import re as _re_pratica
    def _norm_nome_pratica(s):
        s = str(s or '').strip().lower()
        m = _re_pratica.match(r'^(.{2,30}?)\s-\s(.+)$', s)
        if m and len(m.group(1).split()) <= 3:  # só remove prefixo curto (tipo "ENGMAKER -" ou "BIO-FAR -"), não corta nomes reais com " - " no meio
            s = m.group(2)
        s = _re_pratica.sub(r'\s+', ' ', s).strip()
        return s
    oficial_p_to_cat = {}
    oficial_p_to_cat_norm = {}
    for cat, pracs in catalogo_oficial.items():
        for p in pracs:
            oficial_p_to_cat.setdefault(p, cat)
            oficial_p_to_cat_norm.setdefault(_norm_nome_pratica(p), cat)
    catalogo_real = defaultdict(set)
    _bloqueados_categoria_errada = 0
    for _, r in df_p.iterrows():
        chave = str(r.get('_CHAVE', '') or '').strip()
        chave = chave_alias.get(chave, chave)
        proto = r['_PROTO']
        if not chave or chave == 'nan' or not proto or proto == 'nan': continue
        cf = chave_to_cf.get(chave, '')
        if not cf: continue
        for p in proto.split(';'):
            p = p.strip()
            if not p: continue
            cat_oficial = oficial_p_to_cat.get(p) or oficial_p_to_cat_norm.get(_norm_nome_pratica(p))
            if cat_oficial and cat_oficial != cf:
                _bloqueados_categoria_errada += 1
                continue
            catalogo_real[cf].add(p)
    if _bloqueados_categoria_errada:
        print(f"[{ts()}] Práticas bloqueadas de vazar pra categoria errada (nome já conhecido oficialmente em outra): {_bloqueados_categoria_errada}")
    catalogo = {}
    all_cats = set(list(catalogo_oficial.keys()) + list(catalogo_real.keys()))
    for cat in all_cats:
        base = set(catalogo_oficial.get(cat, [])); real = catalogo_real.get(cat, set())
        catalogo[cat] = sorted(base | real)
    print(f"[{ts()}] Catalogo final: {len(catalogo)} cats, {sum(len(v) for v in catalogo.values())} praticas")
    email_to_cf = {}; email_to_chave_tutor = {}
    col_email_t = next((c for c in df_t.columns if 'E-MAIL' in str(c).upper() or 'EMAIL' in str(c).upper()), None)
    _email_chaves_vistas = defaultdict(set)  # PATCH 25b: detectar e-mail duplicado no CONTROLE
    if col_email_t:
        for _, t in df_at.iterrows():
            em = str(t.get(col_email_t, '') or '').strip().lower()
            chave_t = t['_CHAVE']
            cat_raw_ = _normaliza_categoria_bio_duplicado(str(t.get(col_cat, '') or '').strip()) if col_cat else ''
            cf_ = CAT_MAP.get(cat_raw_, cat_raw_)
            if em and em != 'nan':
                _email_chaves_vistas[em].add(chave_t)
                email_to_cf[em] = cf_; email_to_chave_tutor[em] = chave_t
    # PATCH 25b: se o mesmo e-mail aparece em mais de uma linha ATIVA do CONTROLE
    # com chaves (polo+curso) diferentes, a última linha processada sobrescreve
    # silenciosamente as anteriores em email_to_chave_tutor — isso faz submissões
    # de fallback-por-email irem parar no destino errado, sem nenhum aviso. Foi
    # exatamente esse padrão suspeito no caso da tutora Cleya Da Silva Santana
    # Cruz (Diamantina/MG · EMF-ISN). Agora isso gera um aviso alto no log.
    _emails_duplicados = {em: chaves for em, chaves in _email_chaves_vistas.items() if len(chaves) > 1}
    if _emails_duplicados:
        print(f"[{ts()}] ⚠️  AVISO CRÍTICO: {len(_emails_duplicados)} e-mail(s) duplicado(s) no CONTROLE com chaves diferentes — fallback por e-mail pode estar direcionando submissões pro destino errado:")
        for _em, _chaves in _emails_duplicados.items():
            print(f"[{ts()}]     {_em} -> {sorted(_chaves)} (usando apenas a última chave processada: {email_to_chave_tutor.get(_em)!r})")
    col_email_p = next((c for c in df_p.columns if c.upper() in ('EMAIL', 'E-MAIL')), None)
    # ── Mapeamentos de fallback adicionais ──────────────────────────────────
    # Fallback 3: por nome do tutor (coluna "Nome do tutor" no Forms)
    col_nome_tutor_p = None
    for c in df_p.columns:
        cu = str(c).upper()
        if 'NOME' in cu and 'TUTOR' in cu: col_nome_tutor_p = c; break
    if not col_nome_tutor_p:
        for c in df_p.columns:
            cu = str(c).upper()
            if 'TUTOR' in cu and 'PONTOS' not in cu and 'COMENT' not in cu:
                col_nome_tutor_p = c; break

    def _norm_nome_match(s):
        import unicodedata
        s = str(s or '').strip().lower()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return s

    # PATCH 83: mesma lógica de subsequência já usada noutros pontos do arquivo
    # (nome com uma parte a mais ou a menos, tipo "Jose Cicero..." vs "Cicero...")
    # — aqui serve pra resolver submissões do Forms de portfólio que caem em
    # "Aviso de Portfólio" só porque o nome digitado no formulário tem uma
    # grafia levemente diferente do nome oficial no CONTROLE, mesmo a pessoa
    # sendo um tutor real e corretamente cadastrado.
    def _eh_subsequencia_nome_match(curtos, longos):
        i = 0
        for tok in longos:
            if i < len(curtos) and tok == curtos[i]:
                i += 1
        return i == len(curtos)

    def _nomes_batem_match(nome_a, nome_b):
        if nome_a == nome_b:
            return True
        ta, tb = nome_a.split(), nome_b.split()
        if not ta or not tb:
            return False
        if len(ta) >= 2 and len(tb) >= 2 and ta[0] == tb[0] and ta[-1] == tb[-1]:
            return True
        curtos, longos = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
        if len(curtos) < 2:
            return False
        return _eh_subsequencia_nome_match(curtos, longos)

    # Mapear: nome_normalizado → chave CONTROLE
    nome_to_chave_tutor = {}
    for _, t in df_at.iterrows():
        nome_t = str(t.get(col_nome, '') or '').strip()
        chave_t = t['_CHAVE']
        if nome_t and chave_t:
            nome_to_chave_tutor[_norm_nome_match(nome_t)] = chave_t
            # Também mapear primeiro+último nome
            parts = _norm_nome_match(nome_t).split()
            if len(parts) >= 2:
                fl = parts[0] + ' ' + parts[-1]
                nome_to_chave_tutor.setdefault(fl, chave_t)

    sem_match = 0; com_match = 0; match_por_email = 0; match_por_nome = 0

    # ── Constantes de tratamento de portfólios ───────────────────────────────
    # Emails a ignorar completamente (gestão — não são tutores de prática)
    EMAILS_IGNORAR = {'elienai.cesar@uniasselvi.com.br'}

    # Domínio válido para tutores de prática
    DOMINIO_TUTORIA = 'tutorpratica.uniasselvi.com.br'

    # Avisos de portfólio (email incorreto, preenchimento incorreto, etc.)
    _avisos_raw = {}  # key → {email, nome, polo, tipo, msg, count}

    def _dividir_codigo_composto(chave):
        """Divide chave com código composto (BFR-BBI) em sub-chaves individuais."""
        import re as _re2
        m = _re2.search(r'([A-Z]{2,4}-[A-Z]{2,4})$', chave)
        if m:
            polo_base = chave[:-len(m.group(1))]
            partes = m.group(1).split('-')
            candidatos = [polo_base + p for p in partes]
            encontrados = [c for c in candidatos if c in chave_to_cf]
            if encontrados:
                return encontrados
        return []

    def _encontrar_por_polo(chave, cat_subm=''):
        """Opção B: vincula submissão ao tutor ativo no mesmo polo.
        Remove sufixos progressivos da chave para encontrar o polo base."""
        cat_up = str(cat_subm).upper()
        # Palavras-chave por categoria para priorizar match
        _cat_kw = {
            'BIO': ['BFR','BBI','BIO'],
            'FAR': ['BFR','BBI','BIO'],
            'ENF': ['ENF','ISN','INS'],
            'NTR': ['NTR','NUTRI'],
            'ENG': ['EMF','ENG','MEC'],
            'FIS': ['BFI','FIS','FISIO'],
        }
        # Determinar categoria da submissão para priorizar candidatos
        _pref_codigos = []
        for kw, codigos in _cat_kw.items():
            if kw in cat_up:
                _pref_codigos = codigos; break

        for code_len in range(3, min(len(chave), 12)):
            polo_prefix = chave[:-code_len]
            if len(polo_prefix) < 6: break
            candidatos = [k for k in chave_to_cf if k.startswith(polo_prefix)]
            if candidatos:
                # Priorizar por categoria
                if _pref_codigos:
                    pref = [c for c in candidatos if any(cod in c for cod in _pref_codigos)]
                    if pref: return pref[0]
                return candidatos[0]
        return None

    enviados = defaultdict(list)
    polo_sem_tutor = defaultdict(list)  # polo+cat → práticas de tutores desligados
    _correcoes_perfil = 0
    _correcoes_perfil_falhou = set()
    for _, r in df_p.iterrows():
        chave = r['_CHAVE']; proto = r['_PROTO']
        if not chave or chave == 'nan' or not proto or proto == 'nan': continue

        # Ignorar emails de gestão
        _email_subm = str(r.get(col_email_p, '') or '').strip().lower() if col_email_p else ''
        if _email_subm in EMAILS_IGNORAR:
            continue

        chave = chave_alias.get(chave, chave)

        # PATCH 23: o email do remetente identifica a tutora de forma inequívoca.
        # Quando a chave submetida aponta pra outro tutor (casos EMF-ISN2, EMF-ISND,
        # etc. — o Forms só tinha a opção genérica, então a chave veio "errada"),
        # corrige pra chave real do tutor que enviou ANTES de qualquer outra checagem.
        if _email_subm and _email_subm in email_to_chave_tutor:
            _chave_by_email = email_to_chave_tutor[_email_subm]
            if _chave_by_email and _chave_by_email != chave and _chave_by_email in chave_to_cf:
                chave = _chave_by_email
        # que não distingue Fisio (BFI) / T.O. (BTO) / Estética (COS-TIP) corretamente
        # quando os 3 cursos compartilham o mesmo polo/laboratório — corrige usando o
        # nome da prática (inequívoco), que é mais confiável que a chave pronta.
        if NOME_TO_PERFIL and _norm_proto(proto) in NOME_TO_PERFIL:
            perfil_certo = NOME_TO_PERFIL[_norm_proto(proto)]
            for _cod_errado in ('BFI', 'BTO', 'COS-TIP', 'TIP-COS'):
                if chave.endswith(_cod_errado) and _cod_errado != perfil_certo:
                    chave_corrigida = chave[:-len(_cod_errado)] + perfil_certo
                    # Aplica sempre, mesmo sem tutor ativo no destino — é melhor cair
                    # no fluxo de "sem match" (vira anônimo no polo) do que ficar
                    # silenciosamente contado pro curso errado.
                    if chave_corrigida in chave_to_cf:
                        _correcoes_perfil += 1
                    else:
                        _correcoes_perfil_falhou.add((chave, chave_corrigida, proto))
                    chave = chave_corrigida
                    break

        if chave not in chave_to_cf:
            # Fallback 1: por email
            if col_email_p:
                em_p = str(r.get(col_email_p, '') or '').strip().lower()
                if em_p in email_to_chave_tutor:
                    chave = email_to_chave_tutor[em_p]
                    match_por_email += 1
                    # Sinalizar envio por e-mail de Regente de Polo (informativo)
                    if '@regentedepolo.' in em_p:
                        _nome_rp = str(r.get(col_nome_tutor_p, '') or '') if col_nome_tutor_p else ''
                        _aviso_key_rp = f"{em_p}||regente_de_polo"
                        if _aviso_key_rp not in _avisos_raw:
                            _avisos_raw[_aviso_key_rp] = {
                                'email': em_p, 'nome': _nome_rp, 'chave': chave,
                                'tipo': 'regente_de_polo',
                                'msg': 'Envio realizado pelo e-mail de Regente de Polo — tutor também é regente',
                                'count': 0
                            }
                        _avisos_raw[_aviso_key_rp]['count'] += 1

        if chave not in chave_to_cf:
            # Fallback 2: por nome do tutor na coluna específica
            if col_nome_tutor_p:
                nome_p = _norm_nome_match(str(r.get(col_nome_tutor_p, '') or ''))
                if nome_p in nome_to_chave_tutor:
                    chave = nome_to_chave_tutor[nome_p]
                    match_por_nome += 1
                else:
                    # Tentar primeiro+último nome
                    parts_p = nome_p.split()
                    if len(parts_p) >= 2:
                        fl_p = parts_p[0] + ' ' + parts_p[-1]
                        if fl_p in nome_to_chave_tutor:
                            chave = nome_to_chave_tutor[fl_p]
                            match_por_nome += 1
                    # PATCH 83: se nem nome exato nem primeiro+último bateram,
                    # tenta por subsequência (nome com parte a mais/a menos) —
                    # antes disso, esses tutores caíam direto em "Aviso de
                    # Portfólio" mesmo estando corretamente cadastrados, só
                    # porque o nome digitado no Forms tinha uma grafia
                    # levemente diferente da oficial no CONTROLE.
                    if chave not in chave_to_cf:
                        for _nome_ctrl, _chave_ctrl in nome_to_chave_tutor.items():
                            if _nomes_batem_match(nome_p, _nome_ctrl):
                                chave = _chave_ctrl
                                match_por_nome += 1
                                break

        if chave not in chave_to_cf:
            # Fallback 3: normalizar a própria chave (diferenças de espaços/acentos)
            chave_norm = _norm_nome_match(chave.replace(' ', ''))
            for k in chave_to_cf:
                if _norm_nome_match(k.replace(' ', '')) == chave_norm:
                    chave = k; match_por_nome += 1; break

        # Fallback 4: código composto (ex: BFR-BBI → BFR e BBI)
        if chave not in chave_to_cf:
            _sub_chaves = _dividir_codigo_composto(chave)
            if _sub_chaves:
                chave = _sub_chaves[0]  # usar primeira sub-chave
                match_por_nome += 1
                # Se houver mais de uma sub-chave, adicionar às demais depois
                for _sc in _sub_chaves[1:]:
                    if _sc in chave_to_cf:
                        _extra_proto = str(r.get('_PROTO', '') or '').strip()
                        _extra_data  = r['_DATA']
                        _extra_aluno = int(r['_ALUNOS'])
                        _extra_ordem = str(r.get('_ORDEM', 'Ordem 1') or 'Ordem 1').strip()
                        for _p in _extra_proto.split(';'):
                            _p = _p.strip()
                            if _p: enviados[_sc].append({'p':_p,'d':str(_extra_data)[:10] if pd.notna(_extra_data) else None,'a':_extra_aluno,'o':_extra_ordem})

        if chave in chave_to_cf:
            com_match += 1
        else:
            sem_match += 1
            # Criar aviso de portfólio
            _nome_subm  = str(r.get(col_nome_tutor_p, '') or '-') if col_nome_tutor_p else '-'
            _polo_subm  = chave  # chave contém polo+código
            if '@regentedepolo.' in _email_subm:
                _tipo = 'regente_de_polo'
                _msg  = 'Envio por e-mail de Regente de Polo sem correspondência no CONTROLE_TUTORIA'
            elif _email_subm and not _email_subm.endswith('@' + DOMINIO_TUTORIA):
                _dom = _email_subm.split('@')[1] if '@' in _email_subm else 'desconhecido'
                _tipo = 'email_incorreto'
                _msg  = f'E-mail incorreto: @{_dom} ao invés de @{DOMINIO_TUTORIA}'
            else:
                _tipo = 'chave_invalida'
                _msg  = f'Polo/categoria não encontrado no CONTROLE: {chave}'
            _aviso_key = f"{_email_subm}||{_tipo}"
            if _aviso_key not in _avisos_raw:
                _avisos_raw[_aviso_key] = {'email':_email_subm,'nome':_nome_subm,'chave':chave,'tipo':_tipo,'msg':_msg,'count':0}
            _avisos_raw[_aviso_key]['count'] += 1
            # Guardar práticas em polo_sem_tutor para entry anônimo no polo
            _polo_part = chave  # chave = polo+código sem separador
            for _pfb in str(r.get('_PROTO','') or '').split(';'):
                _pfb = _pfb.strip()
                if _pfb:
                    _ordem_pfb = str(r.get('_ORDEM','Ordem 1') or 'Ordem 1').strip()
                    polo_sem_tutor[_polo_part].append({'p':_pfb[:80],'d':None,'a':0,'o':_ordem_pfb})

        data = r['_DATA']; aluno = int(r['_ALUNOS'])
        for p in proto.split(';'):
            p = p.strip()
            if p:
                ordem_val = str(r.get('_ORDEM', 'Ordem 1') or 'Ordem 1').strip()
                if not any(o in ordem_val for o in ['Ordem 1','Ordem 2','Ordem 3','Ordem 4','Ordem 5']):
                    ordem_val = 'Ordem 1'
                _data_str = str(data)[:10] if pd.notna(data) else None
                _sem_envio = _data_para_semestre(_data_str)
                if _sem_envio is None:
                    # Data fora de qualquer janela → semestre mais antigo
                    _sem_envio = sorted(ALL_SEMESTRES.keys())[0]
                    # Log apenas se data estranha (debug)
                    if _data_str and _data_str < '2026-02-01':
                        pass  # silencioso — serão agrupados no 2026/1
                enviados[chave].append({'p': p, 'd': _data_str, 'a': aluno, 'o': ordem_val, 's': _sem_envio})
    # Finalizar avisos
    avisos_portfolio = sorted(_avisos_raw.values(), key=lambda x: -x['count'])
    print(f"[{ts()}] Matching submissões: {com_match} com chave, {match_por_email} por email, {match_por_nome} por nome/código, {sem_match} sem match")
    if _correcoes_perfil:
        print(f"[{ts()}] Correções BFI/BTO/COS-TIP por nome de prática: {_correcoes_perfil}")
    if _correcoes_perfil_falhou:
        print(f"[{ts()}] AVISO: {len(_correcoes_perfil_falhou)} correções de perfil sem tutor ativo no destino (foram para anônimo/polo):")
        for _ch, _chc, _pr in list(_correcoes_perfil_falhou)[:10]:
            print(f"    {_ch!r} -> {_chc!r} (não encontrado) | prática: {_pr!r}")
    if sem_match > 0:
        print(f"[{ts()}] Avisos de portfólio gerados: {len(avisos_portfolio)}")
        for av in avisos_portfolio:
            print(f"[{ts()}]   {av['tipo'].upper()} ({av['count']}x): {av['email']} | {av['nome']} | {av['msg']}")
    tutores = []
    # Pré-computar: para cada polo+CURSO ESPECÍFICO, agregar TODOS os envios
    # (resolve o caso de múltiplos tutores por polo que compartilham o mesmo curso)
    # PATCH 17: agrupar por curso específico (col_cur), não pela categoria ampla —
    # BFI/BTO/COS-TIP têm a mesma categoria ampla (Multidisciplinar III) mas são
    # cursos diferentes; agrupar pela categoria ampla juntava os 3 indevidamente.
    _polo_cat_enviados = {}  # (polo_str, curso_especifico) → lista merged de hist
    for _, t in df_at.iterrows():
        _ch = t['_CHAVE']
        _cursos_t = str(t.get(col_cur, '') or '').strip()
        _polo_str = str(t.get(col_polo, '') or '').strip()
        _key_pc = (_polo_str, _cursos_t)
        if _key_pc not in _polo_cat_enviados:
            # Buscar enviados pela chave canônica E por variantes de chave do mesmo
            # polo que tenham o MESMO curso específico (não só a mesma categoria ampla)
            _hist_merged = list(enviados.get(_ch, []))
            for _k, _h in enviados.items():
                if _k == _ch: continue
                if _k.startswith(_polo_str) and _k[len(_polo_str):] == _cursos_t:
                    for _item in _h:
                        if _item not in _hist_merged:
                            _hist_merged.append(_item)
            _polo_cat_enviados[_key_pc] = _hist_merged
    print(f"[{ts()}] Polo×curso com envios: {len([v for v in _polo_cat_enviados.values() if v])}")

    # PATCH 24: nome_to_perfil também é usado para SEPARAR o catálogo de "previstas"
    # (tp) por curso específico dentro de categorias compartilhadas (Multi III:
    # BFI/BTO/COS-TIP; Multi I: BBI/BFR). Sem isso, um tutor de BFI é cobrado pelas
    # práticas de BTO e COS-TIP também (e vice-versa), porque `catalogo` é indexado
    # só pela categoria AMPLA (cat_form), que soma as práticas dos 3 cursos juntos.
    def _catalogo_por_curso(praticas_full, cursos_t):
        if not NOME_TO_PERFIL or not cursos_t:
            return praticas_full
        filtradas = [p for p in praticas_full if NOME_TO_PERFIL.get(_norm_proto(p)) == cursos_t]
        # Só filtra se o mapa cobrir pelo menos 1 prática desse curso específico —
        # caso contrário (categoria não coberta pelo nome_to_perfil.json), mantém
        # o comportamento antigo em vez de zerar o catálogo do tutor.
        return filtradas if filtradas else praticas_full

    _hist_pre_admissao = 0
    for _, t in df_at.iterrows():
        chave    = t['_CHAVE']
        cat_raw  = _normaliza_categoria_bio_duplicado(str(t.get(col_cat, '') or '').strip()) if col_cat else ''
        cat_form = CAT_MAP.get(cat_raw, cat_raw)
        polo_str = str(t.get(col_polo, '') or '').strip()
        cursos_t = str(t.get(col_cur, '') or '').strip()
        praticas_full = catalogo.get(cat_form, catalogo.get(cat_raw, []))
        praticas = _catalogo_por_curso(praticas_full, cursos_t)  # PATCH 24
        # Enriquecimento MEC / data de admissão (calculado antes do filtro de hist)
        _email_t = str(t.get(col_email, '') or '').strip().lower() if col_email else ''
        _mec = mec_cache.get(_email_t, {})
        _inicio_ctrl = t.get(col_inicio) if col_inicio else None
        _inicio_str = None
        if _inicio_ctrl and str(_inicio_ctrl) not in ('nan','NaT','None',''):
            try:
                _inicio_str = _inicio_ctrl.strftime('%Y-%m-%d') if hasattr(_inicio_ctrl,'strftime') else str(_inicio_ctrl)[:10]
            except: pass
        if not _inicio_str: _inicio_str = _mec.get('admissao')

        # PATCH 17: agregado por polo+CURSO ESPECÍFICO (não categoria ampla) — cobre
        # múltiplos tutores do mesmo curso no mesmo polo, sem juntar BFI/BTO/COS-TIP
        hist_bruto = _polo_cat_enviados.get((polo_str, cursos_t), enviados.get(chave, []))
        # PATCH 141c: mesma proteção do PATCH 139/141 (prática oficialmente
        # conhecida sob OUTRA categoria não pode ficar aqui), mas aplicada
        # nesse TERCEIRO ponto — o histórico individual do tutor, montado por
        # (polo, curso). O diagnóstico do PATCH 140/141b achou 2 tutores de
        # Agronomia (Ana Julia Ribeiro Dos Santos, Otavio Augusto Queiroz Dos
        # Santos) com práticas de ENGMAKER no histórico deles — a causa raiz
        # está na própria planilha fonte (a chave calculada pra essas
        # submissões bateu com a chave AGM desses tutores, por algum motivo
        # ainda não confirmado na origem), mas essa proteção pelo menos evita
        # que o dado errado apareça pro usuário: desvia pro bucket anônimo do
        # polo (mesmo padrão já usado pra práticas pré-admissão), que agora
        # tem resolução de categoria por votação (PATCH 141) — deve
        # recategorizar corretamente como Engenharia lá.
        _hist_categoria_errada = 0
        _hist_filtrado = []
        for h in hist_bruto:
            _cat_da_pratica = oficial_p_to_cat.get(h['p']) or oficial_p_to_cat_norm.get(_norm_nome_pratica(h['p']))
            if _cat_da_pratica and _cat_da_pratica != cat_form and _cat_da_pratica != cat_raw:
                polo_sem_tutor[chave].append(h)
                _hist_categoria_errada += 1
            else:
                _hist_filtrado.append(h)
        hist_bruto = _hist_filtrado
        # PATCH 16: práticas enviadas ANTES da admissão do tutor atual não são dele —
        # provavelmente foram enviadas por quem ocupava essa vaga antes (desligado).
        # Vão pro bucket anônimo do polo em vez de ficarem com o tutor novo.
        if _inicio_str:
            hist = []
            for h in hist_bruto:
                if h.get('d') and h['d'] < _inicio_str:
                    polo_sem_tutor[chave].append(h)
                    _hist_pre_admissao += 1
                else:
                    hist.append(h)
        else:
            hist = hist_bruto
        reais    = set(h['p'] for h in hist)
        pend     = [p for p in praticas if p not in reais]
        te = len(reais); tp = len(praticas)

        # PATCH 90: campo só de EXIBIÇÃO com o curso específico (Fisioterapia/
        # T. Ocupacional/Estética e Cosmética) pra Multi III — sem mexer em
        # 'c' (categoria ampla crua), que o resto do sistema usa pra filtro/
        # agrupamento de grupo. Só a coluna "Categoria" na tela usa isso.
        _cat_exibicao = cat_raw
        if str(cat_raw).strip() == 'BIO-FISIO-EST-TO (Multidisciplinar III)':
            _primeiro_curso = str(t.get(col_cur, '') or '').split('|')[0].strip()
            _cat_exibicao = CURSOS_NOMES.get(_primeiro_curso, cat_raw)

        tutores.append({
            'n': str(t.get(col_nome, '') or ''),
            'p': str(t.get(col_polo, '') or ''),
            'c': cat_raw, 'cf': cat_form or 'Sem mapeamento', 'c_exibicao': _cat_exibicao,
            'cursos': str(t.get(col_cur, '') or ''),  # código específico (BFI, BTO, COS-TIP, etc.)
            'tp': tp, 'te': te,
            'pend': pend, 'real': sorted(reais), 'hist': hist,
            'pct': round(te / tp * 100, 1) if tp else 0,
            'ch_semanal': _parse_ch(t.get(col_ch)) if col_ch else None,
            # Campos MEC / CONTROLE
            'inicio': _inicio_str,
            'lattes_url': _mec.get('lattes_url'),
            'lattes_id': _mec.get('lattes_id'),
            'titulacao': _mec.get('titulacao'),
            'graduacao': _mec.get('graduacao'),
            'especializacao': _mec.get('especializacao'),
            'mestrado': _mec.get('mestrado'),
            'doutorado': _mec.get('doutorado'),
            'exp_fora_meses': _mec.get('exp_fora_meses'),
            'exp_tutor_uni_meses': _mec.get('exp_tutor_uni_meses'),
            'whatsapp': str(t.get(col_whats,'') or '') if col_whats else None,
            'chapa': str(t.get(col_chapa,'') or '') if col_chapa else None,
            '_chave_dbg': chave,  # PATCH 141b: só pra diagnóstico, ver bloco logo abaixo
        })
    if _hist_pre_admissao:
        print(f"[{ts()}] Práticas pré-admissão removidas do tutor atual (vão pro polo): {_hist_pre_admissao}")
    seen = {}; tutores_dedup = []
    for t in tutores:
        key = (t.get('p',''), t.get('n','').strip().lower())
        if key in seen:
            existing = seen[key]
            existing['te'] = max(existing['te'], t['te'])
            existing['hist'] = existing['hist'] + [h for h in t['hist'] if h not in existing['hist']]
            existing['real'] = sorted(set(existing['real']) | set(t['real']))
            existing['pend'] = [p for p in existing['pend'] if p not in existing['real']]
            existing['tp'] = max(existing['tp'], t['tp'])
            existing['pct'] = round(existing['te'] / existing['tp'] * 100, 1) if existing['tp'] else 0
            if t.get('ch_semanal') and not existing.get('ch_semanal'):
                existing['ch_semanal'] = t['ch_semanal']
        else:
            seen[key] = t; tutores_dedup.append(t)
    tutores = tutores_dedup

    # Criar entries anônimos para práticas de tutores desligados (vinculado ao polo)
    for chave_polo, hist_polo in polo_sem_tutor.items():
        if not hist_polo: continue
        _reais_polo = set(h['p'] for h in hist_polo)
        # PATCH 141: dois problemas aqui, achados com o log real (diagnóstico
        # PATCH 140 apontou "Tutor desligado" em Manaus/AM com práticas de
        # ENGMAKER, mas categoria resolvida como ENF-INS): (1) essa busca
        # usava 'oficial_p_to_cat' original, sem a normalização de prefixo do
        # PATCH 139 — nomes como "ENGMAKER - Análise estrutural de uma viga"
        # nunca batiam, ficavam de fora; (2) pegava a categoria da PRIMEIRA
        # prática encontrada num 'set' (ordem não garantida/arbitrária) — se
        # esse polo tivesse QUALQUER prática cujo nome batesse (mesmo por
        # coincidência) com outra categoria, a categoria inteira do bucket
        # saía errada. Trocado por votação: conta a categoria de CADA prática
        # reconhecida (agora com fallback normalizado) e usa a mais frequente.
        from collections import Counter as _Counter_cfpolo
        _votos_cf_polo = _Counter_cfpolo()
        for _p_polo in _reais_polo:
            _cat_p = oficial_p_to_cat.get(_p_polo) or oficial_p_to_cat_norm.get(_norm_nome_pratica(_p_polo))
            if _cat_p: _votos_cf_polo[_cat_p] += 1
        _cf_polo = _votos_cf_polo.most_common(1)[0][0] if _votos_cf_polo else ''
        # Normalizar nome da categoria para o padrão do dashboard (usando CAT_MAP invertido)
        _CAT_MAP_INV = {v: k for k, v in CAT_MAP.items()}
        # Também mapear nomes parciais
        _cf_polo_norm = _cf_polo
        for _full, _short in _CAT_MAP_INV.items():
            if _full.lower() in _cf_polo.lower() or _cf_polo.lower() in _full.lower():
                _cf_polo_norm = _short; break
        # Verificar se já está no formato correto (chave do CAT_MAP)
        if _cf_polo_norm not in CAT_MAP:
            for _short in CAT_MAP:
                if _cf_polo.lower() in CAT_MAP[_short].lower():
                    _cf_polo_norm = _short; break
        if _cf_polo_norm in CAT_MAP:
            _cf_polo = _cf_polo_norm
        _praticas_polo = catalogo.get(_cf_polo, [])
        _pend_polo = [p for p in _praticas_polo if p not in _reais_polo]
        # Extrair nome do polo (remover código de curso do fim da chave)
        import re as _re_anon
        _polo_limpo = _re_anon.sub(r'[A-Z]{2,6}(-[A-Z]{2,6})?$', '', chave_polo).strip()
        if not _polo_limpo: _polo_limpo = chave_polo
        tutores.append({
            'n': 'Tutor desligado', 'p': _polo_limpo, 'c': _cf_polo, 'cf': _cf_polo or 'Sem categoria',
            'tp': len(_praticas_polo), 'te': len(_reais_polo),
            'pend': _pend_polo, 'real': sorted(_reais_polo), 'hist': hist_polo,
            'pct': round(len(_reais_polo)/len(_praticas_polo)*100,1) if _praticas_polo else 0,
            'ch_semanal': None, '_anonimo': True,
        })
    if polo_sem_tutor:
        print(f"[{ts()}] Entries anônimos criados: {len(polo_sem_tutor)} polos com práticas de tutores desligados")

    p_to_cat = {}
    for cat, pracs in catalogo.items():
        for p in pracs:
            if p not in p_to_cat: p_to_cat[p] = cat
    ps = defaultdict(lambda: {'enviou': 0, 'nao_enviou': 0, 'categoria': ''})
    for t in tutores:
        for p in t['real']:  ps[p]['enviou']    += 1
        for p in t['pend']:  ps[p]['nao_enviou'] += 1
    _p_fallback = {}
    for t in tutores:
        for p in t['real'] + t['pend']: _p_fallback.setdefault(p, t['cf'])
    for p in ps: ps[p]['categoria'] = p_to_cat.get(p, _p_fallback.get(p, ''))
    ps_all = sorted([{'nome': k, **v} for k, v in ps.items()], key=lambda x: -x['nao_enviou'])

    # PATCH 141b: refinamento do diagnóstico — o PATCH 140 já achou e o
    # PATCH 141 já corrigiu o caso do bucket anônimo "Tutor desligado" (8 das
    # 13 ocorrências, era resolução de categoria por 1ª prática de um set
    # sem ordem). Sobraram 5 ocorrências em tutores REAIS e ativos (Ana Julia
    # Ribeiro Dos Santos, Otavio Augusto Queiroz Dos Santos — cf "Química e
    # Física - Agronomia") com prática de ENGMAKER no 'real' deles — causa
    # DIFERENTE (a submissão real está sendo casada com a chave errada, não
    # é sobre resolução de categoria). Adiciona a CHAVE de cada tutor
    # diagnosticado, pra conseguir comparar contra a chave que a prática
    # ENGMAKER deveria ter batido, e achar o ponto exato da mistura.
    # PATCH 141d: o diagnóstico anterior (PATCH 140/141b) alertava só por
    # padrão de TEXTO no nome da prática ("começa com QUÍMICA E FÍSICA" etc),
    # sem checar se isso realmente diverge da categoria do tutor — dava falso
    # positivo pra tutores de Química e Física/Agronomia com práticas de
    # Química e Física genuinamente deles (confirmado: "Determinação De Ph...
    # " e "Estequiometria..." SÃO oficialmente da categoria dela no catálogo).
    # Agora compara de verdade: resolve a categoria OFICIAL de cada prática
    # (mesma lógica usada nas 3 correções) contra a categoria do tutor, só
    # alerta se as duas divergirem de fato.
    _diagnostico_vazamento = []
    for t in tutores:
        _cf_tutor = str(t.get('cf', '')).strip()
        for p in (t.get('real', []) + t.get('pend', [])):
            _cat_oficial_pratica = oficial_p_to_cat.get(p) or oficial_p_to_cat_norm.get(_norm_nome_pratica(p))
            if not _cat_oficial_pratica: continue  # prática sem correspondência oficial -- não dá pra afirmar nada
            # compara pela forma longa (mais confiável, já que cf do tutor às
            # vezes vem no formato curto tipo "ENGMAKER+QUÍMICA E FÍSICA")
            _cf_tutor_longo = CAT_MAP.get(_cf_tutor, _cf_tutor)
            if _cat_oficial_pratica != _cf_tutor_longo and _cat_oficial_pratica != _cf_tutor:
                _diagnostico_vazamento.append({
                    'pratica': p, 'tutor': t.get('n',''), 'polo': t.get('p',''),
                    'cf_tutor': _cf_tutor, 'categoria_oficial_pratica': _cat_oficial_pratica,
                    'cursos_tutor': t.get('cursos',''), 'chave_tutor': t.get('_chave_dbg',''),
                    'esta_em': 'real' if p in t.get('real', []) else 'pend',
                })
    if _diagnostico_vazamento:
        print(f"[{ts()}] ⚠️  DIAGNÓSTICO PATCH 141d: {len(_diagnostico_vazamento)} prática(s) com categoria oficial DIFERENTE da categoria do tutor que a tem na lista:")
        for d in _diagnostico_vazamento[:15]:
            print(f"    '{d['pratica']}' está em [{d['esta_em']}] do tutor '{d['tutor']}' (polo: {d['polo']}, chave: '{d['chave_tutor']}') — tutor é '{d['cf_tutor']}', mas a prática é oficialmente de '{d['categoria_oficial_pratica']}'")
    else:
        print(f"[{ts()}] DIAGNÓSTICO PATCH 141d: nenhuma divergência real encontrada — todas as práticas reconhecidas oficialmente batem com a categoria do tutor que as tem")

    ps_list = ps_all[:30]
    cs = defaultdict(lambda: {'total_tutores': 0, 'com_100pct': 0, 'total_previstas': 0, 'total_enviadas': 0})
    for t in tutores:
        if not t['tp']: continue
        c = t['cf']
        cs[c]['total_tutores'] += 1
        if t['pct'] == 100: cs[c]['com_100pct'] += 1
        cs[c]['total_previstas'] += t['tp']; cs[c]['total_enviadas'] += t['te']
    print(f"[{ts()}] {len(tutores)} tutores, {sum(len(v) for v in catalogo.values())} praticas")
    prazos = PRAZOS_ORDENS.copy()
    hoje = datetime.now()
    status_ordem = {}
    # PATCH 3: usar datas de início reais de PERIODOS_ORDENS
    for ordem, prazo_str in prazos.items():
        prazo_date = datetime.strptime(prazo_str, '%d/%m/%Y')
        periodo = PERIODOS_ORDENS.get(ordem, {})
        inicio_str = periodo.get('inicio', '')
        if inicio_str:
            try: inicio_date = datetime.strptime(inicio_str, '%d/%m/%Y')
            except ValueError: inicio_date = prazo_date.replace(day=1)
        else: inicio_date = prazo_date.replace(day=1)
        if hoje > prazo_date: status_ordem[ordem] = 'VENCIDO'
        elif hoje >= inicio_date: status_ordem[ordem] = 'ABERTA'
        else: status_ordem[ordem] = 'FUTURA'
    tutores_out = []
    for t in tutores:
        por_ordem = {}
        for h in t['hist']:
            o = h.get('o', 'Ordem 1') or 'Ordem 1'
            por_ordem[o] = por_ordem.get(o, 0) + 1
        # PATCH 4: situação corrigida quando não há ordens vencidas
        ordens_vencidas = [o for o, s in status_ordem.items() if s == 'VENCIDO']
        ordens_abertas  = [o for o, s in status_ordem.items() if s == 'ABERTA']
        if not ordens_vencidas:
            if any(por_ordem.get(o, 0) > 0 for o in ordens_abertas) if ordens_abertas else False:
                sit = 'ok'
            elif ordens_abertas: sit = 'atrasado'
            else: sit = 'ok'
        else:
            if all(por_ordem.get(o, 0) > 0 for o in ordens_vencidas): sit = 'ok'
            elif any(por_ordem.get(o, 0) > 0 for o in ordens_vencidas): sit = 'atrasado'
            else: sit = 'urgente'
        tutores_out.append({
            **t,
            'nome': t.get('n',''), 'polo': t.get('p',''), 'cat': t.get('c',''),
            'n': t.get('n',''), 'p': t.get('p',''), 'c': t.get('c',''),
            'cf': t.get('cf','Sem mapeamento'),
            'por_ordem': por_ordem, 'porOrdem': por_ordem, 'situacao': sit,
        })
    col_email_key = next((c for c in df_t.columns if 'E-MAIL' in str(c).upper()), None)
    nome_to_email = {}
    if col_email_key:
        for _, row in df_at.iterrows():
            nome = str(row.get(col_nome, '') or '').strip()
            email = str(row.get(col_email_key, '') or '').strip().lower()
            if nome and email and email != 'nan': nome_to_email[nome] = email
    seen = {}; tutores_dedup = []
    for t in tutores_out:
        nome = t['n']; polo = t['p']
        email = nome_to_email.get(nome, '')
        key = email if email else (nome + '|' + polo).lower()
        if key not in seen:
            seen[key] = len(tutores_dedup); tutores_dedup.append(dict(t))
        else:
            ex = tutores_dedup[seen[key]]
            ex['hist'] = ex['hist'] + t['hist']
            merged_real = sorted(set(ex['real']) | set(t['real']))
            ex['real'] = merged_real; ex['te'] = len(merged_real)
            for o, cnt in t['por_ordem'].items():
                ex['por_ordem'][o] = ex['por_ordem'].get(o, 0) + cnt
                ex['porOrdem'][o]  = ex['porOrdem'].get(o, 0) + cnt
            real_set = set(merged_real)
            ex['pend'] = [p for p in ex['pend'] if p not in real_set]
            if ex['tp'] > 0: ex['pct'] = round(ex['te'] / ex['tp'] * 100, 1)
            # PATCH 5: reavalia situação + sincroniza sit
            _orv = [o for o, s in status_ordem.items() if s == 'VENCIDO']
            if not _orv: ex['situacao'] = 'ok'
            elif all(ex['por_ordem'].get(o, 0) > 0 for o in _orv): ex['situacao'] = 'ok'
            elif any(ex['por_ordem'].get(o, 0) > 0 for o in _orv): ex['situacao'] = 'atrasado'
            else: ex['situacao'] = 'urgente'
            ex['sit'] = ex['situacao']  # PATCH 5: sync shortcut
            if t.get('ch_semanal') and not ex.get('ch_semanal'): ex['ch_semanal'] = t['ch_semanal']
    tutores_out = tutores_dedup
    print(f"[{ts()}] Após deduplicação: {len(tutores_out)} tutores únicos")
    for _t_clean in tutores_out:
        _t_clean.pop('_chave_dbg', None)  # PATCH 141b: só era pra diagnóstico, não vai pro JSON final

    # PATCH 13: anexa o relatório de acompanhamento/comunicação na ficha do tutor
    def _norm_nome(s):
        s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
        return ' '.join(s.upper().split())
    com_file = os.path.join(SCRIPT_DIR, 'tutores_comunicacao.json')
    if os.path.isfile(com_file):
        with open(com_file, encoding='utf-8') as f: com_lista = json.load(f)
        # PATCH 14: nomes do relatório costumam vir incompletos (ex: "Gabriella Ribeiro"
        # vs "Gabriella Ribeiro Sousa" no cadastro) — casa por prefixo de tokens, não
        # só por igualdade exata
        com_tokens = [(tuple(_norm_nome(c['nome']).split()), c) for c in com_lista]
        _matches = 0
        for t in tutores_out:
            t_tokens = tuple(_norm_nome(t.get('n','')).split())
            achou = None
            for ctoks, c in com_tokens:
                n = min(len(ctoks), len(t_tokens))
                if n >= 2 and ctoks[:n] == t_tokens[:n]:
                    achou = c; break
            if achou:
                t['comunicacao'] = achou
                _matches += 1
        print(f"[{ts()}] Comunicação de tutores: {_matches}/{len(com_lista)} vinculados por nome")

    total     = len(tutores_out)
    enviaram  = sum(1 for t in tutores_out if t['te'] > 0)
    atrasados = sum(1 for t in tutores_out if t['situacao'] == 'atrasado')
    urgentes  = sum(1 for t in tutores_out if t['situacao'] == 'urgente')
    total_alunos = sum(h['a'] for t in tutores_out for h in t['hist'])
    polo_map = {}
    for t in tutores_out:
        p = t['polo']
        if p not in polo_map:
            polo_map[p] = {'POLO': p, 'polo': p, 'n': p, 'total': 0, 'enviaram': 0, 'atrasados': 0, 'alunos': 0}
        polo_map[p]['total'] += 1
        if t['te'] > 0: polo_map[p]['enviaram'] += 1
        if t['situacao'] == 'atrasado': polo_map[p]['atrasados'] += 1
        polo_map[p]['alunos'] += sum(h['a'] for h in t['hist'])
    polo_envios = {}
    for t in tutores_out:
        p = t['polo']
        polo_envios[p] = polo_envios.get(p, 0) + len(t.get('hist', []))
    polo_stats = sorted(polo_map.values(), key=lambda x: -x['atrasados'])
    for p in polo_stats:
        p['n'] = p.get('polo', p.get('POLO', ''))
        p['t'] = p['total']; p['e'] = p['enviaram']; p['a'] = p['alunos']
        p['pend'] = p['total'] - p['enviaram']
        p['pct'] = round(p['enviaram'] / p['total'] * 100) if p['total'] else 0
        p['envios'] = polo_envios.get(p['POLO'], 0)

    # PATCH 124: contatos por polo (Lista Oficial de Contatos), pedido pelo
    # Leo pra aparecer na aba Polos. Match por nome normalizado (maiúsc/
    # minúsc e acento não importam) — os dois lados usam "Cidade/UF" ou
    # "Cidade/UF - Bairro", mas a grafia exata (acento, maiúscula) pode variar
    # entre a planilha de contatos e o CONTROLE_TUTORIA.
    import unicodedata as _ud6
    def _norm_polo_contato(s):
        s = _ud6.normalize('NFD', str(s or '').strip().lower())
        s = ''.join(c for c in s if _ud6.category(c) != 'Mn')
        return s
    _contatos_file = os.path.join(SCRIPT_DIR, 'contatos_por_polo.json')
    _contatos_raw = {}
    if os.path.isfile(_contatos_file):
        with open(_contatos_file, encoding='utf-8') as f: _contatos_raw = json.load(f)
    _contatos_norm = {_norm_polo_contato(k): v for k, v in _contatos_raw.items()}
    _polos_com_contato = 0
    for p in polo_stats:
        _c = _contatos_norm.get(_norm_polo_contato(p['n']))
        p['contatos'] = _c or []
        if _c: _polos_com_contato += 1
    print(f"[{ts()}] Contatos por polo: {_polos_com_contato}/{len(polo_stats)} polos do CONTROLE com contato encontrado (de {len(_contatos_raw)} polos na lista oficial)")

    ordem_map = {o: {'envios': 0, 'alunos': 0} for o in prazos}
    for t in tutores_out:
        for h in t['hist']:
            o = h.get('o', 'Ordem 1') or 'Ordem 1'
            if o in ordem_map: ordem_map[o]['envios'] += 1; ordem_map[o]['alunos'] += h['a']
    por_ordem = [
        {'ordem': o, 'prazo': prazos[o], 'status': status_ordem[o],
         'envios': ordem_map[o]['envios'], 'alunos': ordem_map[o]['alunos']}
        for o in prazos
    ]
    mes_map = {}
    for t in tutores_out:
        for h in t['hist']:
            d = h.get('d') or ''
            mes = d[:7] if d and len(d) >= 7 else 'Sem data'
            if mes not in mes_map: mes_map[mes] = {'MES': mes, 'mes': mes, 'envios': 0, 'alunos': 0}
            mes_map[mes]['envios'] += 1; mes_map[mes]['alunos'] += h.get('a', 0)
    por_mes = sorted(mes_map.values(), key=lambda x: x['mes'])
    por_ordem_dict = {o: ordem_map[o]['envios'] for o in prazos}
    alunos_por_ordem = {o: ordem_map[o]['alunos'] for o in prazos}
    for p in polo_stats:
        p['n'] = p.get('polo', p.get('POLO', ''))
        p['t'] = p.get('total', 0); p['e'] = p.get('enviaram', 0); p['a'] = p.get('alunos', 0)
    for t in tutores_out:
        t['sit'] = t.get('situacao', 'urgente')
        t['al'] = sum(h.get('a', 0) for h in t.get('hist', []))
        t['email'] = nome_to_email.get(t.get('n', ''), '')
    praticas_template = []
    for p in ps_all:
        total_p = p['enviou'] + p['nao_enviou']
        praticas_template.append({
            'n': p['nome'], 'c': p['categoria'],
            'env_n': p['enviou'], 'pend_n': p['nao_enviou'],
            'pct': round(p['enviou'] / total_p * 100, 1) if total_p else 0,
            'nome': p['nome'], 'enviou': p['enviou'], 'nao_enviou': p['nao_enviou'], 'categoria': p['categoria'],
        })
    # ── Estatísticas por semestre ────────────────────────────────────────────
    def _stats_semestre(sem_key, tutores_list, catalogo_dict, prazos_dict, periodos_dict):
        """Gera o mesmo bloco de dados que processar() retorna, mas filtrado por semestre."""
        from datetime import datetime as _dt2
        _hoje = datetime.now()
        _status_ord = {}
        for _o, _pz in prazos_dict.items():
            try:
                _pz_d = _dt2.strptime(_pz, '%d/%m/%Y')
                _ini  = _dt2.strptime(periodos_dict.get(_o,{}).get('inicio',_pz), '%d/%m/%Y')
                if _hoje > _pz_d: _status_ord[_o] = 'VENCIDO'
                elif _hoje >= _ini: _status_ord[_o] = 'ABERTA'
                else: _status_ord[_o] = 'FUTURA'
            except: _status_ord[_o] = 'FUTURA'

        _por_ordem = {}; _alunos_por_ordem = {}
        _polo_map = {}; _cat_stats = {}
        _ps_sem = {}  # PATCH 108

        for _t in tutores_list:
            _sem_antigo = sorted(ALL_SEMESTRES.keys())[0]
            _hist_sem = [h for h in _t.get('hist', []) if h.get('s', _sem_antigo) == sem_key]
            _reais_sem = set(h['p'] for h in _hist_sem)
            _te_sem = len(_reais_sem)
            _tp = _t.get('tp', 0)
            _pct_sem = round(_te_sem / _tp * 100, 1) if _tp else 0

            # PATCH 108: tally de práticas por semestre -- usa o catálogo
            # COMPLETO do tutor (t['real'] + t['pend'], que juntos representam
            # tudo que ele já teve atribuído, em qualquer semestre) como
            # universo, e classifica cada prática como "enviada" ou "pendente"
            # DENTRO DESTE semestre específico -- sem isso, a página de
            # Práticas mostrava sempre o mesmo dado (global/todos os tempos),
            # ignorando completamente o seletor de semestre no Vinci.
            _catalogo_tutor = set(_t.get('real', [])) | set(_t.get('pend', []))
            for _p_nome in _catalogo_tutor:
                if _p_nome not in _ps_sem:
                    _ps_sem[_p_nome] = {'enviou': 0, 'nao_enviou': 0, 'categoria': _t.get('cf', '')}
                if _p_nome in _reais_sem:
                    _ps_sem[_p_nome]['enviou'] += 1
                else:
                    _ps_sem[_p_nome]['nao_enviou'] += 1

            # por_ordem deste semestre
            _po = {}
            for _h in _hist_sem:
                _o = _h.get('o','Ordem 1') or 'Ordem 1'
                _po[_o] = _po.get(_o, 0) + 1
                _por_ordem[_o] = _por_ordem.get(_o, 0) + 1
                _alunos_por_ordem[_o] = _alunos_por_ordem.get(_o, 0) + _h.get('a', 0)

            # situação neste semestre
            _orv = [_o for _o, _s in _status_ord.items() if _s == 'VENCIDO']
            if not _orv: _sit = 'ok' if _te_sem > 0 else 'atrasado'
            elif all(_po.get(_o,0) > 0 for _o in _orv): _sit = 'ok'
            elif any(_po.get(_o,0) > 0 for _o in _orv): _sit = 'atrasado'
            else: _sit = 'urgente'

            # polo
            _p = _t.get('p','')
            if _p not in _polo_map:
                _polo_map[_p] = {'n':_p,'polo':_p,'POLO':_p,'total':0,'enviaram':0,'atrasados':0,'alunos':0,'pend':0,'pct':0,'envios':0,'t':0,'e':0,'a':0}
            _polo_map[_p]['total'] += 1; _polo_map[_p]['t'] += 1
            if _te_sem > 0: _polo_map[_p]['enviaram'] += 1; _polo_map[_p]['e'] += 1
            if _sit == 'atrasado': _polo_map[_p]['atrasados'] += 1
            _polo_map[_p]['alunos'] += sum(h.get('a',0) for h in _hist_sem)
            _polo_map[_p]['a'] = _polo_map[_p]['alunos']

            # cat_stats
            _cf = _t.get('cf','')
            if _cf not in _cat_stats:
                _cat_stats[_cf] = {'total_tutores':0,'com_100pct':0,'total_previstas':0,'total_enviadas':0}
            if _tp:
                _cat_stats[_cf]['total_tutores'] += 1
                if _pct_sem == 100: _cat_stats[_cf]['com_100pct'] += 1
                _cat_stats[_cf]['total_previstas'] += _tp
                _cat_stats[_cf]['total_enviadas'] += _te_sem

        # Calcular pct e pend nos polos
        for _ps in _polo_map.values():
            _ps['pend'] = _ps['total'] - _ps['enviaram']
            _ps['pct']  = round(_ps['enviaram']/_ps['total']*100) if _ps['total'] else 0

        # PATCH 108: monta pratica_stats/praticas deste semestre, no mesmo
        # formato que a versão global (ps_all/praticas_template) usa --
        # assim o frontend pode ler direto sem precisar de tratamento especial.
        _ps_sem_all = sorted([{'nome': k, **v} for k, v in _ps_sem.items()], key=lambda x: -x['nao_enviou'])
        _praticas_sem_template = []
        for _p in _ps_sem_all:
            _total_p = _p['enviou'] + _p['nao_enviou']
            _praticas_sem_template.append({
                'n': _p['nome'], 'c': _p['categoria'],
                'env_n': _p['enviou'], 'pend_n': _p['nao_enviou'],
                'pct': round(_p['enviou'] / _total_p * 100, 1) if _total_p else 0,
                'nome': _p['nome'], 'enviou': _p['enviou'], 'nao_enviou': _p['nao_enviou'], 'categoria': _p['categoria'],
            })

        _total = len(tutores_list)
        _sem_ant = sorted(ALL_SEMESTRES.keys())[0]
        _enviaram = sum(1 for _t in tutores_list if any(h.get('s',_sem_ant)==sem_key and h.get('p') for h in _t.get('hist',[])))
        # Calcular urgentes e atrasados corretamente
        _orv = [_o for _o, _s in _status_ord.items() if _s == 'VENCIDO']
        _urgentes = 0; _atrasados = 0
        for _t in tutores_list:
            _sem_ant2 = sorted(ALL_SEMESTRES.keys())[0]
            _h = [h for h in _t.get('hist', []) if h.get('s', _sem_ant2) == sem_key]
            _po = {}
            for _hh in _h:
                _o = _hh.get('o', 'Ordem 1') or 'Ordem 1'
                _po[_o] = _po.get(_o, 0) + 1
            if _orv:
                _venc_ok = [_o for _o in _orv if _po.get(_o, 0) > 0]
                if len(_venc_ok) == 0:
                    _urgentes += 1
                elif len(_venc_ok) < len(_orv):
                    _atrasados += 1

        return {
            'semestre': sem_key,
            'kpis': {
                'total': _total, 'enviaram': _enviaram, 'pendentes': _total - _enviaram,
                'urgentes': _urgentes, 'atrasados': _atrasados,
                'total_polos': len(_polo_map),
                'polos_ok': sum(1 for _ps in _polo_map.values() if _ps['pend']==0),
            },
            'polo_stats': sorted(_polo_map.values(), key=lambda x: -x.get('pend',0)),
            'por_ordem': _por_ordem,
            'alunos_por_ordem': _alunos_por_ordem,
            'status_ordem': _status_ord,
            'prazos': prazos_dict,
            'periodos': periodos_dict,
            'cat_stats': [{'categoria':k,**v} for k,v in _cat_stats.items()],
            'pratica_stats': _ps_sem_all[:30], 'praticas': _praticas_sem_template,  # PATCH 108
        }

    _dados_por_semestre = {}
    for _sem_k, _sem_cfg in ALL_SEMESTRES.items():
        _dados_por_semestre[_sem_k] = _stats_semestre(
            _sem_k, tutores_out,
            catalogo, _sem_cfg['prazos'], _sem_cfg['periodos']
        )
        _env_s = _dados_por_semestre[_sem_k]['kpis']['enviaram']
        print(f"[{ts()}] Semestre {_sem_k}: {_env_s} tutores com envios")
    # ── Fim estatísticas por semestre ─────────────────────────────────────────

    BRT = timezone(timedelta(hours=-3))
    gerado = datetime.now(BRT).strftime('%d/%m/%Y %H:%M')
    ch_ok = sum(1 for t in tutores_out if t.get('ch_semanal'))
    print(f"[{ts()}] {total} tutores · {enviaram} enviaram · {atrasados} atrasados · {urgentes} urgentes")
    print(f"[{ts()}] CH SEMANAL preenchida: {ch_ok}/{total} tutores")

    # PATCH 25c: snapshot/regressão — colunas detectadas + contagens-chave desta
    # rodada, comparadas contra a rodada anterior (avisos só em log, nunca na UI)
    _verificar_snapshot_regressao(
        colunas_detectadas={
            'CONTROLE.POLO': col_polo, 'CONTROLE.CURSOS': col_cur,
            'CONTROLE.NOME_TUTOR': col_nome, 'CONTROLE.EMAIL': col_email,
            'CONTROLE.CATEGORIA': col_cat,
            'PORTFOLIO.CHAVE': c_chave, 'PORTFOLIO.PROTOCOLOS': c_proto,
            'PORTFOLIO.DATA': c_data, 'PORTFOLIO.ALUNOS': 'soma_estudantes()',
        },
        contagens={
            'total_tutores': total, 'enviaram': enviaram,
            'com_match': com_match, 'sem_match': sem_match,
            'total_submissoes': sum(len(t.get('hist', [])) for t in tutores_out),
        },
    )

    return limpar({
        'kpis': {
            'total': total, 'enviaram': enviaram, 'pendentes': total - enviaram,
            'atrasados': atrasados, 'urgentes': urgentes,
            'total_alunos': total_alunos, 'total_polos': len(polo_map),
            'polos_ok': sum(1 for p in polo_stats if p['enviaram'] > 0),
        },
        'tutores': tutores_out, 'polo_stats': polo_stats,
        'tutores_desligados': tutores_desligados,  # PATCH 105
        'portfolio_alunos_dedup': {  # PATCH 123
            'geral': _port_alunos_geral,
            'por_polo': _port_alunos_por_polo,
            'por_categoria': _port_alunos_por_cat,
        },
        'portfolio_alunos_dedup_por_semestre': portfolio_alunos_dedup_por_semestre,  # PATCH 127
        'por_ordem': por_ordem_dict, 'por_ordem_lista': por_ordem,
        'alunos_por_ordem': alunos_por_ordem, 'status_ordem': status_ordem,
        'cat_stats': [{'categoria': k, **v} for k, v in cs.items()],
        'pratica_stats': ps_list, 'praticas': praticas_template,
        'catalogo': catalogo, 'prazos': prazos,
        'por_mes': por_mes, 'gerado_em': gerado,
        'avisos_portfolio': avisos_portfolio,
        'semestre': SEMESTRE_ATUAL,
        'todos_semestres': sorted(ALL_SEMESTRES.keys()),
        'dados_por_semestre': _dados_por_semestre,
        'disciplinas_por_ordem': _DISCIPLINAS_POR_ORDEM_GLOBAL,
        'laboratorios': _carregar_laboratorios(),
    })


def _carregar_laboratorios():
    # PATCH 13: dados pré-processados da seção "Laboratórios" (gerados fora do
    # pipeline, a partir das bases da Juliana + do dataset de práticas 2026/1)
    path = os.path.join(SCRIPT_DIR, 'laboratorios_data.json')
    if not os.path.isfile(path):
        print(f"[{ts()}] laboratorios_data.json não encontrado — seção Laboratórios fica vazia")
        return {}
    with open(path, encoding='utf-8') as f:
        lab = json.load(f)
    print(f"[{ts()}] Laboratórios: {len(lab.get('ricos',{}))} categorias ricas, {len(lab.get('simples',{}))} categorias simples")
    return lab


def _detectar_e_corrigir_base64(p4):
    import base64 as _b64
    try:
        with open(str(p4), 'rb') as f: raw = f.read(16)
        if raw[:4] == b'PK\x03\x04': return
        with open(str(p4), 'rb') as f: full = f.read()
        for padded in [full.strip(), full.strip() + b'==']:
            try:
                decoded = _b64.b64decode(padded)
                if decoded[:4] == b'PK\x03\x04':
                    with open(str(p4), 'wb') as fw: fw.write(decoded)
                    print(f"  [FIX] Base64 detectado e corrigido ({len(full)}->{len(decoded)} bytes)")
                    return
            except Exception: continue
        if raw[:5] in (b'\r\n<!D', b'<!DOC', b'<html'):
            print(f"  [ERRO] Arquivo é uma página HTML (login Microsoft)")
        else:
            print(f"  [INFO] Arquivo não é ZIP nem base64 ({raw[:4].hex()})")
    except Exception as e: print(f"  [AVISO] Verificação base64 falhou: {e}")


def _ler_lotacao_xlsx(p4):
    from openpyxl import load_workbook as _lwb
    wb = _lwb(str(p4), read_only=True, data_only=True, keep_vba=False)
    ws = wb['Quadro Geral de Lotação'] if 'Quadro Geral de Lotação' in wb.sheetnames else list(wb.worksheets)[0]
    return list(ws.iter_rows(values_only=True))

def _ler_lotacao_xls(p4):
    import xlrd
    wb = xlrd.open_workbook(str(p4))
    try: ws = wb.sheet_by_name('Quadro Geral de Lotação')
    except xlrd.XLRDError: ws = wb.sheet_by_index(0)
    rows = []
    for i in range(ws.nrows):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_DATE:
                import xlrd.xldate
                row.append(xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode))
            else: row.append(cell.value if cell.ctype != xlrd.XL_CELL_EMPTY else None)
        rows.append(tuple(row))
    return rows

def _ler_lotacao_pandas(p4):
    p = str(p4)
    try: df = pd.read_excel(p, sheet_name='Quadro Geral de Lotação', header=None)
    except Exception: df = pd.read_excel(p, sheet_name=0, header=None)
    return [tuple(row) for row in df.fillna('').values.tolist()]

def carregar_lotacao(p4):
    fname = os.path.basename(str(p4))
    print(f"[{ts()}] Lendo lotação de tutores ({fname})...")
    _detectar_e_corrigir_base64(p4)
    _rows = None
    for estrategia, fn in [('openpyxl', _ler_lotacao_xlsx), ('xlrd', _ler_lotacao_xls), ('pandas', _ler_lotacao_pandas)]:
        try:
            _rows = fn(p4)
            print(f"[{ts()}] Lotação lida via {estrategia}: {len(_rows)} linhas")
            break
        except Exception as e: print(f"[{ts()}] Tentativa {estrategia}: {e}")
    if not _rows: raise RuntimeError(f"Não foi possível ler {fname}")
    lotacao = {}
    # Diagnóstico: mostrar cabeçalhos (linha 0) e linha 2 para verificar estrutura
    if _rows:
        _hdrs = [str(c or '').strip()[:20] for c in _rows[0]] if _rows[0] else []
        print(f"[{ts()}] Lotação colunas (linha 1): {_hdrs[:35]}")
        if len(_rows) > 1:
            _r2 = [str(c or '')[:15] for c in _rows[1]]
            print(f"[{ts()}] Lotação linha 2: {_r2[:35]}")
        # Detectar coluna de total_alunos automaticamente
        _col_alunos = 26  # default
        for _ci, _h in enumerate(_hdrs):
            _hu = _h.upper()
            if any(k in _hu for k in ['TOTAL', 'ALUNOS', 'MATR']):
                _col_alunos = _ci
                print(f"[{ts()}] Coluna alunos detectada: {_ci} = '{_h}'")
                break
    else:
        _col_alunos = 26

    for r in _rows[2:]:
        if not r[8] or str(r[8]).strip() in ('', '-', 'None', 'nan'): continue
        nome_raw = str(r[8]).strip(); nome_lower = nome_raw.lower()
        try: total_al = int(float(str(r[27] if len(r) > 27 else 0) or 0))
        except: total_al = 0
        # Colunas detectadas da planilha Lotação de Tutores_2026_2:
        # [5]=CURSOS, [7]=CONTRATAÇÃO, [8]=TUTOR, [14]=PERFIL, [15]=CH SEMANAL
        # [16]=CH IDEAL, [27]=TOTAL ALUNOS, [30]=CATEGORIA GIOCONDA
        lotacao[nome_lower] = {
            'nome_oficial': nome_raw,
            'perfil':       str(r[14] if len(r) > 14 else '') or '',
            'cursos':       str(r[5]  if len(r) > 5  else '') or '',
            'ch_semanal':   _parse_ch(r[15] if len(r) > 15 else None),
            # CH IDEAL: usar r[16] se preenchido, senão CH PROPOSTA (r[13])
            '_ch_ideal_raw': _parse_ch(r[16] if len(r) > 16 else None) or 0.0,
            '_ch_prop_raw':  _parse_ch(r[13] if len(r) > 13 else None) or 0.0,
            'ch_ideal': (_parse_ch(r[16] if len(r) > 16 else None) or
                         _parse_ch(r[13] if len(r) > 13 else None) or 0.0),
            'contratacao':  str(r[7]  if len(r) > 7  else '') or '',
            'polo_hub':     str(r[4]  if len(r) > 4  else '') or '',
            'categoria_gio':str(r[30] if len(r) > 30 else '') or '',
            'total_alunos': total_al,
        }
        # Indexar também por nome primeiro+último para match mais abrangente
        _parts = nome_lower.split()
        if len(_parts) >= 2:
            _nfl = _parts[0] + ' ' + _parts[-1]
            if _nfl not in lotacao:
                lotacao[_nfl] = lotacao[nome_lower]
    print(f"[{ts()}] Lotação: {len(lotacao)} tutores mapeados")
    return lotacao


# PATCH 29: página de Vagas — extrai posições em aberto (Aumento de Quadro /
# Substituição) diretamente da aba "Quadro Geral de Lotação". Usa a coluna
# LOTAÇÃO como sinal de pendência (vazia = posição preenchida, sem vaga).
# IMPORTANTE: por decisão explícita do Leo, NENHUM dado financeiro é lido ou
# exposto aqui — a coluna "Salário Phill" e a aba "Controle Orçamento" ficam
# de fora por completo, mesmo que estejam na mesma planilha-fonte.
def processar_vagas(p4):
    print(f"[{ts()}] Lendo vagas (Lotação)...")
    _rows = None
    for estrategia, fn in [('openpyxl', _ler_lotacao_xlsx), ('xlrd', _ler_lotacao_xls), ('pandas', _ler_lotacao_pandas)]:
        try:
            _rows = fn(p4)
            break
        except Exception as e:
            print(f"[{ts()}] Vagas — tentativa {estrategia}: {e}")
    if not _rows or len(_rows) < 3:
        print(f"[{ts()}] Vagas: não foi possível ler a planilha de lotação")
        return {'vagas': [], 'kpis': {}}

    def _gv(r, i):
        try:
            v = r[i]
            return v if v is not None else ''
        except IndexError:
            return ''

    vagas = []
    for r in _rows[2:]:
        lotacao_status = str(_gv(r, 6)).strip()  # coluna 6 = LOTAÇÃO
        if not lotacao_status:
            continue  # posição preenchida — sem pendência de vaga
        polo = str(_gv(r, 4)).strip()  # coluna 4 = POLO HUB
        if not polo:
            continue
        cursos = str(_gv(r, 5)).strip()  # coluna 5 = CURSOS
        contratacao = str(_gv(r, 7)).strip()  # coluna 7 = CONTRATAÇÃO
        tutor_atual = str(_gv(r, 8)).strip()  # coluna 8 = TUTOR DE PRATICA
        if tutor_atual in ('-', 'None', 'nan'):
            tutor_atual = ''
        status = 'Substituição' if 'Substitui' in lotacao_status else 'Aumento de Quadro'
        # PATCH 29a: 'Aumento de Quadro' só conta como vaga de verdade quando a
        # posição está 100% vazia (sem tutor atribuído) — confirmado pelo Leo.
        # Linhas de "Aumento de Quadro" com tutor já preenchido representam outra
        # coisa (CH a ampliar pra quem já está lá), não uma vaga em aberto.
        # 'Substituição' continua contando mesmo com tutor preenchido (é normal
        # o substituído ainda aparecer ativo até a troca de fato acontecer).
        if status == 'Aumento de Quadro' and tutor_atual:
            continue
        chamado_sydle = str(_gv(r, 10)).strip()  # coluna 10 = CHAMADO SYDLE
        if chamado_sydle in ('None', 'nan', '0'):
            chamado_sydle = ''
        status_chamado = str(_gv(r, 11)).strip()  # coluna 11 = STATUS CHAMADO
        perfil = str(_gv(r, 14)).strip()  # coluna 14 = PERFIL DO TUTOR
        ch_semanal = _parse_ch(_gv(r, 15))  # coluna 15 = CH SEMANAL
        ch_ideal = _parse_ch(_gv(r, 16))  # coluna 16 = CH IDEAL
        prioridade = str(_gv(r, 34)).strip() or 'Sem Prioridade'  # coluna 34
        autorizado = str(_gv(r, 35)).strip()  # coluna 35 = Aumento de Quadro
        vagas.append({
            'polo': polo, 'cursos': cursos, 'perfil': perfil,
            'status': status,
            'contratacao': contratacao, 'tutor_atual': tutor_atual,
            'chamado_sydle': chamado_sydle, 'status_chamado': status_chamado,
            'ch_semanal': ch_semanal, 'ch_ideal': ch_ideal,
            'prioridade': prioridade, 'autorizado': autorizado,
            'alunos_polo': 0,  # PATCH 126: preenchido depois com o hub CSV, se disponível
        })

    total = len(vagas)
    kpis = {
        'total_vagas': total,
        'aumento_quadro': sum(1 for v in vagas if v['status'] == 'Aumento de Quadro'),
        'substituicao': sum(1 for v in vagas if v['status'] == 'Substituição'),
        'com_previsao': sum(1 for v in vagas if 'Com previsão' in v['contratacao']),
        'sem_previsao': sum(1 for v in vagas if 'Sem previsão' in v['contratacao']),
        'nao_liberada': sum(1 for v in vagas if 'liberada' in v['contratacao'].lower()),
        'autorizadas': sum(1 for v in vagas if v['autorizado'].startswith('Autorizado')),
        'prioridade_alta': sum(1 for v in vagas if v['prioridade'] == 'Alta'),
        'com_chamado_aberto': sum(1 for v in vagas if v['chamado_sydle']),
    }
    print(f"[{ts()}] Vagas: {total} pendentes ({kpis['aumento_quadro']} aumento de quadro, {kpis['substituicao']} substituição, {kpis['com_previsao']} com previsão)")
    return {'vagas': vagas, 'kpis': kpis}


CURSOS_NOMES = {
    'EMF-ISN': 'Enfermagem e Instrumentação Cirúrgica', 'EMF-ISN2': 'Enfermagem e Instrumentação Cirúrgica',
    'BFR': 'Farmácia', 'BBI': 'Biomedicina', 'BFI': 'Fisioterapia', 'BTO': 'T. Ocupacional',
    'COS-TIP': 'Estética e Cosmética', 'NTR': 'Nutrição', 'AGM': 'Agronomia',
    'BAU': 'Arquitetura e Urbanismo', 'ECE-ENM-ENS-ENG-EEA-GPI-CDE-OBR-SAN-TER-FSA-SLF-QUI': 'Engenharias e Licenciaturas',
    'BIOMEDICINA': 'Biomedicina', 'FARMÁCIA': 'Farmácia', 'FISIOTERAPIA': 'Fisioterapia',
    'TERAPIA OCUPACIONAL': 'T. Ocupacional', 'NUTRIÇÃO': 'Nutrição', 'AGRONOMIA': 'Agronomia',
    'ARQUITETURA E URBANISMO': 'Arquitetura e Urbanismo',
}

def gerar_onboarding_atualizado(p1, p6, destino):
    """
    PATCH 89: gera/atualiza a planilha de Acompanhamento de Onboarding.
    - p1: caminho do 01_CONTROLE_TUTORIA.xlsx (fonte dos tutores ativos)
    - p6: caminho da Acompanhamento_Onboarding.xlsx ATUAL (pode ser None na
      primeira vez) — usado só pra preservar os flags já marcados
    - destino: onde salvar a planilha atualizada

    Regra: tutor ativo com "INÍCIO" nos últimos 2 meses entra na lista
    automaticamente. Se já estava na planilha antiga, mantém os flags dele
    (Trilha/Checklist/1:1/Observações) tal como estavam. Tutor que já tem as
    3 colunas em "Sim" (virou "apto") sai da lista — não precisa de uma
    coluna a mais pra isso, o próprio critério das 3 colunas já resolve.
    """
    import pandas as _pd
    import openpyxl as _oxl
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
    from openpyxl.utils import get_column_letter as _gcl

    df = _pd.read_excel(p1, sheet_name='Base de Tutores', header=1)
    ativos = df[df['SITUAÇÃO'].astype(str).str.strip().str.upper() == 'ATIVO'].copy()

    def _categoria_exibicao(categoria, cursos):
        cat = str(categoria or '').strip()
        if cat == 'BIO-FISIO-EST-TO (Multidisciplinar III)':
            primeiro = str(cursos or '').split('|')[0].strip()
            return CURSOS_NOMES.get(primeiro, cat)
        return cat

    ativos['_CAT_EXIB'] = ativos.apply(lambda r: _categoria_exibicao(r.get('CATEGORIA'), r.get('CURSOS')), axis=1)

    hoje = _pd.Timestamp.now().normalize()
    corte = hoje - _pd.Timedelta(days=60)
    novos = ativos[_pd.to_datetime(ativos['INÍCIO'], errors='coerce') >= corte].copy()

    # Carrega flags já existentes (se a planilha anterior estiver disponível)
    flags_por_chapa = {}
    if p6:
        try:
            df_old = _pd.read_excel(p6, sheet_name='Onboarding Tutores')
            for _, r in df_old.iterrows():
                chapa = str(r.get('Chapa', '') or '').strip()
                if chapa:
                    flags_por_chapa[chapa] = {
                        'trilha': str(r.get('Trilha de Aprendizagem', '') or '').strip(),
                        'checklist': str(r.get('Checklist Realizado', '') or '').strip(),
                        'um_a_um': str(r.get('1:1 de Gerenciamento', '') or '').strip(),
                        'obs': r.get('Observações', '') or '',
                    }
        except Exception as e:
            print(f"[{ts()}] AVISO: não consegui ler onboarding anterior pra preservar flags: {e}")

    linhas = []
    for _, t in novos.iterrows():
        chapa = str(t.get('CHAPA', '') or '').strip()
        antigos = flags_por_chapa.get(chapa, {})
        trilha = antigos.get('trilha', 'Não') or 'Não'
        checklist = antigos.get('checklist', 'Não') or 'Não'
        um_a_um = antigos.get('um_a_um', 'Não') or 'Não'
        # PATCH 89: tutor com as 3 colunas em "Sim" virou apto — não entra
        # (ou sai) da planilha de acompanhamento.
        if trilha.lower() == 'sim' and checklist.lower() == 'sim' and um_a_um.lower() == 'sim':
            continue
        inicio_val = t.get('INÍCIO')
        inicio_str = _pd.to_datetime(inicio_val).strftime('%d/%m/%Y') if _pd.notna(inicio_val) else ''
        linhas.append({
            'Chapa': chapa, 'Nome do Tutor': t.get('NOME DO TUTOR', ''), 'Polo': t.get('POLO', ''),
            'Categoria': t.get('_CAT_EXIB', ''), 'Data de Início': inicio_str,
            'Trilha de Aprendizagem': trilha, 'Checklist Realizado': checklist,
            '1:1 de Gerenciamento': um_a_um, 'Observações': antigos.get('obs', ''),
        })
    linhas.sort(key=lambda x: x['Data de Início'], reverse=True)

    wb = _oxl.Workbook()
    ws = wb.active
    ws.title = 'Onboarding Tutores'
    headers = ['Chapa', 'Nome do Tutor', 'Polo', 'Categoria', 'Data de Início',
               'Trilha de Aprendizagem', 'Checklist Realizado', '1:1 de Gerenciamento', 'Observações']
    font_header = _Font(name='Arial', bold=True, color='FFFFFF', size=10)
    fill_header = _Fill('solid', fgColor='1B4D3E')
    font_body = _Font(name='Arial', size=10)
    align_center = _Align(horizontal='center', vertical='center')
    align_left = _Align(horizontal='left', vertical='center')
    thin = _Side(style='thin', color='D9D9D9')
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_pendente = _Fill('solid', fgColor='FCE8B2')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font_header; c.fill = fill_header; c.alignment = align_center; c.border = border
    ws.freeze_panes = 'A2'

    for row_idx, linha in enumerate(linhas, start=2):
        for col, h in enumerate(headers, 1):
            v = linha[h]
            c = ws.cell(row=row_idx, column=col, value=v)
            c.font = font_body; c.border = border
            c.alignment = align_left if h in ('Nome do Tutor', 'Polo', 'Categoria', 'Observações') else align_center
            if h in ('Trilha de Aprendizagem', 'Checklist Realizado', '1:1 de Gerenciamento') and str(v).strip().lower() != 'sim':
                c.fill = fill_pendente

    widths = [14, 30, 30, 26, 14, 20, 18, 18, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_gcl(i)].width = w
    ws.row_dimensions[1].height = 28

    ws2 = wb.create_sheet('Instruções')
    instrucoes = [
        'Como preencher — Acompanhamento de Onboarding', '',
        '• Esta planilha lista os tutores contratados nos últimos 2 meses (calculado automaticamente a partir da Data de Início).',
        '• A categoria mostra o curso ESPECÍFICO do tutor (ex: Fisioterapia, T. Ocupacional, Estética e Cosmética dentro do Multi III) — não a categoria ampla.',
        '• Marque "Sim" nas 3 colunas conforme cada etapa for concluída: Trilha de Aprendizagem, Checklist Realizado, 1:1 de Gerenciamento.',
        '• Quando as 3 colunas estiverem "Sim", o tutor sai da lista automaticamente na próxima atualização — é considerado "apto".',
        '• Não altere Chapa, Nome, Polo, Categoria ou Data de Início — são usados para localizar o tutor certo.',
        '• Esta planilha é atualizada automaticamente a cada ciclo — tutores novos entram sozinhos, e o que você já preencheu NÃO é apagado.',
    ]
    for r, texto in enumerate(instrucoes, 1):
        cell = ws2.cell(row=r, column=1, value=texto if texto else None)
        if r == 1:
            cell.font = _Font(name='Arial', bold=True, size=13)
    ws2.column_dimensions['A'].width = 100

    wb.save(destino)
    print(f"[{ts()}] Onboarding atualizado: {len(linhas)} tutores em acompanhamento -> {destino}")

def enriquecer_tutores(dados, lotacao):
    tutores = dados.get('tutores', [])
    matched = 0
    LAB_PARA_CAT = {
        'ENFERMAGEM,INSTRUMENTAÇÃO CIRÚRGICA': 'Enfermagem e Instrumentação Cirúrgica',
        'ENFERMAGEM,INSTRUMENTAÇÃO CIRÚRGICA2': 'Enfermagem e Instrumentação Cirúrgica',
        'BIOMEDICINA': 'Biomedicina', 'FARMÁCIA': 'Farmácia', 'FISIOTERAPIA': 'Fisioterapia',
        'TERAPIA OCUPACIONAL': 'T. Ocupacional',
        'TECNOLOGIA EM ESTÉTICA E COSMÉTICA,ESTÉTICA E IMAGEM PESSOAL': 'Estética e Cosmética',
        'NUTRIÇÃO': 'Nutrição', 'AGRONOMIA': 'Agronomia', 'ARQUITETURA E URBANISMO': 'Arquitetura e Urbanismo',
        'CONSTRUÇÃO DE EDIFÍCIOS,ENGENHARIA CIVIL,ENGENHARIA ELÉTRICA,ENGENHARIA DE PRODUÇÃO,ENGENHARIA MECÂNICA,ENGENHARIA AMBIENTAL E SANITÁRIA,FORMAÇÃO PEDAGÓGICA EM FÍSICA,FÍSICA,GESTÃO DA PRODUÇÃO INDUSTRIAL,CONTROLE DE OBRAS,QUÍMICA,SANEAMENTO AMBIENTAL,SEGUNDA LICENCIATURA EM FÍSICA,TECNOLOGIA EM ENERGIAS RENOVÁVEIS': 'Engenharias e Licenciaturas',
        'ENGENHARIA CIVIL,ENGENHARIA ELÉTRICA,ENGENHARIA DE PRODUÇÃO,ENGENHARIA MECÂNICA': 'Engenharias (Civil/Elét./Prod./Mec.)',
        'EMF-ISN': 'Enfermagem e Instrumentação Cirúrgica', 'BFR': 'Farmácia',
        'BBI': 'Biomedicina', 'BFI': 'Fisioterapia', 'BTO': 'T. Ocupacional',
        'COS-TIP': 'Estética e Cosmética', 'NTR': 'Nutrição', 'AGM': 'Agronomia',
        'BAU': 'Arquitetura e Urbanismo',
    }
    polo_lab_seen = set(); alunos_por_lab_raw = {}
    for nome_lower, info in lotacao.items():
        cursos_raw = info.get('cursos', '').strip().upper()
        polo_hub   = info.get('polo_hub', '').strip()
        total_al   = info.get('total_alunos', 0)
        if not cursos_raw or not total_al: continue
        chave_polo_lab = f"{polo_hub}||{cursos_raw}"
        if chave_polo_lab in polo_lab_seen: continue
        polo_lab_seen.add(chave_polo_lab)
        sep = '+' if '+' in cursos_raw else ','
        componentes = sorted([c.strip() for c in cursos_raw.split(sep)])
        lab_key = (','.join(componentes) if any(len(c) > 8 for c in componentes) else '+'.join(componentes))
        alunos_por_lab_raw[lab_key] = alunos_por_lab_raw.get(lab_key, 0) + total_al
    def _norm_lab_key(k):
        sep = '+' if '+' in k else ','
        partes = sorted([p.strip().upper() for p in k.split(sep)])
        return (','.join(partes) if any(len(p) > 8 for p in partes) else '+'.join(partes))
    lab_cat_norm = {_norm_lab_key(k): v for k, v in LAB_PARA_CAT.items()}
    alunos_por_curso = []
    if alunos_por_lab_raw:
        # Fonte 1: TOTAL ALUNOS da lotação (quando preenchido)
        for lab_key, total in sorted(alunos_por_lab_raw.items(), key=lambda x: -x[1]):
            nome = lab_cat_norm.get(lab_key)
            if not nome:
                primeiro = lab_key.split(',')[0].split('+')[0].strip()
                nome = CURSOS_NOMES.get(primeiro, primeiro.title())
            alunos_por_curso.append({'sigla': lab_key, 'curso': nome, 'alunos': total})
    else:
        # Fonte 2 (fallback): agrupar matrículas do hub CSV por categoria
        # Usa alunos_hub_por_grupo gerado no processar_alunos_hub()
        # Fallback: usar por_cat do hub CSV (matrículas únicas por categoria)
        # Disponível em dados['hub']['por_cat'] após processar_alunos_hub()
        _hub_por_cat = (dados.get('hub') or {}).get('por_cat', {})
        _CAT_NOME = {
            'ENF-INS (Multidisciplinar II)':        'Enfermagem e Instrumentação Cirúrgica',
            'BIO-FAR (Multidisciplinar I)':         'Biomedicina e Farmácia',
            'BIO-FISIO-EST-TO (Multidisciplinar III)': 'Fisioterapia, T.Ocupacional e Estética',
            'NUTRI (Multidisciplinar IV)':          'Nutrição',
            'ENGMAKER':                             'Engenharias e Licenciaturas',
            'QUÍMICA E FÍSICA':                     'Química e Física',
        }
        if _hub_por_cat:
            for sigla, total in sorted(_hub_por_cat.items(), key=lambda x: -x[1]):
                if total > 0:
                    alunos_por_curso.append({
                        'sigla': sigla,
                        'curso': _CAT_NOME.get(sigla, sigla),
                        'alunos': int(total)
                    })
            print(f"[{ts()}] Alunos por curso (hub CSV): {len(alunos_por_curso)} categorias, total {sum(x['alunos'] for x in alunos_por_curso)}")
        else:
            print(f"[{ts()}] Alunos por curso: sem dados disponíveis")
    dados['alunos_por_curso'] = alunos_por_curso
    total_al_sum = sum(x['alunos'] for x in alunos_por_curso)
    print(f"[{ts()}] Alunos por lab: {len(alunos_por_curso)} labs, total {total_al_sum:,}")

    import unicodedata as _ud_ch, re as _re_ch
    def _norm_ch(s):
        s = str(s or '').strip().lower()
        s = _ud_ch.normalize('NFD', s)
        s = ''.join(c for c in s if _ud_ch.category(c) != 'Mn')
        return _re_ch.sub(r'\s+', ' ', s)
    def _eh_subsequencia_ch(curtos, longos):
        i = 0
        for tok in longos:
            if i < len(curtos) and tok == curtos[i]:
                i += 1
        return i == len(curtos)
    def _nomes_batem_ch(nome_a, nome_b):
        if nome_a == nome_b:
            return True
        ta, tb = nome_a.split(), nome_b.split()
        if not ta or not tb:
            return False
        if len(ta) >= 2 and len(tb) >= 2 and ta[0] == tb[0] and ta[-1] == tb[-1]:
            return True
        curtos, longos = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
        if len(curtos) < 2:
            return False
        return _eh_subsequencia_ch(curtos, longos)
    _lotacao_norm = {_norm_ch(k): v for k, v in lotacao.items()}
    _matched_subsequencia_ch = 0
    for t in tutores:
        nome_lower = str(t.get('n', '')).lower()
        info = lotacao.get(nome_lower)
        if not info:
            nome_norm_t = _norm_ch(nome_lower)
            info = _lotacao_norm.get(nome_norm_t)
            if not info:
                # PATCH 96: fallback por subsequência (nome com parte a mais/a
                # menos, ou uma palavra do meio diferente) — antes disso, só
                # existia um teste de substring simples, que não cobre o caso
                # de palavra SUBSTITUÍDA no meio (ex: "Souza da Silva" vs
                # "Souza De Silva").
                for k_norm, v in _lotacao_norm.items():
                    if _nomes_batem_ch(nome_norm_t, k_norm):
                        info = v
                        _matched_subsequencia_ch += 1
                        break
        if info:
            t['perfil'] = info['perfil']; t['cursos'] = info['cursos']
            # Lotação tem prioridade sobre CH da planilha de controle
            if info.get('ch_semanal') and info['ch_semanal'] > 0:
                t['ch_semanal'] = info['ch_semanal']
            t['ch_ideal'] = info.get('ch_ideal', 0)
            t['contratacao_lot'] = info['contratacao']
            t['lab'] = info.get('cursos', '')  # curso da planilha de lotação (para Multi 3)
            t['polo_hub_lot'] = info.get('polo_hub', '')
            matched += 1
    print(f"[{ts()}] Enriquecimento: {matched}/{len(tutores)} tutores com perfil/CH ({_matched_subsequencia_ch} via correspondência por subsequência)")
    # ── Adicionar tutores sintéticos para avisos (aparecem na aba Tutores) ────
    # PATCH 120: dois casos confirmados pelo Leo em 21/08 como erro de
    # preenchimento no Forms (não erro de matching) — o tutor É real e a
    # submissão de portfólio deve contar pra ele normalmente, não cair em
    # "Aviso de Portfólio". Categoria/polo confirmados via atividade real de
    # gerenciamento no GIOCONDA (não veio do CONTROLE, que não tem registro
    # localizável pra essas duas pessoas sob esse nome). Correção pontual e
    # nominal — não altera a lógica geral de matching de chave/e-mail/nome,
    # só esses dois casos já investigados e confirmados.
    _CORRECOES_MANUAIS_AVISO = {
        'ingrid schroeder pineiro': {
            'polo': 'Imbituba/SC',
            'c': 'ENF-INS (Multidisciplinar II)', 'cf': 'ENF-INS (Multidisciplinar II)',
        },
        'kellen ferreira nascimento': {
            'polo': 'Goiânia/GO - Jardim Europa',
            'c': 'NUTRI (Multidisciplinar IV)', 'cf': 'NUTRI (Multidisciplinar IV)',
        },
    }
    def _norm_nome_aviso(s):
        import unicodedata as _ud
        s = str(s or '').strip().lower()
        s = _ud.normalize('NFD', s)
        return ''.join(c for c in s if _ud.category(c) != 'Mn')

    _avisos_enr = dados.get('avisos_portfolio', [])
    if _avisos_enr:
        for av in _avisos_enr:
            if av['nome'] and av['nome'] not in ('nan', '-', ''):
                nome_display = av['nome']
            else:
                # Extrair nome do email
                local = av['email'].split('@')[0] if '@' in av['email'] else av['email']
                nome_display = local.replace('.', ' ').replace('_', ' ').title()
            _correcao = _CORRECOES_MANUAIS_AVISO.get(_norm_nome_aviso(nome_display))
            if _correcao:
                tutores.append({
                    'n': nome_display,
                    'p': _correcao['polo'],
                    'c': _correcao['c'], 'cf': _correcao['cf'], 'c_exibicao': _correcao['c'],
                    'cursos': '',
                    'tp': 0, 'te': av['count'],
                    'pend': [], 'real': [], 'hist': [],
                    'pct': None,  # tp desconhecido (não temos o catálogo real dela) — não fabricar 100%
                    'ch_semanal': None,
                    'correcao_manual': True,
                    'correcao_manual_motivo': 'Erro de preenchimento no Forms (categoria/polo digitados errado) — confirmado com o Leo em 21/08, submissão é real.',
                })
                continue
            tutores.append({
                'n': nome_display,
                'p': av.get('chave', '').replace('BFR-BBI','').replace('EMF-ISN','').replace('NTR','').strip(),
                'c': 'Aviso de Portfólio', 'cf': 'Aviso de Portfólio',
                'tp': 0, 'te': av['count'],
                'pend': [], 'real': [], 'hist': [],
                'pct': 0,
                'ch_semanal': None,
                'aviso_tipo': av['tipo'],
                'aviso_msg': av['msg'],
                'aviso_email': av['email'],
                'aviso_count': av['count'],
            })
    dados['tutores'] = tutores


    # avisos_portfolio já está em dados (vindo de processar()) — não sobrescrever com []
    if 'avisos_portfolio' not in dados:
        dados['avisos_portfolio'] = []
    return dados


def processar_gerenciamento_csv(p5):
    """Processa CSV detalhado de gerenciamento (REL_DETALHADO.csv)."""
    import csv, re as _re
    print(f"[{ts()}] Lendo gerenciamento (CSV)...")
    for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            with open(str(p5), 'r', encoding=enc, errors='replace') as f:
                rows = list(csv.reader(f, delimiter=';'))
            break
        except: continue
    header = rows[0]; data = rows[1:]
    col = {h.strip().upper(): i for i, h in enumerate(header)}
    def gc(name): return col.get(name.upper())
    ci_polo = gc('LABORATORIO'); ci_cat = gc('CATEGORIA'); ci_exp = gc('NOME_EXPERIMENTO')
    ci_tutor = gc('TUTOR'); ci_mat = gc('ALUNOS_MATRICULADOS'); ci_agend = gc('ALUNOS_AGENDADOS')
    ci_capa = gc('CAPACIDADE_TOTAL'); ci_ofe = gc('OFERTAS_CADASTRADAS')
    ci_situ = gc('SITU_OFERTA'); ci_dt_ag = gc('DT_GERENCIADA'); ci_hr_ag = gc('HR_GERENCIADA')
    def gv(row, ci, default=''):
        try: return str(row[ci]).strip() if ci is not None and ci < len(row) else default
        except: return default
    def gn(row, ci):
        try: return float(str(row[ci]).replace(',','.').strip()) if ci is not None and ci < len(row) and row[ci] else 0
        except: return 0
    print(f"[{ts()}] Gerenciamento CSV: {len(data)} linhas, {len(header)} colunas")
    def extrair_ordem_exp(val):
        m = _re.match(r'O\.(\d+):\s*(.*)', str(val or ''))
        if m: return f'Ordem {m.group(1)}', m.group(2).strip()
        return '', str(val or '').strip()
    registros = []
    for r in data:
        polo = gv(r, ci_polo)
        if not polo: continue
        cat = gv(r, ci_cat); exp = gv(r, ci_exp); tutor = gv(r, ci_tutor); situ = gv(r, ci_situ)
        ordem, pratica = extrair_ordem_exp(exp)
        mat = int(gn(r, ci_mat)); agend = int(gn(r, ci_agend))
        capa = int(gn(r, ci_capa)); ofe = int(gn(r, ci_ofe))
        dt_ag = gv(r, ci_dt_ag); hr_ag = gv(r, ci_hr_ag)
        dt_ag_iso = ''
        if dt_ag and '/' in dt_ag:
            try:
                parts = dt_ag.split('/')
                dt_ag_iso = f"{parts[2]}-{parts[1]}-{parts[0]}"
            except: pass
        registros.append({
            'polo': polo, 'categoria': cat, 'pratica': pratica, 'ordem': ordem,
            'tutor': tutor if bool(tutor) else '',
            'tem_tutor': bool(tutor), 'tem_agenda': bool(dt_ag_iso),
            'gerenciado': ofe > 0, 'situ': situ,
            'alunos_mat': mat, 'alunos_agend': agend, 'capacidade': capa,
            'dt_agenda_iso': dt_ag_iso, 'hr_agenda': hr_ag,
        })
    df_r = pd.DataFrame(registros)
    if df_r.empty:
        return {'ger_kpis': {}, 'ger_polo': [], 'ger_cat': [], 'ger_ordem': [],
                'ger_contratacao': [], 'ger_agendas': [], 'ger_ofertas': []}
    polo_cat_tem_tutor = (
        df_r[df_r['tem_tutor']].groupby(['polo','categoria']).size()
        .reset_index(name='_qt').assign(_tem=True)
        .set_index(['polo','categoria'])['_tem']
    )
    def _fix_tem_tutor(row):
        return polo_cat_tem_tutor.get((row['polo'], row['categoria']), row['tem_tutor'])
    df_r['tem_tutor'] = df_r.apply(_fix_tem_tutor, axis=1)
    total = len(df_r); com_tutor = int(df_r['tem_tutor'].sum()); sem_tutor = total - com_tutor
    gerenciadas = int(df_r['gerenciado'].sum()); com_agenda = int(df_r['tem_agenda'].sum())
    tot_mat = int(df_r['alunos_mat'].sum()); tot_agend = int(df_r['alunos_agend'].sum())
    tot_capa = int(df_r['capacidade'].sum())
    print(f"[{ts()}] Gerenciamento: {total} ofertas, {gerenciadas} ger., {sem_tutor} sem tutor")
    ger_kpis = {
        'total_ofertas': total, 'ofertas_gerenciadas': gerenciadas,
        'ofertas_nao_gerenciadas': total - gerenciadas,
        'pct_gerenciado': round(gerenciadas/total*100,1) if total else 0,
        'ofertas_com_tutor': com_tutor, 'ofertas_sem_tutor': sem_tutor,
        'pct_com_tutor': round(com_tutor/total*100,1) if total else 0,
        'ofertas_com_agenda': com_agenda, 'total_alunos_matriculados': tot_mat,
        'total_alunos_agendados': tot_agend, 'total_capacidade': tot_capa,
        'pct_ocupacao': round(tot_agend/tot_capa*100,1) if tot_capa else 0,
        'polos_total': df_r['polo'].nunique(),
        'polos_sem_tutor': int(df_r[~df_r['tem_tutor']].groupby('polo').ngroups),
    }
    ger_polo = []
    for polo, grp in df_r.groupby('polo'):
        tuts = list(grp[grp['tem_tutor']]['tutor'].dropna().unique())
        ger_polo.append({
            'polo': str(polo), 'total_ofertas': len(grp),
            'gerenciadas': int(grp['gerenciado'].sum()),
            'pct_gerenciado': round(grp['gerenciado'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['tem_tutor'].sum()), 'sem_tutor': int((~grp['tem_tutor']).sum()),
            'com_agenda': int(grp['tem_agenda'].sum()),
            'alunos_matriculados': int(grp['alunos_mat'].sum()), 'alunos_agendados': int(grp['alunos_agend'].sum()),
            'capacidade': int(grp['capacidade'].sum()), 'tutores_unicos': [str(t) for t in tuts],
        })
    ger_polo.sort(key=lambda x: -x['sem_tutor'])
    ger_cat = []
    for cat, grp in df_r.groupby('categoria'):
        ger_cat.append({
            'categoria': str(cat), 'total_ofertas': len(grp),
            'gerenciadas': int(grp['gerenciado'].sum()),
            'pct_gerenciado': round(grp['gerenciado'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['tem_tutor'].sum()), 'sem_tutor': int((~grp['tem_tutor']).sum()),
            'alunos_matriculados': int(grp['alunos_mat'].sum()), 'alunos_agendados': int(grp['alunos_agend'].sum()),
        })
    ger_cat.sort(key=lambda x: -x['total_ofertas'])
    ger_ordem = []
    ordem_sort = {'Ordem 1':1,'Ordem 2':2,'Ordem 3':3,'Ordem 4':4,'Ordem 5':5}
    for ordem in sorted(df_r['ordem'].unique(), key=lambda x: ordem_sort.get(x,9)):
        if not ordem: continue
        grp = df_r[df_r['ordem']==ordem]
        ger_ordem.append({
            'ordem': ordem, 'total_ofertas': len(grp),
            'gerenciadas': int(grp['gerenciado'].sum()),
            'pct_gerenciado': round(grp['gerenciado'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['tem_tutor'].sum()),
            'alunos_matriculados': int(grp['alunos_mat'].sum()), 'alunos_agendados': int(grp['alunos_agend'].sum()),
            'dt_inicio': '', 'dt_fim': PRAZOS_ORDENS.get(ordem,''),
            # PATCH 86 (P6): quantos tutores distintos gerenciaram QUALQUER coisa
            # nesta ordem — base pro gráfico "tutores gerenciaram por ordem",
            # atualizando conforme cada ordem acontece (não é % de ofertas, é
            # contagem de PESSOAS, seguindo a mesma regra já validada de
            # "geriu alguma coisa = gerenciou" — não precisa bater a capacidade.
            'tutores_gerenciaram': int(grp[grp['gerenciado']]['tutor'].dropna().nunique()),
        })
    ger_contratacao = []
    for (polo, cat), grp in df_r.groupby(['polo','categoria']):
        tuts = list(grp[grp['tem_tutor']]['tutor'].dropna().unique())
        ger_contratacao.append({
            'polo': str(polo), 'categoria': str(cat), 'total_ofertas': len(grp),
            'tem_tutor': len(tuts)>0, 'tutores': [str(t) for t in tuts],
            'status': 'Contratado' if len(tuts)>0 else 'Sem tutor',
        })
    ger_agendas = []
    for polo, grp in df_r.groupby('polo'):
        total_p = len(grp); com_ag = int(grp['tem_agenda'].sum()); sem_ag = total_p - com_ag
        datas_por_cat = {}; datas_por_tutor = {}
        for _, row in grp[grp['tem_agenda']].iterrows():
            d = row['dt_agenda_iso']; c = row['categoria']; t = row['tutor'] or ''
            if d:
                if d not in datas_por_cat: datas_por_cat[d] = {'cats': [], 'tutores': []}
                if c and c not in datas_por_cat[d]['cats']: datas_por_cat[d]['cats'].append(c)
                if t and t not in datas_por_cat[d]['tutores']: datas_por_cat[d]['tutores'].append(t)
        # PATCH 7: estrutura completa de datas_por_tutor
        for d, v in datas_por_cat.items():
            datas_por_tutor[d] = v['tutores']
        ger_agendas.append({
            'polo': str(polo), 'total': total_p, 'com_agenda': com_ag, 'sem_agenda': sem_ag,
            'pct_agendado': round(com_ag/total_p*100, 1) if total_p else 0,
            'datas_agenda': sorted(datas_por_cat.keys()),
            'datas_por_cat': {d: v['cats'] for d, v in datas_por_cat.items()},
            'datas_por_tutor': datas_por_tutor,  # PATCH 7
        })
    ger_agendas.sort(key=lambda x: -x['sem_agenda'])
    ger_ofertas_detalhe = []
    for _, row in df_r.iterrows():
        ger_ofertas_detalhe.append({
            'polo': row['polo'], 'categoria': row['categoria'],
            'ordem': row['ordem'], 'pratica': row['pratica'],
            'tutor': row['tutor'], 'tem_tutor': row['tem_tutor'],
            'tem_agenda': row['tem_agenda'], 'gerenciado': row['gerenciado'],
            'alunos_mat': row['alunos_mat'], 'alunos_agend': row['alunos_agend'],
            'dt_agenda': row['dt_agenda_iso'], 'hr_agenda': row['hr_agenda'],
        })
    return {
        'ger_kpis': ger_kpis, 'ger_polo': ger_polo, 'ger_cat': ger_cat,
        'ger_ordem': ger_ordem, 'ger_contratacao': ger_contratacao,
        'ger_agendas': ger_agendas, 'ger_ofertas': ger_ofertas_detalhe,
    }


def _processar_gerenciamento_novo(df_g):
    import re as _re
    col = {str(c).strip().upper(): c for c in df_g.columns}
    def gc(name): return col.get(name.upper())
    c_polo = gc('LABORATORIO'); c_cat = gc('CATEGORIA'); c_exp = gc('NOME_EXPERIMENTO')
    c_tutor = gc('TUTOR'); c_mat = gc('ALUNOS_MATRICULADOS'); c_agend = gc('ALUNOS_AGENDADOS')
    c_capa = gc('CAPACIDADE_TOTAL'); c_ofe = gc('OFERTAS_CADASTRADAS'); c_situ = gc('SITU_OFERTA')
    c_dt_ag = gc('DT_GERENCIADA'); c_hr_ag = gc('HR_GERENCIADA'); c_cursos = gc('CURSOS')
    def extrair_ordem_exp(val):
        m = _re.match(r'O\.(\d+):\s*(.*)', str(val or ''))
        if m: return f'Ordem {m.group(1)}', m.group(2).strip()
        return '', str(val or '').strip()
    df = df_g.copy()
    df['_POLO']  = df[c_polo].astype(str).str.strip() if c_polo else ''
    # PATCH 19/82: o export do gerenciamento já trocou de rótulo pra essa mesma
    # categoria (Fisio/T.O./Estética) mais de uma vez — primeiro "FISIO-TO-EST-
    # BIO", agora "BIO-BIO-FISIO-EST-TO" (prefixo "BIO-" duplicado num export
    # mais recente). Cada variante nova virava uma "categoria fantasma" separada
    # nos filtros/agregações em vez de cair na mesma categoria de sempre. Além
    # da lista de variantes conhecidas, adiciona uma regra geral: se o rótulo
    # começar com "BIO-" duplicado (ex: "BIO-BIO-..."), colapsa pro rótulo
    # correto — proteção pra a PRÓXIMA vez que isso acontecer de novo.
    _CAT_RAW_NORM = {
        'FISIO-TO-EST-BIO (Multidisciplinar III)': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
        'BIO-BIO-FISIO-EST-TO (Multidisciplinar III)': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
    }
    def _corrige_prefixo_bio_duplicado(s):
        s2 = str(s or '').strip()
        while s2.upper().startswith('BIO-BIO-'):
            s2 = s2[4:]  # remove um "BIO-" duplicado da frente, pode acontecer mais de uma vez
        return s2
    df['_CAT']   = (df[c_cat].astype(str).str.strip().replace(_CAT_RAW_NORM).apply(_corrige_prefixo_bio_duplicado)) if c_cat  else ''
    # PATCH 42: curso específico (BFI/BTO/COS-TIP/etc.) — pedido do Leo pra poder
    # filtrar Multidisciplinar III por especialidade (Fisioterapia/T.O./Estética)
    # em vez de só pela categoria ampla, que mistura as três no mesmo filtro.
    _SUBCURSO_LABEL = {
        'BFI': 'Fisioterapia', 'BTO': 'Terapia Ocupacional',
        'COS-TIP': 'Estética e Cosmética', 'TIP-COS': 'Estética e Cosmética', 'COS': 'Estética e Cosmética',
        'BBI': 'Biomedicina', 'BFR': 'Farmácia',
        'EMF-ISN': 'Enfermagem/Instrumentação', 'NTR': 'Nutrição',
    }
    def _extrair_curso(v):
        s = str(v or '').strip()
        if not s or s == 'nan': return ''
        primeiro = s.split('|')[0].strip()
        return primeiro
    df['_CURSO'] = df[c_cursos].apply(_extrair_curso) if c_cursos else ''
    df['_SUBCURSO'] = df['_CURSO'].map(lambda c: _SUBCURSO_LABEL.get(c, c))
    df['_TUTOR'] = df[c_tutor].fillna('').astype(str).str.strip().replace('nan','') if c_tutor else ''
    # PATCH 115: a fonte do GIOCONDA passou a trazer o nome do tutor com um
    # número de chapa colado no final, tipo "Beatriz Henkels (17124304)" --
    # isso quebrava silenciosamente qualquer comparação com o nome limpo em
    # DB.tutores (nenhum é igual ao outro), inflando a contagem de "tutores
    # únicos" na aba Detalhe (365 em vez dos 342/347 certos) e provavelmente
    # atrapalhando outros cruzamentos por nome também. Remove o sufixo antes
    # de qualquer outro processamento usar esse nome.
    df['_TUTOR'] = df['_TUTOR'].str.replace(r'\s*\(\d{4,}\)\s*$', '', regex=True).str.strip()
    df['_MAT']   = pd.to_numeric(df[c_mat],  errors='coerce').fillna(0).astype(int) if c_mat  else 0
    df['_AGEND'] = pd.to_numeric(df[c_agend],errors='coerce').fillna(0).astype(int) if c_agend else 0
    df['_CAPA']  = pd.to_numeric(df[c_capa], errors='coerce').fillna(0).astype(int) if c_capa  else 0
    df['_OFE']   = pd.to_numeric(df[c_ofe],  errors='coerce').fillna(0).astype(int) if c_ofe   else 0
    df['_TEM_TUTOR'] = df['_TUTOR'].str.len() > 0
    _situ_col = df[c_situ].fillna('').astype(str).str.strip() if c_situ else pd.Series([''] * len(df))
    dt_col = df[c_dt_ag] if c_dt_ag else pd.Series([''] * len(df))
    def to_iso(v):
        if v is None: return ''
        try:
            import datetime as _dt
            if isinstance(v, (_dt.datetime, _dt.date)): return v.strftime('%Y-%m-%d')
        except: pass
        sv = str(v).strip()
        if not sv or sv == 'nan': return ''
        if '/' in sv:
            try:
                parts = sv.split('/')
                if len(parts) == 3: return f'{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}'
            except: pass
        if '-' in sv and len(sv) >= 10: return sv[:10]
        try:
            n = float(sv)
            import datetime as _dt
            base = _dt.date(1899, 12, 30)
            return (base + _dt.timedelta(days=int(n))).strftime('%Y-%m-%d')
        except: pass
        return ''
    df['_DT_AG_ISO'] = dt_col.apply(to_iso)
    df['_TEM_AGENDA'] = df['_DT_AG_ISO'].str.len() > 0
    # PATCH 22: GERENCIADO = tem tutor E tem data de gerenciamento (DT_GERENCIADA
    # preenchida) — confirmado por Leo: TUTOR preenchido só indica quem está
    # responsável, não que o gerenciamento foi feito; o sinal real de conclusão
    # é a data em DT_GERENCIADA. O critério anterior (tutor + ofertas cadastradas)
    # inflava a contagem de "gerenciadas" pra muito além do que foi feito de fato.
    df['_GERENCIADO'] = df['_TEM_TUTOR'] & df['_TEM_AGENDA']
    df['_HR_AG'] = df[c_hr_ag].fillna('').astype(str).str.strip().replace('nan','').replace('NaT','') if c_hr_ag else ''
    # PATCH 30: dia da semana e turno derivados de DT/HR_GERENCIADA — usados na
    # nova seção "Análise de Agendas" (horários incomuns + sessões sem aluno).
    df['_DIA_SEMANA'] = df['_DT_AG_ISO'].apply(_dia_semana_pt)
    df['_TURNO'] = df['_HR_AG'].apply(_turno_de_horario)
    df['_HORARIO_INCOMUM'] = df['_TEM_AGENDA'] & ((df['_DIA_SEMANA'] == 'Domingo') | (df['_TURNO'] == 'Madrugada'))
    df['_SEM_ALUNOS'] = df['_GERENCIADO'] & (df['_AGEND'] == 0)
    parsed = (df[c_exp] if c_exp else pd.Series([''] * len(df))).apply(extrair_ordem_exp)
    df['_ORDEM'] = parsed.apply(lambda x: x[0])
    df['_PRATICA'] = parsed.apply(lambda x: x[1])
    df = df[df['_POLO'].str.len() > 0].copy()
    total = len(df); com_tutor = int(df['_TEM_TUTOR'].sum()); gerenciadas = int(df['_GERENCIADO'].sum())
    com_agenda = int(df['_TEM_AGENDA'].sum())
    # FIX: Alunos Matriculados — deduplicar por polo×categoria (remove contagem múltipla por ordem)
    _mat_col = '_MAT'; _agend_col = '_AGEND'; _capa_col = '_CAPA'
    _raw_mat = int(df[_mat_col].sum()) if _mat_col in df.columns else 0
    _grp_cols_ok = ['_POLO','_CAT']
    if all(c in df.columns for c in _grp_cols_ok + [_mat_col, _agend_col, _capa_col]):
        _dedup = df.groupby(_grp_cols_ok)[[_mat_col, _agend_col, _capa_col]].max()
        tot_mat   = int(_dedup[_mat_col].sum())
        tot_agend = int(_dedup[_agend_col].sum())
        tot_capa  = int(_dedup[_capa_col].sum())
        print(f"[{ts()}] Alunos DEDUPLICADOS por polo×cat: {tot_mat:,} (bruto era {_raw_mat:,}, redução: {_raw_mat-tot_mat:,})")
    else:
        tot_mat   = _raw_mat
        tot_agend = int(df[_agend_col].sum()) if _agend_col in df.columns else 0
        tot_capa  = int(df[_capa_col].sum())  if _capa_col  in df.columns else 0
        print(f"[{ts()}] Alunos sem dedup: {tot_mat:,}")
    print(f"[{ts()}] Gerenciamento: {total} ofertas, {gerenciadas} ger., {total-com_tutor} sem tutor")
    print(f"[{ts()}] Agendas: {com_agenda} · datas: {sorted(df[df['_TEM_AGENDA']]['_DT_AG_ISO'].head(3).tolist())}")
    print(f"[{ts()}] {df['_POLO'].nunique()} polos, {df['_CAT'].nunique()} cats, {df['_ORDEM'].nunique()} ordens")
    ger_kpis = {
        'total_ofertas': total, 'ofertas_gerenciadas': gerenciadas,
        'ofertas_nao_gerenciadas': total - gerenciadas,
        'pct_gerenciado': round(gerenciadas/total*100,1) if total else 0,
        'ofertas_com_tutor': com_tutor, 'ofertas_sem_tutor': total-com_tutor,
        'pct_com_tutor': round(com_tutor/total*100,1) if total else 0,
        'ofertas_com_agenda': com_agenda, 'total_alunos_matriculados': tot_mat,
        'total_alunos_agendados': tot_agend, 'total_capacidade': tot_capa,
        'pct_ocupacao': round(tot_agend/tot_capa*100,1) if tot_capa else 0,
        'polos_total': df['_POLO'].nunique(),
        'polos_sem_tutor': int(df[~df['_TEM_TUTOR']].groupby('_POLO').ngroups),
    }
    ger_polo = []
    for polo, grp in df.groupby('_POLO'):
        tuts = list(grp[grp['_TEM_TUTOR']]['_TUTOR'].dropna().unique())
        # PATCH 38: dedup por categoria dentro do polo antes de somar alunos —
        # a mesma linha de ALUNOS_MATRICULADOS se repete em cada prática/ordem
        # do mesmo polo+categoria no GIOCONDA; somar direto inflava o total em
        # ~5x (bug reportado pelo Leo: 42 mil vs 8 mil no KPI geral).
        _dedup_p = grp.groupby('_CAT')[['_MAT','_AGEND','_CAPA']].max()
        ger_polo.append({
            'polo': str(polo), 'total_ofertas': len(grp),
            'gerenciadas': int(grp['_GERENCIADO'].sum()),
            'pct_gerenciado': round(grp['_GERENCIADO'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['_TEM_TUTOR'].sum()), 'sem_tutor': int((~grp['_TEM_TUTOR']).sum()),
            'com_agenda': int(grp['_TEM_AGENDA'].sum()),
            'alunos_matriculados': int(_dedup_p['_MAT'].sum()), 'alunos_agendados': int(_dedup_p['_AGEND'].sum()),
            'capacidade': int(_dedup_p['_CAPA'].sum()), 'tutores_unicos': [str(t) for t in tuts],
        })
    ger_polo.sort(key=lambda x: -x['sem_tutor'])
    ger_cat = []
    for cat, grp in df.groupby('_CAT'):
        # PATCH 38: mesma correção — dedup por polo dentro da categoria antes de somar
        _dedup_c = grp.groupby('_POLO')[['_MAT','_AGEND']].max()
        ger_cat.append({
            'categoria': str(cat), 'total_ofertas': len(grp),
            'gerenciadas': int(grp['_GERENCIADO'].sum()),
            'pct_gerenciado': round(grp['_GERENCIADO'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['_TEM_TUTOR'].sum()), 'sem_tutor': int((~grp['_TEM_TUTOR']).sum()),
            'alunos_matriculados': int(_dedup_c['_MAT'].sum()), 'alunos_agendados': int(_dedup_c['_AGEND'].sum()),
        })
    ger_cat.sort(key=lambda x: -x['total_ofertas'])
    ger_ordem = []; ordem_sort = {'Ordem 1':1,'Ordem 2':2,'Ordem 3':3,'Ordem 4':4,'Ordem 5':5}
    for ordem in sorted(df['_ORDEM'].unique(), key=lambda x: ordem_sort.get(x,9)):
        if not ordem: continue
        grp = df[df['_ORDEM']==ordem]
        # PATCH 38: mesma correção de dedup — dentro de uma ordem, um polo+categoria
        # tem várias práticas distintas, todas repetindo o mesmo ALUNOS_MATRICULADOS
        _dedup_o = grp.groupby(['_POLO','_CAT'])[['_MAT','_AGEND']].max()
        ger_ordem.append({
            'ordem': ordem, 'total_ofertas': len(grp),
            'gerenciadas': int(grp['_GERENCIADO'].sum()),
            'pct_gerenciado': round(grp['_GERENCIADO'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['_TEM_TUTOR'].sum()),
            'alunos_matriculados': int(_dedup_o['_MAT'].sum()), 'alunos_agendados': int(_dedup_o['_AGEND'].sum()),
            'dt_inicio': '', 'dt_fim': PRAZOS_ORDENS.get(ordem,''),
            'tutores_gerenciaram': int(grp[grp['_GERENCIADO']]['_TUTOR'].dropna().nunique()),  # PATCH 86 (P6)
        })
    ger_contratacao = []
    for (polo, cat), grp in df.groupby(['_POLO','_CAT']):
        tuts = list(grp[grp['_TEM_TUTOR']]['_TUTOR'].dropna().unique())
        ger_contratacao.append({
            'polo': str(polo), 'categoria': str(cat), 'total_ofertas': len(grp),
            'tem_tutor': len(tuts)>0, 'tutores': [str(t) for t in tuts],
            'status': 'Contratado' if len(tuts)>0 else 'Sem tutor',
        })
    ger_agendas = []
    for polo, grp in df.groupby('_POLO'):
        total_p = len(grp); com_ag = int(grp['_TEM_AGENDA'].sum())
        datas_por_cat = {}; datas_por_tutor = {}
        for _, row in grp[grp['_TEM_AGENDA']].iterrows():
            d = row['_DT_AG_ISO']; c = row['_CAT']; t = row['_TUTOR']
            if d:
                if d not in datas_por_cat: datas_por_cat[d]=[]
                if c and c not in datas_por_cat[d]: datas_por_cat[d].append(c)
                if d not in datas_por_tutor: datas_por_tutor[d]=[]
                if t and t not in datas_por_tutor[d]: datas_por_tutor[d].append(t)
        ger_agendas.append({
            'polo': str(polo), 'total': total_p, 'com_agenda': com_ag,
            'sem_agenda': total_p-com_ag,
            'pct_agendado': round(com_ag/total_p*100,1) if total_p else 0,
            'datas_agenda': sorted(datas_por_cat.keys()),
            'datas_por_cat': datas_por_cat,
            'datas_por_tutor': datas_por_tutor,  # PATCH 7: preservado
        })
    ger_agendas.sort(key=lambda x: -x['sem_agenda'])
    ger_ofertas = []
    for _, row in df.iterrows():
        ger_ofertas.append({
            'polo': row['_POLO'], 'categoria': row['_CAT'],
            'ordem': row['_ORDEM'], 'pratica': row['_PRATICA'],
            'tutor': row['_TUTOR'], 'tem_tutor': bool(row['_TEM_TUTOR']),
            'tem_agenda': bool(row['_TEM_AGENDA']), 'gerenciado': bool(row['_GERENCIADO']),
            'alunos_mat': int(row['_MAT']), 'alunos_agend': int(row['_AGEND']),
            'dt_agenda': row['_DT_AG_ISO'], 'hr_agenda': row['_HR_AG'],
            'dia_semana': row['_DIA_SEMANA'], 'turno': row['_TURNO'],
            'horario_incomum': bool(row['_HORARIO_INCOMUM']), 'sem_alunos': bool(row['_SEM_ALUNOS']),
            'curso': row['_CURSO'], 'subcurso': row['_SUBCURSO'],
        })
    return {
        'ger_kpis': ger_kpis, 'ger_polo': ger_polo, 'ger_cat': ger_cat,
        'ger_ordem': ger_ordem, 'ger_contratacao': ger_contratacao,
        'ger_agendas': ger_agendas, 'ger_ofertas': ger_ofertas,
    }


# PATCH 32: garante que TODO tutor ativo apareça no gerenciamento (Contratação,
# polo, heatmap, Detalhe), mesmo quando o GIOCONDA ainda não tem NENHUMA oferta
# cadastrada pro polo+categoria dele (lab novo/não provisionado no sistema deles).
# Antes só existia o backfill do PATCH 21 (preenche TUTOR em branco numa oferta
# JÁ existente) — se a oferta nem existisse, o tutor simplesmente não aparecia
# em lugar nenhum do gerenciamento, mesmo estando ativo em todo o resto do
# VinciLab (Ficha dos Tutores, Portfólios etc.).
def _recalcular_agregados_de_ofertas(ofertas):
    """Recalcula ger_kpis/ger_polo/ger_cat/ger_ordem/ger_contratacao/ger_agendas
    a partir de uma lista de ofertas (dicts) — usado depois de injetar ofertas
    sintéticas pra tutores sem nenhuma oferta cadastrada no GIOCONDA ainda."""
    total = len(ofertas)
    com_tutor = sum(1 for o in ofertas if o['tem_tutor'])
    gerenciadas = sum(1 for o in ofertas if o['gerenciado'])
    com_agenda = sum(1 for o in ofertas if o['tem_agenda'])
    tot_mat = sum(o.get('alunos_mat', 0) for o in ofertas)
    tot_agend = sum(o.get('alunos_agend', 0) for o in ofertas)
    polos_set = set(o['polo'] for o in ofertas)
    polos_sem_tutor = set(o['polo'] for o in ofertas if not o['tem_tutor'])

    ger_kpis = {
        'total_ofertas': total, 'ofertas_gerenciadas': gerenciadas,
        'ofertas_nao_gerenciadas': total - gerenciadas,
        'pct_gerenciado': round(gerenciadas / total * 100, 1) if total else 0,
        'ofertas_com_tutor': com_tutor, 'ofertas_sem_tutor': total - com_tutor,
        'pct_com_tutor': round(com_tutor / total * 100, 1) if total else 0,
        'ofertas_com_agenda': com_agenda, 'total_alunos_matriculados': tot_mat,
        'total_alunos_agendados': tot_agend, 'total_capacidade': 0, 'pct_ocupacao': 0,
        'polos_total': len(polos_set), 'polos_sem_tutor': len(polos_sem_tutor),
    }

    polo_map = {}; cat_map = {}; ordem_map = {}; contr_map = {}; agenda_map = {}
    for o in ofertas:
        p = o['polo'] or '—'
        if p not in polo_map:
            polo_map[p] = {'polo': p, 'total_ofertas': 0, 'gerenciadas': 0, 'com_tutor': 0, 'sem_tutor': 0,
                           'com_agenda': 0, 'alunos_matriculados': 0, 'alunos_agendados': 0, 'capacidade': 0, 'tutores_unicos': []}
        pm = polo_map[p]
        pm['total_ofertas'] += 1
        if o['gerenciado']: pm['gerenciadas'] += 1
        if o['tem_tutor']: pm['com_tutor'] += 1
        else: pm['sem_tutor'] += 1
        if o['tem_agenda']: pm['com_agenda'] += 1
        pm['alunos_matriculados'] += o.get('alunos_mat', 0)
        pm['alunos_agendados'] += o.get('alunos_agend', 0)
        if o.get('tutor') and o['tutor'] not in pm['tutores_unicos']:
            pm['tutores_unicos'].append(o['tutor'])

        c = o['categoria'] or '—'
        if c not in cat_map:
            cat_map[c] = {'categoria': c, 'total_ofertas': 0, 'gerenciadas': 0, 'com_tutor': 0, 'sem_tutor': 0,
                          'alunos_matriculados': 0, 'alunos_agendados': 0}
        cm = cat_map[c]
        cm['total_ofertas'] += 1
        if o['gerenciado']: cm['gerenciadas'] += 1
        if o['tem_tutor']: cm['com_tutor'] += 1
        else: cm['sem_tutor'] += 1
        cm['alunos_matriculados'] += o.get('alunos_mat', 0)
        cm['alunos_agendados'] += o.get('alunos_agend', 0)

        od = o.get('ordem') or ''
        if od:
            if od not in ordem_map:
                ordem_map[od] = {'ordem': od, 'total_ofertas': 0, 'gerenciadas': 0, 'com_tutor': 0,
                                  'alunos_matriculados': 0, 'alunos_agendados': 0, 'dt_inicio': '', 'dt_fim': PRAZOS_ORDENS.get(od, ''),
                                  '_tutores_ger_set': set()}
            omp = ordem_map[od]
            omp['total_ofertas'] += 1
            if o['gerenciado']: omp['gerenciadas'] += 1
            if o['tem_tutor']: omp['com_tutor'] += 1
            omp['alunos_matriculados'] += o.get('alunos_mat', 0)
            omp['alunos_agendados'] += o.get('alunos_agend', 0)
            # PATCH 86 (P6): tutores distintos que gerenciaram QUALQUER coisa
            # nesta ordem específica — usado no gráfico "Tutores Gerenciaram
            # por Ordem". Essa função roda DEPOIS da injeção de ofertas
            # sintéticas, então é aqui que o dado realmente chega pro
            # dashboard ao vivo — faltava justamente aqui.
            if o['gerenciado'] and o.get('tutor'):
                omp['_tutores_ger_set'].add(o['tutor'])

        trk = (p, c)
        if trk not in contr_map:
            contr_map[trk] = {'polo': p, 'categoria': c, 'total_ofertas': 0, 'tutores': []}
        contr_map[trk]['total_ofertas'] += 1
        if o.get('tutor') and o['tutor'] not in contr_map[trk]['tutores']:
            contr_map[trk]['tutores'].append(o['tutor'])

        if p not in agenda_map:
            agenda_map[p] = {'polo': p, 'total': 0, 'com_agenda': 0, 'datas_por_cat': {}, 'datas_por_tutor': {}, 'datas_por_horario': {}}
        am = agenda_map[p]
        am['total'] += 1
        if o['tem_agenda']:
            am['com_agenda'] += 1
            d = o.get('dt_agenda')
            if d:
                am['datas_por_cat'].setdefault(d, [])
                if c and c not in am['datas_por_cat'][d]: am['datas_por_cat'][d].append(c)
                am['datas_por_tutor'].setdefault(d, [])
                if o.get('tutor') and o['tutor'] not in am['datas_por_tutor'][d]: am['datas_por_tutor'][d].append(o['tutor'])
                # PATCH 85 (P7): horário do agendamento junto do tutor, pra dar
                # pra ver não só QUEM agendou naquele dia mas A QUE HORAS.
                am['datas_por_horario'].setdefault(d, [])
                _hr_cal = o.get('hr_agenda')
                if _hr_cal and o.get('tutor'):
                    _entrada_cal = f"{o['tutor']} · {_hr_cal}"
                    if _entrada_cal not in am['datas_por_horario'][d]: am['datas_por_horario'][d].append(_entrada_cal)

    for pm in polo_map.values():
        pm['pct_gerenciado'] = round(pm['gerenciadas'] / pm['total_ofertas'] * 100, 1) if pm['total_ofertas'] else 0
    for cm in cat_map.values():
        cm['pct_gerenciado'] = round(cm['gerenciadas'] / cm['total_ofertas'] * 100, 1) if cm['total_ofertas'] else 0
    for omp in ordem_map.values():
        omp['pct_gerenciado'] = round(omp['gerenciadas'] / omp['total_ofertas'] * 100, 1) if omp['total_ofertas'] else 0
        omp['tutores_gerenciaram'] = len(omp.pop('_tutores_ger_set', set()))  # PATCH 86 (P6)

    ger_polo = sorted(polo_map.values(), key=lambda x: -x['sem_tutor'])
    ger_cat = sorted(cat_map.values(), key=lambda x: -x['total_ofertas'])
    _ordem_sort = {'Ordem 1': 1, 'Ordem 2': 2, 'Ordem 3': 3, 'Ordem 4': 4, 'Ordem 5': 5}
    ger_ordem = sorted(ordem_map.values(), key=lambda x: _ordem_sort.get(x['ordem'], 9))
    ger_contratacao = []
    for trk, v in contr_map.items():
        tem_tutor = len(v['tutores']) > 0
        ger_contratacao.append({**v, 'tem_tutor': tem_tutor, 'status': 'Contratado' if tem_tutor else 'Sem tutor'})
    ger_agendas = []
    for p, am in agenda_map.items():
        sem_agenda = am['total'] - am['com_agenda']
        ger_agendas.append({
            'polo': p, 'total': am['total'], 'com_agenda': am['com_agenda'], 'sem_agenda': sem_agenda,
            'pct_agendado': round(am['com_agenda'] / am['total'] * 100, 1) if am['total'] else 0,
            'datas_agenda': sorted(am['datas_por_cat'].keys()),
            'datas_por_cat': am['datas_por_cat'], 'datas_por_tutor': am['datas_por_tutor'],
            'datas_por_horario': am['datas_por_horario'],  # PATCH 85 (P7)
        })
    ger_agendas.sort(key=lambda x: -x['sem_agenda'])

    return {
        'ger_kpis': ger_kpis, 'ger_polo': ger_polo, 'ger_cat': ger_cat,
        'ger_ordem': ger_ordem, 'ger_contratacao': ger_contratacao,
        'ger_agendas': ger_agendas, 'ger_ofertas': ofertas,
    }


def _detectar_gerenciamento_fora_ordem(ofertas, periodos):
    """
    PATCH 97: pra cada oferta gerida com data de gerenciamento conhecida,
    verifica se essa data cai dentro do período oficial de uma ordem ANTERIOR
    à ordem da própria prática — sinal de que o tutor gerenciou uma ordem
    mais avançada antes da hora (agora possível desde que o GIOCONDA parou de
    travar isso). Marca cada oferta com '_anomalia_ordem' (bool) e, quando
    True, '_ordem_esperada_na_data' (qual ordem o período correspondia).
    Não modifica 'gerenciado' nem nenhum outro campo — só sinaliza.

    PATCH 130: segunda checagem, DIFERENTE da de cima — o Leo reportou dado
    de "Ordem 3" aparecendo no Engajamento por Ordem quando a Ordem 3 daquele
    semestre nem começou ainda (hoje ainda está na janela da Ordem 1/2). A
    checagem original só compara a data de gerenciamento contra períodos
    ANTERIORES à própria ordem — não cobre "o período da própria ordem
    começa no futuro, em relação a hoje". Marca '_anomalia_ordem_futura'
    (bool) nesses casos — normalmente indica ordem mal preenchida na origem
    (GIOCONDA/CONTROLE), não gerenciamento real adiantado.
    """
    import datetime as _dt_ord

    def _parse_data_br_ou_iso(s):
        if not s:
            return None
        s = str(s).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                return _dt_ord.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    _ordem_num = {'Ordem 1': 1, 'Ordem 2': 2, 'Ordem 3': 3, 'Ordem 4': 4, 'Ordem 5': 5}
    _periodos_parsed = []
    for _ord_nome, _cfg in (periodos or {}).items():
        _ini = _parse_data_br_ou_iso(_cfg.get('inicio', ''))
        _fim = _parse_data_br_ou_iso(_cfg.get('fim', ''))
        if _ini and _fim:
            _periodos_parsed.append((_ordem_num.get(_ord_nome, 99), _ord_nome, _ini, _fim))
    _periodos_parsed.sort()
    _inicio_por_ordem = {_nome: _ini for _num, _nome, _ini, _fim in _periodos_parsed}
    _hoje = _dt_ord.date.today()

    for o in ofertas:
        o['_anomalia_ordem'] = False
        o['_anomalia_ordem_futura'] = False
        if not o.get('gerenciado'):
            continue
        _ordem_propria = o.get('ordem', '')
        _ordem_propria_num = _ordem_num.get(_ordem_propria)

        # PATCH 130: a própria ordem da prática ainda nem começou (hoje < início do período dela)
        _ini_propria = _inicio_por_ordem.get(_ordem_propria)
        if _ini_propria and _hoje < _ini_propria:
            o['_anomalia_ordem_futura'] = True

        _data_ger = _parse_data_br_ou_iso(o.get('dt_agenda', ''))
        if not _data_ger or not _ordem_propria_num:
            continue
        for _num, _nome, _ini, _fim in _periodos_parsed:
            if _ini <= _data_ger <= _fim and _num < _ordem_propria_num:
                o['_anomalia_ordem'] = True
                o['_ordem_esperada_na_data'] = _nome
                break
    return ofertas

def _injetar_tutores_sem_oferta(ger_dados, tutores_ativos):
    """PATCH 32: injeta uma oferta-placeholder pra cada tutor ativo cujo
    polo+categoria não tem NENHUMA linha no GIOCONDA, e recalcula todos os
    agregados a partir da lista de ofertas resultante."""
    import re as _re_inj, unicodedata as _ud_inj
    def _norm_polo_inj(s):
        # PATCH 45: mesma robustez do PATCH 33 (JS) — sem remover parênteses/
        # acentos aqui, tutores com dado REAL no GIOCONDA (que costuma usar a
        # grafia curta do polo, ex: "Blumenau/SC - Salto Do Norte") recebiam uma
        # linha sintética extra "sem oferta" só porque o CONTROLE usa a grafia
        # mais completa (ex: "...Salto Do Norte (Centro Universitário Dante)"),
        # duplicando o tutor na tela (uma linha real + uma fantasma).
        s = str(s or '').strip()
        s = _re_inj.sub(r'^LAP\s*[-–]\s*', '', s, flags=_re_inj.IGNORECASE)
        s = _re_inj.sub(r'\([^)]*\)', '', s)  # remove parênteses e conteúdo
        s = _ud_inj.normalize('NFD', s)
        s = ''.join(c for c in s if _ud_inj.category(c) != 'Mn')  # remove acentos
        return _re_inj.sub(r'\s+', ' ', s).strip().lower()

    ofertas = list(ger_dados.get('ger_ofertas', []))
    polo_cat_existentes = set((_norm_polo_inj(o['polo']), o['categoria']) for o in ofertas)

    # PATCH 75: a checagem acima (polo, categoria) deveria ter bastado — mas na
    # prática, um tutor real (ex: "Cicero Rosendo da Silva Filho" no GIOCONDA)
    # ainda recebia um placeholder duplicado sob o nome do CONTROLE ("Jose
    # Cicero Rosendo Da Silva Filho", com um nome a mais na frente), mesmo
    # sendo a mesma pessoa e o mesmo polo+categoria. Adiciona uma segunda
    # checagem independente, por NOME (mesma lógica de subsequência do PATCH
    # 74 no JS), como blindagem extra.
    def _normaliza_nome_inj(s):
        s = str(s or '').strip()
        s = _re_inj.sub(r'\s*\(\d+\)\s*$', '', s)  # remove chapa entre parênteses no final
        s = _ud_inj.normalize('NFD', s)
        s = ''.join(c for c in s if _ud_inj.category(c) != 'Mn')
        return _re_inj.sub(r'\s+', ' ', s).strip().lower()

    def _eh_subsequencia_inj(curtos, longos):
        i = 0
        for tok in longos:
            if i < len(curtos) and tok == curtos[i]:
                i += 1
        return i == len(curtos)

    def _nomes_batem_inj(nome_a, nome_b):
        if nome_a == nome_b:
            return True
        ta = nome_a.split()
        tb = nome_b.split()
        if not ta or not tb:
            return False
        # PATCH 76: além da subsequência (nome com parte a mais/a menos), cobre
        # também o caso de MESMO número de partes com uma palavra do meio
        # diferente (ex: "Renata Souza da Silva" vs "Renata Souza De Silva") —
        # a subsequência sozinha não pega isso porque nenhum token "sobra" ou
        # "falta", só troca; primeiro+último nome bate mesmo assim.
        if len(ta) >= 2 and len(tb) >= 2 and ta[0] == tb[0] and ta[-1] == tb[-1]:
            return True
        curtos, longos = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
        if len(curtos) < 2:
            return False  # nome de 1 token só é arriscado demais pra casar por subsequência
        return _eh_subsequencia_inj(curtos, longos)

    # Nomes de tutores que JÁ têm alguma oferta real no GIOCONDA (independente
    # de polo/categoria) — usado como segunda checagem, complementar à de
    # (polo, categoria), pra pegar exatamente o caso acima (nome do CONTROLE
    # com uma parte a mais/a menos do que o nome usado no GIOCONDA).
    nomes_reais_existentes = set()
    for o in ofertas:
        _nm = _normaliza_nome_inj(o.get('tutor', ''))
        if _nm:
            nomes_reais_existentes.add(_nm)

    # PATCH 41: categorias reais válidas (as mesmas que aparecem de verdade no
    # GIOCONDA) — usado pra "abrir" categorias compostas do CONTROLE (ex:
    # "ENGMAKER+QUÍMICA E FÍSICA", usada pra tutores que cobrem os dois cursos)
    # em entradas separadas de categoria real, em vez de vazar a string
    # composta inteira como se fosse uma categoria válida no filtro.
    _CATEGORIAS_REAIS_GIOCONDA = {
        'ENF-INS (Multidisciplinar II)', 'BIO-FAR (Multidisciplinar I)',
        'BIO-FISIO-EST-TO (Multidisciplinar III)', 'QUÍMICA E FÍSICA',
        'ENGMAKER', 'NUTRI (Multidisciplinar IV)',
    }
    def _categorias_validas_para(cat_raw):
        cat_raw = (cat_raw or '').strip()
        if cat_raw in _CATEGORIAS_REAIS_GIOCONDA:
            return [cat_raw]
        partes = [p.strip() for p in cat_raw.split('+') if p.strip() in _CATEGORIAS_REAIS_GIOCONDA]
        return partes or ([cat_raw] if cat_raw else [])

    injetadas = 0
    for t in tutores_ativos:
        # PATCH 40: pseudo-tutores de "Aviso de Portfólio" (submissões que não
        # bateram com nenhum tutor real) não têm _anonimo=True, mas também não
        # são tutores de verdade — sem esse filtro extra, eles vazavam uma
        # categoria fantasma "Aviso de Portfólio" pro filtro de Gerenciamento.
        if t.get('_anonimo') or t.get('c') == 'Aviso de Portfólio' or not t.get('n') or not t.get('p'):
            continue
        for cat_valida in _categorias_validas_para(t.get('c', '')):
            chave = (_norm_polo_inj(t['p']), cat_valida)
            if chave in polo_cat_existentes:
                continue
            nome_alvo = _normaliza_nome_inj(t['n'])
            if any(_nomes_batem_inj(nome_alvo, nm) for nm in nomes_reais_existentes):
                continue
            _cursos_t = t.get('cursos', '') or ''
            _SUBCURSO_LABEL_INJ = {
                'BFI': 'Fisioterapia', 'BTO': 'Terapia Ocupacional',
                'COS-TIP': 'Estética e Cosmética', 'TIP-COS': 'Estética e Cosmética', 'COS': 'Estética e Cosmética',
                'BBI': 'Biomedicina', 'BFR': 'Farmácia',
                'EMF-ISN': 'Enfermagem/Instrumentação', 'NTR': 'Nutrição',
            }
            ofertas.append({
                'polo': t['p'], 'categoria': cat_valida, 'ordem': '',
                'pratica': 'Sem oferta cadastrada no GIOCONDA', 'tutor': t['n'],
                'tem_tutor': True, 'tem_agenda': False, 'gerenciado': False,
                'alunos_mat': 0, 'alunos_agend': 0, 'dt_agenda': '', 'hr_agenda': '',
                'dia_semana': '', 'turno': '', 'horario_incomum': False, 'sem_alunos': False,
                'curso': _cursos_t, 'subcurso': _SUBCURSO_LABEL_INJ.get(_cursos_t, _cursos_t),
                '_sintetico': True,
            })
            polo_cat_existentes.add(chave)
            injetadas += 1
    if injetadas:
        print(f"[{ts()}] Tutores ativos sem nenhuma oferta no GIOCONDA (injetados como placeholder): {injetadas}")
    return _recalcular_agregados_de_ofertas(ofertas)


def processar_gerenciamento_semestres(arquivos, controle_tutor_lookup=None):
    """
    PATCH 18: lê 1+ arquivos de gerenciamento (cada um com um semestre padrão de
    fallback) e devolve {semestre: ger_dados_dict}. Quando o arquivo tem coluna
    SEMESTRE (export novo), usa o valor da própria linha como fonte de verdade —
    não confia só em "qual arquivo é qual semestre".
    arquivos: lista de (path, semestre_fallback)

    PATCH 21: controle_tutor_lookup, se fornecido, é um dict {(polo_norm, categoria): nome}
    usado para preencher o TUTOR quando o próprio GIOCONDA ainda não tem esse campo
    preenchido pra aquela oferta, mas o CONTROLE já tem alguém contratado ali —
    sem isso, um tutor recém-contratado fica invisível na tabela de Detalhe até o
    GIOCONDA "alcançar" o cadastro, mesmo já estando ativo em todas as outras telas.
    """
    import re as _re_local, unicodedata as _ud_local
    def _norm_polo_ger(s):
        # PATCH 45: consistente com _norm_polo_bf — remove parênteses e acentos também
        s = str(s or '').strip()
        s = _re_local.sub(r'^LAP\s*[-–]\s*', '', s, flags=_re_local.IGNORECASE)
        s = _re_local.sub(r'\([^)]*\)', '', s)
        s = _ud_local.normalize('NFD', s)
        s = ''.join(c for c in s if _ud_local.category(c) != 'Mn')
        return _re_local.sub(r'\s+', ' ', s).strip().lower()

    # PATCH 32: nome da prática -> curso específico (BFI/BTO/COS-TIP/BBI/BFR),
    # usado no backfill de tutor pra distinguir múltiplos tutores no mesmo polo
    import unicodedata as _ud_bf
    def _norm_proto_bf(s):
        s = _ud_bf.normalize('NFKC', str(s or ''))
        s = s.replace('–', '-').replace('—', '-')
        return ' '.join(s.split()).strip()
    _NOME_TO_PERFIL_BF = {}
    _nomep_path_bf = os.path.join(SCRIPT_DIR, 'nome_to_perfil.json')
    if os.path.isfile(_nomep_path_bf):
        with open(_nomep_path_bf, encoding='utf-8') as _f_bf:
            _ntp_raw_bf = json.load(_f_bf)
        _NOME_TO_PERFIL_BF = {_norm_proto_bf(k): v for k, v in _ntp_raw_bf.items()}
    def _pratica_de_experimento_bf(v):
        m = _re_local.match(r'O\.\d+:\s*(.*)', str(v or ''))
        return _norm_proto_bf(m.group(1) if m else v)

    frames_novo = []
    resultado = {}
    _backfill_count = 0
    for path, fallback_sem in arquivos:
        if not path or not os.path.isfile(path):
            continue
        try:
            df = _ler_arquivo_gerenciamento(path)
        except Exception as e:
            print(f"[{ts()}] ERRO ao ler {os.path.basename(path)}: {e}")
            continue
        cols_upper = [str(c).upper() for c in df.columns]
        is_novo = 'LABORATORIO' in cols_upper and 'NOME_EXPERIMENTO' in cols_upper
        if not is_novo:
            print(f"[{ts()}] {os.path.basename(path)}: formato ANTIGO — todo o arquivo tratado como {fallback_sem}")
            resultado[fallback_sem] = processar_gerenciamento(path)
            continue
        sem_col = next((c for c in df.columns if str(c).upper() == 'SEMESTRE'), None)
        if sem_col:
            df['_SEM_ROW'] = df[sem_col].astype(str).str.strip()
            _fora = ~df['_SEM_ROW'].isin(ALL_SEMESTRES.keys())
            if _fora.any():
                print(f"[{ts()}] {os.path.basename(path)}: {int(_fora.sum())} linhas com SEMESTRE não reconhecido — usando fallback {fallback_sem}")
            df.loc[_fora, '_SEM_ROW'] = fallback_sem
        else:
            df['_SEM_ROW'] = fallback_sem

        # PATCH 21 + PATCH 32: backfill de TUTOR a partir do CONTROLE quando o
        # GIOCONDA está vazio — tenta primeiro pelo curso específico (inferido a
        # partir do nome da prática via nome_to_perfil.json), caindo pra
        # categoria ampla só quando isso não é possível (categorias sem
        # ambiguidade de curso, ou prática não mapeada).
        if controle_tutor_lookup:
            c_lab_bf = next((c for c in df.columns if str(c).upper() == 'LABORATORIO'), None)
            c_cat_bf = next((c for c in df.columns if str(c).upper() == 'CATEGORIA'), None)
            c_tut_bf = next((c for c in df.columns if str(c).upper() == 'TUTOR'), None)
            c_exp_bf = next((c for c in df.columns if str(c).upper() == 'NOME_EXPERIMENTO'), None)
            if c_lab_bf and c_cat_bf and c_tut_bf:
                _CAT_RAW_NORM_BF = {
                    'FISIO-TO-EST-BIO (Multidisciplinar III)': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
                    'BIO-BIO-FISIO-EST-TO (Multidisciplinar III)': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
                }
                def _corrige_prefixo_bio_duplicado_bf(s):
                    s2 = str(s or '').strip()
                    while s2.upper().startswith('BIO-BIO-'):
                        s2 = s2[4:]
                    return s2
                _tutor_vazio = df[c_tut_bf].isna() | (df[c_tut_bf].astype(str).str.strip().isin(['', 'nan']))
                if _tutor_vazio.any():
                    _polo_norm = df.loc[_tutor_vazio, c_lab_bf].map(_norm_polo_ger)
                    _cat_norm  = df.loc[_tutor_vazio, c_cat_bf].astype(str).str.strip().replace(_CAT_RAW_NORM_BF).apply(_corrige_prefixo_bio_duplicado_bf)
                    if c_exp_bf and _NOME_TO_PERFIL_BF:
                        _praticas_norm = df.loc[_tutor_vazio, c_exp_bf].map(_pratica_de_experimento_bf)
                        _cursos_esp = _praticas_norm.map(lambda p: _NOME_TO_PERFIL_BF.get(p))
                    else:
                        _cursos_esp = [None] * int(_tutor_vazio.sum())
                    _preenchido = []
                    for _pn, _cn, _ce in zip(_polo_norm, _cat_norm, _cursos_esp):
                        _val = controle_tutor_lookup.get((_pn, _ce), '') if _ce else ''
                        if not _val:
                            _val = controle_tutor_lookup.get((_pn, _cn), '')
                        _preenchido.append(_val)
                    df.loc[_tutor_vazio, c_tut_bf] = _preenchido
                    _backfill_count += sum(1 for v in _preenchido if v)
        frames_novo.append(df)
    if _backfill_count:
        print(f"[{ts()}] Backfill de tutor via CONTROLE (GIOCONDA sem tutor preenchido): {_backfill_count} ofertas")
    if frames_novo:
        df_all = pd.concat(frames_novo, ignore_index=True)
        for sem, grp in df_all.groupby('_SEM_ROW'):
            grp2 = grp.drop(columns=['_SEM_ROW'])
            print(f"[{ts()}] Gerenciamento {sem}: {len(grp2)} linhas")
            resultado[sem] = _processar_gerenciamento_novo(grp2)
    return resultado


def processar_gerenciamento(p3):
    print(f"[{ts()}] Lendo gerenciamento...")
    df_g = ler_excel(p3)
    print(f"[{ts()}] Gerenciamento: {len(df_g)} linhas, {len(df_g.columns)} colunas")
    cols_upper = [str(c).upper() for c in df_g.columns]
    is_novo = 'LABORATORIO' in cols_upper and 'NOME_EXPERIMENTO' in cols_upper
    if is_novo:
        print(f"[{ts()}] Formato: NOVO (relatório detalhado)")
        return _processar_gerenciamento_novo(df_g)
    print(f"[{ts()}] Formato: ANTIGO (GIOCONDA)")
    def gcol(df, *partes):
        for c in df.columns:
            cu = str(c).upper()
            if all(p.upper() in cu for p in partes): return c
        return None
    c_polo = gcol(df_g, 'CEEM', 'RSOC') or 'CEEM_RSOC'
    c_cat  = gcol(df_g, 'CATP', 'NOME') or 'CATP_NOME'
    c_lab  = gcol(df_g, 'LABE', 'NOME') or 'LABE_NOME'
    c_curso = gcol(df_g, 'NOME', 'CURS') or 'NOME_CURS'
    c_situ = gcol(df_g, 'SITU') or 'SITU'
    c_alunos = gcol(df_g, 'ALUNOS', 'MATRIC') or 'ALUNOS_MATRICULADOS'
    c_capa_exp = gcol(df_g, 'CAPA', 'EXP') or 'CAPA_EXP'
    c_ofe_cad = gcol(df_g, 'OFE', 'CAD') or 'OFE_CAD'
    c_qtd_alun = gcol(df_g, 'QTD', 'ALUN') or 'QTD_ALUN'
    c_tutor = gcol(df_g, 'TUTOR') or 'TUTOR'
    c_dt_agenda = gcol(df_g, 'DT', 'GERENCIADA') or 'DT_GERENCIADA'
    c_hr_agenda = gcol(df_g, 'HR', 'GERENCIADA') or 'HR_GERENCIADA'
    c_ofex_dtin = gcol(df_g, 'OFEX', 'DTIN') or 'OFEX_DTIN'
    c_ofex_dtfi = gcol(df_g, 'OFEX', 'DTFI') or 'OFEX_DTFI'
    if c_situ in df_g.columns:
        df_g = df_g[df_g[c_situ].astype(str).str.strip().str.upper() == 'ATIVO'].copy()
    print(f"[{ts()}] Gerenciamento após filtro ativos: {len(df_g)} linhas")
    df_g['_ORDEM_G'] = ''; df_g['_PRATICA_G'] = ''
    if c_lab in df_g.columns:
        import re
        def extrair_ordem(val):
            val = str(val or '')
            m = re.match(r'O\.(\d+):\s*(.*)', val)
            if m: return f'Ordem {m.group(1)}', m.group(2).strip()
            return '', val.strip()
        parsed = df_g[c_lab].apply(extrair_ordem)
        df_g['_ORDEM_G'] = parsed.apply(lambda x: x[0])
        df_g['_PRATICA_G'] = parsed.apply(lambda x: x[1])
    # FIX BUG 1: _TEM_TUTOR deve ser definido ANTES de _GERENCIADO
    df_g['_TEM_TUTOR'] = df_g[c_tutor].notna() & (df_g[c_tutor].astype(str).str.strip() != '') & (df_g[c_tutor].astype(str).str.strip().str.upper() != 'NAN')
    df_g['_TEM_AGENDA'] = df_g.get(c_dt_agenda, pd.Series(dtype='object')).notna()
    # PATCH 22: GERENCIADO = tem tutor E tem data de gerenciamento — mesmo
    # critério corrigido do formato NOVO (ver comentário lá). OFERTAS_CADASTRADAS
    # ou status CONCLUÍDO não confirmam que o gerenciamento foi feito de fato.
    df_g['_GERENCIADO'] = df_g['_TEM_TUTOR'] & df_g['_TEM_AGENDA']
    df_g['_ALUNOS_MAT'] = pd.to_numeric(df_g.get(c_alunos, 0), errors='coerce').fillna(0).astype(int)
    df_g['_QTD_ALUN'] = pd.to_numeric(df_g.get(c_qtd_alun, 0), errors='coerce').fillna(0).astype(int)
    df_g['_CAPA'] = pd.to_numeric(df_g.get(c_capa_exp, 0), errors='coerce').fillna(0).astype(int)
    total_ofertas = len(df_g); gerenciadas = int(df_g['_GERENCIADO'].sum())
    com_tutor = int(df_g['_TEM_TUTOR'].sum()); sem_tutor = total_ofertas - com_tutor
    # FIX: Alunos Matriculados — deduplicar por polo×categoria (soma bruta conta os mesmos alunos por ordem)
    # Usar apenas colunas que REALMENTE existem (não fallbacks)
    _c_polo_real = c_polo if (c_polo and c_polo in df_g.columns) else None
    _c_cat_real  = c_cat  if (c_cat  and c_cat  in df_g.columns) else None
    # Se nenhuma das buscas primárias funcionou, tentar qualquer coluna polo/cat
    if not _c_polo_real:
        _c_polo_real = next((c for c in df_g.columns if 'POLO' in str(c).upper() or 'CEEM' in str(c).upper()), None)
    if not _c_cat_real:
        _c_cat_real = next((c for c in df_g.columns if 'CATEG' in str(c).upper() or 'CATP' in str(c).upper()), None)
    _raw_mat = int(df_g['_ALUNOS_MAT'].sum())
    if _c_polo_real and _c_cat_real:
        _dedup_g = df_g.groupby([_c_polo_real, _c_cat_real])[['_ALUNOS_MAT','_QTD_ALUN','_CAPA']].max()
        tot_mat   = int(_dedup_g['_ALUNOS_MAT'].sum())
        tot_agend = int(_dedup_g['_QTD_ALUN'].sum())
        tot_capa  = int(_dedup_g['_CAPA'].sum())
        print(f"[{ts()}] Alunos DEDUPLICADOS por polo×cat: {tot_mat:,} (bruto era {_raw_mat:,}, redução: {_raw_mat-tot_mat:,})")
    else:
        tot_mat = _raw_mat
        tot_agend = int(df_g['_QTD_ALUN'].sum())
        tot_capa  = int(df_g['_CAPA'].sum())
        print(f"[{ts()}] Alunos sem dedup (colunas polo/cat não encontradas): {tot_mat:,}")
    polos_total = df_g[c_polo].nunique() if c_polo in df_g.columns else 0
    polos_sem_tutor_count = int(df_g[~df_g['_TEM_TUTOR']].groupby(c_polo).ngroups) if c_polo in df_g.columns else 0
    ger_kpis = {
        'total_ofertas': total_ofertas, 'ofertas_gerenciadas': gerenciadas,
        'ofertas_nao_gerenciadas': total_ofertas - gerenciadas,
        'pct_gerenciado': round(gerenciadas/total_ofertas*100,1) if total_ofertas else 0,
        'ofertas_com_tutor': com_tutor, 'ofertas_sem_tutor': sem_tutor,
        'pct_com_tutor': round(com_tutor/total_ofertas*100,1) if total_ofertas else 0,
        'ofertas_com_agenda': int(df_g['_TEM_AGENDA'].sum()),
        'total_alunos_matriculados': tot_mat, 'total_alunos_agendados': tot_agend,
        'total_capacidade': tot_capa,
        'pct_ocupacao': round(tot_agend/tot_capa*100,1) if tot_capa else 0,
        'polos_total': polos_total, 'polos_sem_tutor': polos_sem_tutor_count,
    }
    print(f"[{ts()}] Gerenciamento: {total_ofertas} ofertas, {gerenciadas} ger., {sem_tutor} sem tutor")
    ger_polo = []
    if c_polo in df_g.columns:
        for polo, grp in df_g.groupby(c_polo):
            # PATCH 38: dedup por categoria dentro do polo antes de somar alunos
            # (mesma correção aplicada em _processar_gerenciamento_novo)
            _dedup_p = grp.groupby(c_cat)[['_ALUNOS_MAT','_QTD_ALUN','_CAPA']].max() if c_cat in grp.columns else grp[['_ALUNOS_MAT','_QTD_ALUN','_CAPA']].max().to_frame().T
            ger_polo.append({
                'polo': str(polo), 'total_ofertas': len(grp),
                'gerenciadas': int(grp['_GERENCIADO'].sum()),
                'pct_gerenciado': round(grp['_GERENCIADO'].sum()/len(grp)*100,1) if len(grp) else 0,
                'com_tutor': int(grp['_TEM_TUTOR'].sum()), 'sem_tutor': int((~grp['_TEM_TUTOR']).sum()),
                'com_agenda': int(grp['_TEM_AGENDA'].sum()),
                'alunos_matriculados': int(_dedup_p['_ALUNOS_MAT'].sum()), 'alunos_agendados': int(_dedup_p['_QTD_ALUN'].sum()),
                'capacidade': int(_dedup_p['_CAPA'].sum()),
                'tutores_unicos': list(grp[grp['_TEM_TUTOR']][c_tutor].dropna().unique()),
            })
        ger_polo.sort(key=lambda x: -x['sem_tutor'])
    ger_cat = []
    if c_cat in df_g.columns:
        for cat, grp in df_g.groupby(c_cat):
            # PATCH 38: dedup por polo dentro da categoria antes de somar alunos
            _dedup_c = grp.groupby(c_polo)[['_ALUNOS_MAT','_QTD_ALUN']].max() if c_polo in grp.columns else grp[['_ALUNOS_MAT','_QTD_ALUN']].max().to_frame().T
            ger_cat.append({
                'categoria': str(cat), 'total_ofertas': len(grp),
                'gerenciadas': int(grp['_GERENCIADO'].sum()),
                'pct_gerenciado': round(grp['_GERENCIADO'].sum()/len(grp)*100,1) if len(grp) else 0,
                'com_tutor': int(grp['_TEM_TUTOR'].sum()), 'sem_tutor': int((~grp['_TEM_TUTOR']).sum()),
                'alunos_matriculados': int(_dedup_c['_ALUNOS_MAT'].sum()), 'alunos_agendados': int(_dedup_c['_QTD_ALUN'].sum()),
            })
        ger_cat.sort(key=lambda x: -x['total_ofertas'])
    ger_ordem = []
    ordens_validas = [o for o in sorted(df_g['_ORDEM_G'].unique()) if o and 'Ordem' in str(o)]
    for ordem in ordens_validas:
        grp = df_g[df_g['_ORDEM_G'] == ordem]
        datas_inicio = pd.to_datetime(grp.get(c_ofex_dtin, pd.Series(dtype='object')), errors='coerce').dropna()
        datas_fim = pd.to_datetime(grp.get(c_ofex_dtfi, pd.Series(dtype='object')), errors='coerce').dropna()
        dt_inicio = datas_inicio.min().strftime('%d/%m/%Y') if len(datas_inicio) > 0 else ''
        dt_fim = datas_fim.max().strftime('%d/%m/%Y') if len(datas_fim) > 0 else ''
        # PATCH 38: mesma correção de dedup por polo×categoria dentro da ordem
        if c_polo in grp.columns and c_cat in grp.columns:
            _dedup_o = grp.groupby([c_polo, c_cat])[['_ALUNOS_MAT','_QTD_ALUN']].max()
        else:
            _dedup_o = grp[['_ALUNOS_MAT','_QTD_ALUN']].max().to_frame().T
        ger_ordem.append({
            'ordem': ordem, 'total_ofertas': len(grp),
            'gerenciadas': int(grp['_GERENCIADO'].sum()),
            'pct_gerenciado': round(grp['_GERENCIADO'].sum()/len(grp)*100,1) if len(grp) else 0,
            'com_tutor': int(grp['_TEM_TUTOR'].sum()),
            'alunos_matriculados': int(_dedup_o['_ALUNOS_MAT'].sum()), 'alunos_agendados': int(_dedup_o['_QTD_ALUN'].sum()),
            'dt_inicio': dt_inicio, 'dt_fim': dt_fim,
            'tutores_gerenciaram': int(grp[grp['_GERENCIADO']][c_tutor].dropna().nunique()) if c_tutor in grp.columns else 0,  # PATCH 86 (P6)
        })
    ger_contratacao = []
    if c_polo in df_g.columns and c_cat in df_g.columns:
        for (polo, cat), grp in df_g.groupby([c_polo, c_cat]):
            tutores_list = list(grp[grp['_TEM_TUTOR']][c_tutor].dropna().unique())
            ger_contratacao.append({
                'polo': str(polo), 'categoria': str(cat), 'total_ofertas': len(grp),
                'tem_tutor': len(tutores_list)>0, 'tutores': [str(t) for t in tutores_list],
                'status': 'Contratado' if len(tutores_list)>0 else 'Sem tutor',
            })
        ger_contratacao.sort(key=lambda x: (0 if x['tem_tutor'] else 1, x['polo']))
    ger_agendas = []
    if c_polo in df_g.columns:
        for polo, grp in df_g.groupby(c_polo):
            total = len(grp); com_agenda = int(grp['_TEM_AGENDA'].sum()); sem_agenda = total - com_agenda
            datas = []; datas_por_cat = {}; datas_por_tutor = {}
            if c_dt_agenda and c_dt_agenda in grp.columns:
                for _, ag_row in grp[grp['_TEM_AGENDA']].iterrows():
                    dt_val = pd.to_datetime(ag_row.get(c_dt_agenda), errors='coerce')
                    if pd.notna(dt_val):
                        dt_str = dt_val.strftime('%Y-%m-%d')
                        cat_val = str(ag_row.get(c_cat, '') or '')
                        tutor_val = str(ag_row.get(c_tutor, '') or '')
                        if dt_str not in datas: datas.append(dt_str)
                        if cat_val:
                            if dt_str not in datas_por_cat: datas_por_cat[dt_str] = []
                            if cat_val not in datas_por_cat[dt_str]: datas_por_cat[dt_str].append(cat_val)
                        if tutor_val and tutor_val != 'nan':
                            if dt_str not in datas_por_tutor: datas_por_tutor[dt_str] = []
                            if tutor_val not in datas_por_tutor[dt_str]: datas_por_tutor[dt_str].append(tutor_val)
                datas = sorted(set(datas))
            ger_agendas.append({
                'polo': str(polo), 'total': total, 'com_agenda': com_agenda, 'sem_agenda': sem_agenda,
                'pct_agendado': round(com_agenda/total*100, 1) if total else 0,
                'datas_agenda': datas, 'datas_por_cat': datas_por_cat,
                'datas_por_tutor': datas_por_tutor,  # PATCH 7: preservado
            })
        ger_agendas.sort(key=lambda x: -x['sem_agenda'])
    ger_ofertas_detalhe = []
    for _, row in df_g.iterrows():
        ger_ofertas_detalhe.append({
            'polo': str(row.get(c_polo, '')), 'categoria': str(row.get(c_cat, '')),
            'ordem': str(row.get('_ORDEM_G', '')), 'pratica': str(row.get('_PRATICA_G', '')),
            'curso': str(row.get(c_curso, '')),
            # PATCH 115: mesma limpeza do sufixo de chapa aplicada no formato
            # NOVO -- protege esse caminho (formato ANTIGO/GIOCONDA) também.
            'tutor': re.sub(r'\s*\(\d{4,}\)\s*$', '', str(row.get(c_tutor, '')).strip()) if pd.notna(row.get(c_tutor)) else '',
            'gerenciado': bool(row.get('_GERENCIADO', False)),
            'tem_agenda': bool(row.get('_TEM_AGENDA', False)),
            'alunos_mat': int(row.get('_ALUNOS_MAT', 0)), 'alunos_agend': int(row.get('_QTD_ALUN', 0)),
            'capacidade': int(row.get('_CAPA', 0)),
        })
    print(f"[{ts()}] Gerenciamento: {len(ger_polo)} polos, {len(ger_cat)} cats, {len(ger_ordem)} ordens")
    return {
        'ger_kpis': ger_kpis, 'ger_polo': ger_polo, 'ger_cat': ger_cat,
        'ger_ordem': ger_ordem, 'ger_contratacao': ger_contratacao,
        'ger_agendas': ger_agendas, 'ger_ofertas': ger_ofertas_detalhe,
    }





def carregar_alunos_hub(path_csv):
    """
    Lê o relatório de alunos (matrículas) e retorna dict com matrículas distintas
    por polo e por categoria — substitui a contagem inflacionada do GIOCONDA.

    PATCH 43: detecta automaticamente entre dois esquemas de coluna diferentes
    que já circularam com esse mesmo nome de arquivo:
      - Esquema ANTIGO: POLO_HUB, GRUPO_HUB, TUTOR_PRATICA, SITUACAO_SEMESTRE
        (granularidade: 1 linha por matrícula no semestre)
      - Esquema NOVO: POLO, CATEGORIA_LABORATORIO, TUTOR, SITUACAO_OFERTA
        (granularidade: 1 linha por aluno × experimento/prática — o mesmo aluno
        aparece várias vezes, uma por prática; dedup por MATRICULA continua
        sendo o jeito certo de contar "alunos distintos")
    """
    import unicodedata as _ud, re as _re
    if not path_csv or not os.path.isfile(path_csv):
        print(f"[{ts()}] Alunos hub: arquivo não encontrado ({path_csv})")
        return None
    print(f"[{ts()}] Lendo alunos por hub: {os.path.basename(path_csv)}")
    # Verificar se o arquivo é HTML (download falhou) e não CSV real
    try:
        with open(path_csv, 'rb') as _f: _head = _f.read(500)
        if b'<!DOCTYPE' in _head or b'<html' in _head.lower() or b'<HTML' in _head:
            sz = os.path.getsize(path_csv)
            print(f"[{ts()}] ERRO: Relatorio_alunos_por_hub.csv é HTML ({sz} bytes) — o download falhou")
            print(f"[{ts()}] SOLUÇÃO: Atualize o secret URL_ALUNOS_HUB para o formato download.aspx:")
            print(f"[{ts()}]   https://uniasselvi01-my.sharepoint.com/personal/[USUARIO]/_layouts/15/download.aspx?share=[TOKEN]")
            return None
    except: pass
    for enc in ['latin-1', 'utf-8', 'cp1252']:
        try:
            df = pd.read_csv(path_csv, sep=';', encoding=enc, dtype=str)
            if 'MATRICULA' in df.columns: break
        except: continue
    else:
        print(f"[{ts()}] ERRO: não foi possível ler {path_csv}")
        return None

    # PATCH 43: detectar esquema de colunas
    esquema_novo = 'POLO' in df.columns and 'CATEGORIA_LABORATORIO' in df.columns and 'POLO_HUB' not in df.columns
    if esquema_novo:
        col_polo, col_grupo, col_tutor_pratica = 'POLO', 'CATEGORIA_LABORATORIO', 'TUTOR'
        print(f"[{ts()}] Alunos hub: esquema NOVO detectado (POLO/CATEGORIA_LABORATORIO/TUTOR)")
    else:
        col_polo, col_grupo, col_tutor_pratica = 'POLO_HUB', 'GRUPO_HUB', 'TUTOR_PRATICA'
        print(f"[{ts()}] Alunos hub: esquema ANTIGO detectado (POLO_HUB/GRUPO_HUB/TUTOR_PRATICA)")

    # Apenas matrículas confirmadas (só existe no esquema antigo)
    if 'SITUACAO_SEMESTRE' in df.columns:
        df = df[df['SITUACAO_SEMESTRE'].str.strip() == 'Matrícula Confirmada'].copy()

    def _norm(s):
        s = _ud.normalize('NFD', str(s or '').upper().strip())
        s = ''.join(c for c in s if _ud.category(c) != 'Mn')
        s = _re.sub(r'^LAP\s*[-–]\s*', '', s).strip()
        return _re.sub(r'\s+', ' ', s)

    # Mapear categoria bruta (de qualquer um dos dois esquemas) → nossas categorias
    GRUPO_CAT = {
        'MULTIDISCIPLINAR II':              'ENF-INS (Multidisciplinar II)',
        'ENF-INS (MULTIDISCIPLINAR II)':    'ENF-INS (Multidisciplinar II)',
        'MULTIDISCIPLINAR I':               'BIO-FAR (Multidisciplinar I)',
        'BIO-FAR (MULTIDISCIPLINAR I)':     'BIO-FAR (Multidisciplinar I)',
        'MULTIDISCIPLINAR III':             'BIO-FISIO-EST-TO (Multidisciplinar III)',
        'FISIO-TO-EST-BIO (MULTIDISCIPLINAR III)': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
        'BIO-FISIO-EST-TO (MULTIDISCIPLINAR III)':  'BIO-FISIO-EST-TO (Multidisciplinar III)',
        'ENGMAKER+QUIMICA E FISICA':         'QUÍMICA E FÍSICA',
        'QUIMICA E FISICA':                  'QUÍMICA E FÍSICA',
        'ENGMAKER':                          'ENGMAKER',
        'MULTIDISCIPLINAR IV':               'NUTRI (Multidisciplinar IV)',
        'NUTRI (MULTIDISCIPLINAR IV)':       'NUTRI (Multidisciplinar IV)',
    }
    def _grupo_para_cat(g):
        gn = _norm(g)
        # Match EXATO primeiro (evita 'MULTIDISCIPLINAR I' casar com 'MULTIDISCIPLINAR II')
        for k, v in GRUPO_CAT.items():
            if _norm(k) == gn: return v
        # Fallback: contém (só para casos como 'ENGMAKER+...' vs 'ENGMAKER')
        for k, v in GRUPO_CAT.items():
            kn = _norm(k)
            if kn in gn and len(kn) > 8: return v
        return g

    df['_POLO_NORM'] = df[col_polo].apply(_norm)
    df['_CAT']       = df[col_grupo].apply(_grupo_para_cat)

    total_distintos = df['MATRICULA'].nunique()
    print(f"[{ts()}] Matrículas DISTINTAS (ativos): {total_distintos:,} (de {len(df):,} linhas)")

    # Por polo (chave normalizada)
    por_polo = (df.groupby('_POLO_NORM')['MATRICULA']
                  .nunique().to_dict())

    # Por polo × categoria
    por_polo_cat = {}
    for (polo, cat), grp in df.groupby(['_POLO_NORM', '_CAT']):
        por_polo_cat[f"{polo}||{cat}"] = int(grp['MATRICULA'].nunique())

    # Por categoria (totais)
    por_cat = (df.groupby('_CAT')['MATRICULA']
                 .nunique().to_dict())

    # ── Mapear TUTOR_PRATICA/TUTOR → subcurso para Multi 3 ──────────────
    tutor_subcurso = {}  # nome_norm → 'Fisio'/'T.Oc'/'Est'
    if col_tutor_pratica in df.columns and 'DISCIPLINA' in df.columns and col_grupo in df.columns:
        import re as _re
        from collections import Counter as _Counter
        _FISIO = ['FISIOTERAPIA','CINESIOTERAPIA','ELETROTERM','CARDIORRESPIR',
                  'PROTESE','ORTESE','RECURSOS TERAPEUTICOS','MOVIMENTO FUNCIONAL',
                  'AVALIACAO FISICO','REABILITACAO','NEUROFUNC','ORTOPEDIC','RESPIRATORIA']
        _TO    = ['TERAPIA OCUPACIONAL','PSICOMOTRICIDADE','INTEGRACAO SENSORIAL',
                  'TRANSTORNOS MENTAIS','COMPORTAMENTO HUMANO','VIDA DIARIA','TRABALHO EM GRUPO']
        _EST   = ['ESTETICA','COSMETOLOGIA','BIOMEDICINA ESTETICA','PIGMENTAC',
                  'DEPILAC','FACIAL CORPORAL','MICROAGULH']
        def _classif_disc(d):
            d2 = _norm(d) if d else ''
            if any(k in d2 for k in _FISIO): return 'Fisio'
            if any(k in d2 for k in _TO):    return 'T.Oc'
            if any(k in d2 for k in _EST):   return 'Est'
            return None
        def _norm_tutor(s):
            s = _re.sub(r'\s*\(\d+\)\s*$', '', str(s or '')).strip()
            return _norm(s)
        df3 = df[df[col_grupo].str.upper().str.contains('MULTIDISCIPLINAR III|MULTI.*3|BIO-FISIO|FISIO-TO-EST', na=False)].copy()
        df3 = df3[df3[col_tutor_pratica].notna() & (df3[col_tutor_pratica].astype(str).str.strip().str.upper() != 'NAN')]
        df3['_sub'] = df3['DISCIPLINA'].apply(_classif_disc)
        df3['_tnorm'] = df3[col_tutor_pratica].apply(_norm_tutor)
        for tutor, grp in df3[df3['_sub'].notna()].groupby('_tnorm'):
            subs = list(grp['_sub'])
            if subs:
                # Guardar em MINÚSCULAS para o JS (que usa normN = toLowerCase)
                tutor_lower = tutor.lower()
                tutor_subcurso[tutor_lower] = _Counter(subs).most_common(1)[0][0]
                # Também guardar primeiro+último nome (fallback)
                parts = tutor_lower.split()
                if len(parts) >= 2:
                    fl = parts[0] + ' ' + parts[-1]
                    if fl not in tutor_subcurso:
                        tutor_subcurso[fl] = tutor_subcurso[tutor_lower]
        print(f"[{ts()}] Subcursos Multi 3 mapeados: {len(tutor_subcurso)} tutores")

    return {
        'total_distintos': int(total_distintos),
        'por_polo': {k: int(v) for k, v in por_polo.items()},
        'por_polo_cat': por_polo_cat,
        'por_cat': {k: int(v) for k, v in por_cat.items()},
        'tutor_subcurso': tutor_subcurso,  # Multi 3: nome_tutor → Fisio/T.Oc/Est
    }

# Senha de acesso ao dashboard (mesma da tela de login)
SENHA_DASHBOARD = "uniasselvi2026"

# PATCH 8: cifra o JSON antes de injetar no HTML — sem isso, dava pra ver
# tudo no Ctrl+U mesmo sem digitar a senha
def cifrar_dados(dados_json_str, senha):
    chave = hashlib.sha256(senha.encode('utf-8')).digest()  # 32 bytes → AES-256
    aesgcm = AESGCM(chave)
    iv = os.urandom(12)
    ct = aesgcm.encrypt(iv, dados_json_str.encode('utf-8'), None)
    iv_b64 = base64.b64encode(iv).decode('ascii')
    ct_b64 = base64.b64encode(ct).decode('ascii')
    return f"{iv_b64}:{ct_b64}"

def gerar_html_coordenadores(dados):
    """
    PATCH 110: gera coordenadores.html -- segunda página no mesmo GitHub Pages,
    mesma senha e mesmo dado cifrado do dashboard principal, mas com uma
    interface separada e simplificada (template_coordenadores.html), travada
    por curso, focada só em "esse polo teve oferta dessa prática ou não".
    """
    saida = os.path.join(SCRIPT_DIR, "saida")
    os.makedirs(saida, exist_ok=True)
    output = os.path.join(saida, "coordenadores.html")
    tmpl   = os.path.join(SCRIPT_DIR, "template_coordenadores.html")
    if not os.path.isfile(tmpl):
        print(f"[{ts()}] AVISO: template_coordenadores.html não encontrado -- pulando geração do portal de coordenadores")
        return
    with open(tmpl, encoding='utf-8') as f: html = f.read()
    json_str = json.dumps(dados, ensure_ascii=False)
    payload_cifrado = cifrar_dados(json_str, SENHA_DASHBOARD)
    html = html.replace("'DATA_GOES_HERE'", json.dumps(payload_cifrado))
    with open(output, 'w', encoding='utf-8') as f: f.write(html)
    print(f"[{ts()}] Salvo: {output} (portal de coordenadores, mesma cifra AES-256-GCM)")

def gerar_html(dados):
    saida = os.path.join(SCRIPT_DIR, "saida")
    os.makedirs(saida, exist_ok=True)
    output = os.path.join(saida, "dashboard.html")
    tmpl   = os.path.join(SCRIPT_DIR, "template_dashboard.html")
    with open(tmpl, encoding='utf-8') as f: html = f.read()
    json_str = json.dumps(dados, ensure_ascii=False)
    payload_cifrado = cifrar_dados(json_str, SENHA_DASHBOARD)
    html = html.replace("'DATA_GOES_HERE'", json.dumps(payload_cifrado))
    html = html.replace("TIMESTAMP_GOES_HERE", dados['gerado_em'])
    with open(output, 'w', encoding='utf-8') as f: f.write(html)
    print(f"[{ts()}] Salvo: {output} (JSON cifrado com AES-256-GCM, {len(payload_cifrado)} chars)")

    # PATCH 9: lookup público (não cifrado) só com email/nome/polo/categoria,
    # pro portfolio_form.html autopreencher sem precisar da senha do dashboard
    lookup = [
        {'email': t.get('email',''), 'n': t.get('n',''), 'p': t.get('p',''), 'c': t.get('c','')}
        for t in dados.get('tutores', [])
        if t.get('email') and not t.get('_anonimo') and t.get('c') != 'Aviso de Portfólio'
    ]
    lookup_path = os.path.join(saida, "lookup.json")
    with open(lookup_path, 'w', encoding='utf-8') as f:
        json.dump(lookup, f, ensure_ascii=False)
    print(f"[{ts()}] Salvo: {lookup_path} ({len(lookup)} tutores, sem cifra)")
    return output


def modo_watch(p1, p2):
    print(f"[{ts()}] Monitorando a cada 30s — feche a janela para parar")
    mods = {p1: 0.0, p2: 0.0}
    def loop():
        while True:
            try:
                mudou = any(os.path.getmtime(a) != mods[a] for a in mods if os.path.isfile(a))
                if mudou:
                    for a in mods:
                        if os.path.isfile(a): mods[a] = os.path.getmtime(a)
                    print(f"[{ts()}] Mudança detectada, atualizando...")
                    gerar_html(processar(p1, p2))
            except Exception as e: print(f"[{ts()}] Erro: {e}")
            time.sleep(30)
    threading.Thread(target=loop, daemon=True).start()
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt: print(f"\n[{ts()}] Encerrado.")


if __name__ == '__main__':
    print()
    print(" Verificando arquivos...")
    print()
    p1, p2, tmpl, p3, p3b, p4, p5, p6 = verificar_e_localizar()
    if not p1 or not p2 or not os.path.isfile(tmpl):
        print()
        print(" Coloque as planilhas na pasta planilhas\\")
        print(" e tente novamente.")
        print()
        if '--sem-browser' not in sys.argv:
            input(" Pressione Enter para sair...")
        sys.exit(1)
    print()
    dados = processar(p1, p2)
    if p4:
        try:
            lotacao = carregar_lotacao(p4)
            dados = enriquecer_tutores(dados, lotacao)
        except Exception as e:
            print(f"[{ts()}] AVISO: Erro ao processar lotação: {e}")
            dados['alunos_por_curso'] = []
    # PATCH 88: aplica o acompanhamento de onboarding, se o arquivo já estiver
    # disponível — cruza por CHAPA (mais confiável) e, se não achar, por
    # nome+polo. Tutor com as 3 colunas em "Sim" vira apto=True (some da lista
    # de treinamento no Vinci); do contrário, apto=False com os 3 flags
    # individuais, pra mostrar os quadradinhos na tela.
    if p6:
        try:
            import pandas as _pd_onb
            _df_onb = _pd_onb.read_excel(p6, sheet_name='Onboarding Tutores')
            _onb_por_chapa = {}
            _onb_por_nomepolo = {}
            for _, _row in _df_onb.iterrows():
                _chapa_onb = str(_row.get('Chapa', '') or '').strip()
                _nome_onb = str(_row.get('Nome do Tutor', '') or '').strip()
                _polo_onb = str(_row.get('Polo', '') or '').strip()
                _flags = {
                    'trilha': str(_row.get('Trilha de Aprendizagem', '') or '').strip().lower() == 'sim',
                    'checklist': str(_row.get('Checklist Realizado', '') or '').strip().lower() == 'sim',
                    'um_a_um': str(_row.get('1:1 de Gerenciamento', '') or '').strip().lower() == 'sim',
                }
                if _chapa_onb:
                    _onb_por_chapa[_chapa_onb] = _flags
                if _nome_onb and _polo_onb:
                    _onb_por_nomepolo[(_nome_onb.lower(), _polo_onb.lower())] = _flags
            _n_aplicados = 0
            for _t in dados.get('tutores', []):
                _chapa_t = str(_t.get('chapa', '') or '').strip()
                _flags_t = _onb_por_chapa.get(_chapa_t)
                if not _flags_t:
                    _key = ((_t.get('n') or '').lower(), (_t.get('p') or '').lower())
                    _flags_t = _onb_por_nomepolo.get(_key)
                if _flags_t:
                    _t['onboarding_trilha'] = _flags_t['trilha']
                    _t['onboarding_checklist'] = _flags_t['checklist']
                    _t['onboarding_1a1'] = _flags_t['um_a_um']
                    _t['onboarding_apto'] = all(_flags_t.values())
                    _n_aplicados += 1
            print(f"[{ts()}] Onboarding: {_n_aplicados} tutores com acompanhamento aplicado")
        except Exception as e:
            print(f"[{ts()}] AVISO: Erro ao processar onboarding: {e}")
    # PATCH 107: processar_vagas(p4) estava aninhado dentro do "if p6:" acima
    # por engano — vagas depende só da Lotação (p4), não tem NADA a ver com o
    # arquivo de onboarding (p6). Como o secret do onboarding ainda não foi
    # configurado (p6 = None), esse bloco inteiro nunca rodava, e a seção RH/
    # Vagas ficava vazia/escondida no Vinci mesmo com a Lotação lida com
    # sucesso. Agora roda independente, sempre que p4 existir.
    if p4:
        try:
            dados['vagas'] = processar_vagas(p4)
        except Exception as e:
            print(f"[{ts()}] AVISO: Erro ao processar vagas: {e}")
            dados['vagas'] = {'vagas': [], 'kpis': {}}
    else:
        dados['alunos_por_curso'] = []
        dados['vagas'] = {'vagas': [], 'kpis': {}}
    # PATCH 89: gera/atualiza a planilha de Acompanhamento de Onboarding —
    # roda sempre (mesmo na primeira vez, quando ainda não existe um p6 pra
    # ler flags antigos) pra garantir que a lista sempre nasce e se mantém
    # sozinha. Preserva flags de quem já está sendo acompanhado, adiciona
    # tutor novo automaticamente, remove quem já ficou "apto".
    try:
        gerar_onboarding_atualizado(p1, p6, os.path.join(SCRIPT_DIR, 'Acompanhamento_Onboarding.xlsx'))
    except Exception as e:
        print(f"[{ts()}] AVISO: Erro ao gerar onboarding atualizado: {e}")
    # PATCH 2: tem_lotacao baseado em dados reais (CH > 0 em pelo menos 1 tutor)
    _ch_ok = sum(1 for t in dados.get('tutores', []) if t.get('ch_semanal') and t['ch_semanal'] > 0)
    dados['tem_lotacao'] = _ch_ok > 0
    print(f"[{ts()}] tem_lotacao={dados['tem_lotacao']} ({_ch_ok} tutores com CH SEMANAL)")
    # PATCH 121: estudo do Racional de Insumos por Experimento (aba Insumos).
    # ESTÁTICO por decisão do Leo (21/08) — não é uma fonte que atualiza a
    # cada ciclo de 2h, é um estudo pontual (v10_FINAL) embutido uma vez. Se
    # um dia vier uma versão nova do estudo, é só substituir o
    # insumos_estudo.json no repo (mesmo padrão do catalogo_oficial.json) e
    # rodar o pipeline — não precisa mexer em código.
    _insumos_file = os.path.join(SCRIPT_DIR, 'insumos_estudo.json')
    if os.path.isfile(_insumos_file):
        with open(_insumos_file, encoding='utf-8') as f:
            dados['insumos'] = json.load(f)
        print(f"[{ts()}] Insumos (estudo estático v10_FINAL): {len(dados['insumos'].get('base', []))} linhas na base unificada")
    else:
        dados['insumos'] = None
        print(f"[{ts()}] AVISO: insumos_estudo.json não encontrado — aba Insumos ficará vazia")
    if p3 or p3b:
        try:
            # PATCH 21 + PATCH 32: tutor ativo no CONTROLE pra cada (polo, categoria)
            # E também (polo, curso específico) — usado como backfill quando o
            # GIOCONDA ainda não tem TUTOR preenchido pra essa oferta, mesmo a
            # pessoa já estando contratada/ativa de verdade. A chave por curso
            # específico (ex: COS-TIP) resolve o caso de múltiplos tutores
            # dividindo o mesmo polo sob a mesma categoria ampla (Multi III:
            # Fisioterapia/T.O./Estética no mesmo laboratório) — sem ela, o
            # backfill por (polo,categoria) só conseguia acertar UM dos tutores
            # do polo, e os outros ficavam sem oferta atribuída e sumiam do
            # Detalhe (caso real: Suzieli Alves Rumpel, Estética, Novo
            # Hamburgo/RS, dividindo o polo com uma tutora de Fisio e outra de
            # T.O., todas sob "BIO-FISIO-EST-TO (Multidisciplinar III)").
            import re as _re_bf, unicodedata as _ud_bf
            def _norm_polo_bf(s):
                # PATCH 45: mesma robustez aplicada em _norm_polo_inj — remove
                # parênteses e acentos, não só o prefixo "LAP -".
                s = str(s or '').strip()
                s = _re_bf.sub(r'^LAP\s*[-–]\s*', '', s, flags=_re_bf.IGNORECASE)
                s = _re_bf.sub(r'\([^)]*\)', '', s)
                s = _ud_bf.normalize('NFD', s)
                s = ''.join(c for c in s if _ud_bf.category(c) != 'Mn')
                return _re_bf.sub(r'\s+', ' ', s).strip().lower()
            controle_tutor_lookup = {}
            for _t in dados.get('tutores', []):
                if _t.get('_anonimo') or not _t.get('n') or not _t.get('p'): continue
                _polo_bf = _norm_polo_bf(_t['p'])
                _cursos_t = _t.get('cursos', '') or ''
                if _cursos_t:
                    controle_tutor_lookup.setdefault((_polo_bf, _cursos_t), _t['n'])
                controle_tutor_lookup.setdefault((_polo_bf, _t.get('c', '')), _t['n'])
            print(f"[{ts()}] Lookup de tutores ativos (CONTROLE) pra backfill: {len(controle_tutor_lookup)} chaves polo+categoria/curso")

            # PATCH 18: cada arquivo tem um semestre de fallback (usado só quando a
            # linha não tem coluna SEMESTRE reconhecível) — arquivo antigo -> mais
            # antigo dos semestres carregados; arquivo "_26_02" -> 2026/2 explícito
            _sem_mais_antigo = sorted(ALL_SEMESTRES.keys())[0]
            ger_por_semestre = processar_gerenciamento_semestres([
                (p3,  _sem_mais_antigo),
                (p3b, '2026/2' if '2026/2' in ALL_SEMESTRES else sorted(ALL_SEMESTRES.keys())[-1]),
            ], controle_tutor_lookup=controle_tutor_lookup)
            # PATCH 118 (movido pro backend — antes só existia no JS, duplicado
            # nos dois templates): remove da ORIGEM as linhas do GIOCONDA de
            # tutores sem nenhum vínculo ativo hoje (desligados sem nenhuma
            # correspondência ativa, ou categorizados como "Aviso de
            # Portfólio") — evita que "órfãos" do GIOCONDA sejam contados como
            # tutores extras no Detalhe/Pizza de Gerenciamento (bug reportado
            # pelo Leo: 355 no gráfico vs 348 no Total Tutores). Aplicado ANTES
            # de _injetar_tutores_sem_oferta, senão um tutor ativo de verdade
            # pode ficar "escondido" atrás de um órfão com nome parecido
            # (mesmo efeito colateral encontrado testando a versão em JS).
            # Cuidado: só exclui quando não existe NENHUM registro ativo
            # correspondente ao nome — um tutor desligado pode ter voltado ou
            # trocado de polo e seguir ativo de verdade (casos reais: Aline
            # Camurça Mesquita, Magno Luis Das Neves Rosa — não excluir).
            # Confirmado com o Leo em 21/08.
            import unicodedata as _ud_orf, re as _re_orf
            def _norm_tutor_key_orf(s):
                s = str(s or '').strip()
                s = _re_orf.sub(r'\s*\(\d+\)\s*$', '', s)
                s = _ud_orf.normalize('NFD', s)
                s = ''.join(c for c in s if _ud_orf.category(c) != 'Mn')
                return _re_orf.sub(r'\s+', ' ', s).strip().lower()

            _tutores_ativos_keys_orf = set()
            for _t in dados.get('tutores', []):
                if _t.get('_anonimo') or not _t.get('n') or _t.get('n') == 'Tutor desligado':
                    continue
                _tutores_ativos_keys_orf.add(_norm_tutor_key_orf(_t['n']))

            _sem_vinculo_ativo_keys_orf = set()
            for _td in dados.get('tutores_desligados', []):
                _k_orf = _norm_tutor_key_orf(_td.get('n'))
                if _k_orf and _k_orf not in _tutores_ativos_keys_orf:
                    _sem_vinculo_ativo_keys_orf.add(_k_orf)
            for _t in dados.get('tutores', []):
                if _t.get('c') == 'Aviso de Portfólio':
                    _sem_vinculo_ativo_keys_orf.add(_norm_tutor_key_orf(_t.get('n')))

            if _sem_vinculo_ativo_keys_orf:
                for _sk in list(ger_por_semestre.keys()):
                    _antes_orf = len(ger_por_semestre[_sk].get('ger_ofertas', []))
                    ger_por_semestre[_sk]['ger_ofertas'] = [
                        o for o in ger_por_semestre[_sk].get('ger_ofertas', [])
                        if not o.get('tutor') or _norm_tutor_key_orf(o['tutor']) not in _sem_vinculo_ativo_keys_orf
                    ]
                    _removidos_orf = _antes_orf - len(ger_por_semestre[_sk]['ger_ofertas'])
                    if _removidos_orf:
                        print(f"[{ts()}] {_sk}: {_removidos_orf} linha(s) de tutor sem vínculo ativo removida(s) do gerenciamento (desligado sem vínculo, ou Aviso de Portfólio)")

            # PATCH 32: garantir que todo tutor ativo apareça no gerenciamento,
            # mesmo sem nenhuma oferta cadastrada no GIOCONDA pro polo dele
            for _sk in list(ger_por_semestre.keys()):
                ger_por_semestre[_sk] = _injetar_tutores_sem_oferta(ger_por_semestre[_sk], dados.get('tutores', []))
            # PATCH 97: o GIOCONDA passou a permitir gerenciar qualquer ordem a
            # qualquer momento (antes travava fora do período vigente). Detecta
            # e sinaliza quando uma prática de uma ordem MAIS AVANÇADA foi
            # gerida enquanto a data ainda caía dentro do período de uma ordem
            # ANTERIOR — ex: gerenciar Ordem 3 enquanto ainda estamos no
            # período oficial da Ordem 2. Não bloqueia nada, só sinaliza pra
            # visualização/auditoria.
            for _sk in list(ger_por_semestre.keys()):
                _periodos_sk = (ALL_SEMESTRES.get(_sk) or {}).get('periodos', {})
                ger_por_semestre[_sk]['ger_ofertas'] = _detectar_gerenciamento_fora_ordem(
                    ger_por_semestre[_sk].get('ger_ofertas', []), _periodos_sk)
                _anomalias_sk = [o for o in ger_por_semestre[_sk]['ger_ofertas'] if o.get('_anomalia_ordem')]
                ger_por_semestre[_sk]['ger_anomalias_ordem'] = _anomalias_sk
                if _anomalias_sk:
                    print(f"[{ts()}] ⚠️  {_sk}: {len(_anomalias_sk)} gerenciamento(s) fora do período esperado da ordem")
                _anomalias_futuras_sk = [o for o in ger_por_semestre[_sk]['ger_ofertas'] if o.get('_anomalia_ordem_futura')]
                if _anomalias_futuras_sk:
                    _ords_futuras = sorted(set(o.get('ordem','?') for o in _anomalias_futuras_sk))
                    print(f"[{ts()}] ⚠️  {_sk}: {len(_anomalias_futuras_sk)} gerenciamento(s) marcado(s) numa ordem que AINDA NÃO COMEÇOU ({', '.join(_ords_futuras)}) — provável erro de preenchimento de ordem na origem, não gerenciamento real adiantado")
            dados['gerenciamento_por_semestre'] = ger_por_semestre
            for _sk, _sv in ger_por_semestre.items():
                print(f"[{ts()}] Gerenciamento {_sk}: {_sv['ger_kpis']['total_ofertas']} ofertas, {_sv['ger_kpis']['ofertas_gerenciadas']} ger.")
            # dados['ger_*'] no nível raiz = semestre ativo do dashboard (compat
            # com todo o código de enriquecimento abaixo, que sempre operou em
            # cima de um único conjunto de ofertas)
            ger_dados = ger_por_semestre.get(SEMESTRE_ATUAL) or next(iter(ger_por_semestre.values()), {})
            dados.update(ger_dados)
            dados['tem_gerenciamento'] = True

            # PATCH 123/127: cruzamento "Registrado no Portfólio × Agendado" —
            # pedido pelo Leo, nos 3 níveis (geral, por polo, por categoria).
            # Duas fontes com vocabulário de categoria DIFERENTE: o Portfólio
            # guarda o texto bruto do formulário ("Multidisciplinar III -
            # Fisioterapia"), o GIOCONDA guarda o rótulo amplo do CAT_MAP
            # ("BIO-FISIO-EST-TO (Multidisciplinar III)"). Reconciliado via
            # categoria_para_curso.json (texto do formulário -> código fino,
            # ex: BFI) + um mapa código fino -> rótulo amplo do CAT_MAP.
            #
            # BUG CORRIGIDO (Leo reportou "647% Agendado sobre Registrado",
            # matematicamente impossível): antes comparava TODO o histórico de
            # portfólio (2026/1+2026/2 somados) contra agendados de UM
            # semestre só. Agora calcula o cruzamento PRA CADA semestre
            # separadamente (mesma fonte por_semestre dos dois lados), e o
            # front troca de recorte junto com a aba de semestre — do mesmo
            # jeito que já faz pra gerenciamento_por_semestre.
            _FINO_PARA_AMPLO = {
                'EMF-ISN': 'ENF-INS (Multidisciplinar II)', 'BFR': 'BIO-FAR (Multidisciplinar I)',
                'NTR': 'NUTRI (Multidisciplinar IV)', 'BFI': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
                'BTO': 'BIO-FISIO-EST-TO (Multidisciplinar III)', 'COS-TIP': 'BIO-FISIO-EST-TO (Multidisciplinar III)',
                'BBI': 'BIO-FISIO-EST-TO (Multidisciplinar III)', 'AGM': 'QUÍMICA E FÍSICA', 'BAU': 'ENGMAKER',
                'ECE-ENM-ENS-ENG-EEA-GPI-CDE-OBR-SAN-TER-FSA-SLF-QUI': 'ENGMAKER+QUÍMICA E FÍSICA',
            }
            _cpc_file2 = os.path.join(SCRIPT_DIR, 'categoria_para_curso.json')
            _cat_para_curso2 = json.load(open(_cpc_file2, encoding='utf-8')) if os.path.isfile(_cpc_file2) else {}

            def _cruza(port, agend):
                # PATCH 133: a direção estava invertida — o Leo esclareceu:
                # "de X agendados, Y (Z% dos agendados) foi registrado no
                # portfólio/compareceu à prática". Ou seja, AGENDADO é a base
                # (quem deveria comparecer) e REGISTRADO é quanto disso
                # realmente aconteceu — uma taxa de comparecimento, não o
                # contrário. Antes: agendado/portfolio (podia passar de 100%
                # facilmente, sem sentido). Agora: portfolio/agendado (0-100%
                # na imensa maioria dos casos, só passa de 100% se sobrar
                # registro sem agendamento prévio, o que é raro e legítimo).
                return {'registrado_portfolio': port, 'agendado': agend,
                        'diferenca': port - agend,
                        'pct_registrado_sobre_agendado': round(port / agend * 100, 1) if agend else None}

            def _monta_cruzamento(_port_dedup_sem, _ger_dados_sem):
                _port_por_cat_amplo = {}
                for _cat_bruta, _val in _port_dedup_sem.get('por_categoria', {}).items():
                    _cod_fino = _cat_para_curso2.get(_cat_bruta)
                    _amplo = _FINO_PARA_AMPLO.get(_cod_fino) if _cod_fino else None
                    if _amplo:
                        _port_por_cat_amplo[_amplo] = _port_por_cat_amplo.get(_amplo, 0) + _val
                _agend_geral_sem = int((_ger_dados_sem.get('ger_kpis') or {}).get('total_alunos_agendados', 0))
                _agend_por_polo_sem = {p['polo']: p.get('alunos_agendados', 0) for p in _ger_dados_sem.get('ger_polo', [])}
                _agend_por_cat_sem = {c['categoria']: c.get('alunos_agendados', 0) for c in _ger_dados_sem.get('ger_cat', [])}
                _por_polo_c = sorted(
                    [{'polo': _p, **_cruza(_port_dedup_sem.get('por_polo', {}).get(_p, 0), _agend_por_polo_sem.get(_p, 0))}
                     for _p in set(_port_dedup_sem.get('por_polo', {}).keys()) | set(_agend_por_polo_sem.keys())],
                    key=lambda x: -x['registrado_portfolio'])
                _por_cat_c = sorted(
                    [{'categoria': _c, **_cruza(_port_por_cat_amplo.get(_c, 0), _agend_por_cat_sem.get(_c, 0))}
                     for _c in set(_port_por_cat_amplo.keys()) | set(_agend_por_cat_sem.keys())],
                    key=lambda x: -x['registrado_portfolio'])
                return {'geral': _cruza(_port_dedup_sem.get('geral', 0), _agend_geral_sem),
                        'por_polo': _por_polo_c, 'por_categoria': _por_cat_c}

            _portfolio_dedup_sem_dict = dados.get('portfolio_alunos_dedup_por_semestre', {})
            cruzamento_por_semestre = {}
            for _sk in ger_por_semestre.keys():
                _port_sem = _portfolio_dedup_sem_dict.get(_sk, {'geral': 0, 'por_polo': {}, 'por_categoria': {}})
                cruzamento_por_semestre[_sk] = _monta_cruzamento(_port_sem, ger_por_semestre[_sk])
            # "Ambos": soma geral dos dois lados através de todos os semestres,
            # inclusive por polo e por categoria (não deixar como lista vazia,
            # senão a tabela mostra "Agendado: 0" que parece dado real mas é
            # só ausência de soma)
            _agend_por_polo_todos = {}
            _agend_por_cat_todos = {}
            for _sk2 in ger_por_semestre.keys():
                for _p in ger_por_semestre[_sk2].get('ger_polo', []):
                    _agend_por_polo_todos[_p['polo']] = _agend_por_polo_todos.get(_p['polo'], 0) + _p.get('alunos_agendados', 0)
                for _c in ger_por_semestre[_sk2].get('ger_cat', []):
                    _agend_por_cat_todos[_c['categoria']] = _agend_por_cat_todos.get(_c['categoria'], 0) + _c.get('alunos_agendados', 0)
            cruzamento_por_semestre['Ambos'] = _monta_cruzamento(
                _portfolio_dedup_sem_dict.get('Ambos', {'geral': 0, 'por_polo': {}, 'por_categoria': {}}),
                {'ger_kpis': {'total_alunos_agendados': sum((ger_por_semestre[s].get('ger_kpis') or {}).get('total_alunos_agendados', 0) for s in ger_por_semestre)},
                 'ger_polo': [{'polo': k, 'alunos_agendados': v} for k, v in _agend_por_polo_todos.items()],
                 'ger_cat': [{'categoria': k, 'alunos_agendados': v} for k, v in _agend_por_cat_todos.items()]})
            dados['cruzamento_portfolio_agendado_por_semestre'] = cruzamento_por_semestre
            dados['cruzamento_portfolio_agendado'] = cruzamento_por_semestre.get(SEMESTRE_ATUAL, {'geral': _cruza(0,0), 'por_polo': [], 'por_categoria': []})
            print(f"[{ts()}] Cruzamento Portfólio×Agendado ({SEMESTRE_ATUAL}): {dados['cruzamento_portfolio_agendado']['geral']}")
            # ── Injetar gerenciamento nos tutores (ger_pct, ger_ok, ger_total) ──────
            import unicodedata as _ud5, re as _re6
            def _norm_ger2(s):
                s = _ud5.normalize('NFD', str(s or '').lower().strip())
                s = ''.join(ch for ch in s if _ud5.category(ch) != 'Mn')
                s = _re6.sub(r'\s*\(\d+\)\s*$', '', s).strip()
                return _re6.sub(r'\s+', ' ', s)
            def _fl_ger2(s):
                pts = _norm_ger2(s).split()
                return f"{pts[0]} {pts[-1]}" if len(pts) >= 2 else _norm_ger2(s)
            # PATCH 95: mesma lógica de subsequência já usada em outros pontos
            # (nome com parte a mais/a menos, ou uma palavra do meio diferente)
            # — usada aqui como ÚLTIMO fallback, só quando nem o nome exato nem
            # primeiro+último batem. Isso resolve exatamente o padrão dos "32
            # tutores sem gerenciamento vinculado" reportado — provável causa
            # em tutores novos, cujo nome ainda pode estar digitado de forma
            # levemente diferente entre o Controle e o GIOCONDA.
            def _eh_subsequencia_ger2(curtos, longos):
                i = 0
                for tok in longos:
                    if i < len(curtos) and tok == curtos[i]:
                        i += 1
                return i == len(curtos)
            def _nomes_batem_ger2(nome_a, nome_b):
                if nome_a == nome_b:
                    return True
                ta, tb = nome_a.split(), nome_b.split()
                if not ta or not tb:
                    return False
                if len(ta) >= 2 and len(tb) >= 2 and ta[0] == tb[0] and ta[-1] == tb[-1]:
                    return True
                curtos, longos = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
                if len(curtos) < 2:
                    return False
                return _eh_subsequencia_ger2(curtos, longos)
            _ger_idx2 = {}
            for _g2 in dados.get('ger_ofertas', []):
                _gn2 = (_g2.get('tutor') or '').strip()
                if not _gn2: continue
                for _k2 in [_norm_ger2(_gn2), _fl_ger2(_gn2)]:
                    if _k2 not in _ger_idx2:
                        _ger_idx2[_k2] = {'ger': 0, 'total': 0}
                    _ger_idx2[_k2]['total'] += 1
                    if _g2.get('gerenciado'): _ger_idx2[_k2]['ger'] += 1
            _ger_matched2 = 0
            _ger_matched2_subsequencia = 0
            for _t2 in dados.get('tutores', []):
                _tn2 = _t2.get('n', '')
                _tn2_norm = _norm_ger2(_tn2)
                _gd2 = _ger_idx2.get(_tn2_norm) or _ger_idx2.get(_fl_ger2(_tn2))
                if not _gd2:
                    for _k2n in _ger_idx2:
                        if _nomes_batem_ger2(_tn2_norm, _k2n):
                            _gd2 = _ger_idx2[_k2n]
                            _ger_matched2_subsequencia += 1
                            break
                if _gd2 and _gd2['total'] > 0:
                    _t2['ger_total'] = _gd2['total']
                    _t2['ger_ok']    = _gd2['ger']
                    _t2['ger_pct']   = round(_gd2['ger'] / _gd2['total'] * 100)
                    _ger_matched2 += 1
                else:
                    _t2['ger_total'] = 0
                    _t2['ger_ok']    = 0
                    _t2['ger_pct']   = None
            print(f"[{ts()}] Gerenciamento injetado nos tutores: {_ger_matched2}/{len(dados.get('tutores',[]))} matches ({_ger_matched2_subsequencia} via correspondência por subsequência)")
            # ────────────────────────────────────────────────────────────────────────
            # Enriquecer ger_ofertas com ch_semanal (join por nome normalizado)
            def _norm_nome(s):
                import unicodedata
                s = str(s or '').lower().split('(')[0].strip()
                s = unicodedata.normalize('NFD', s)
                s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
                return ' '.join(s.split())
            def _nome_fl(s):
                pts = _norm_nome(s).split()
                return (pts[0] + ' ' + pts[-1]) if len(pts) >= 2 else _norm_nome(s)
            # Mapear ch_semanal por nome completo E por primeiro+último nome
            _ch_map = {}; _ch_map_fl = {}
            # Fonte 1: portfólio tutores (com CH já enriquecida pela lotação)
            for t in dados.get('tutores', []):
                if t.get('ch_semanal') and t.get('n'):
                    _ch_map[_norm_nome(t['n'])] = t['ch_semanal']
                    _ch_map_fl[_nome_fl(t['n'])] = t['ch_semanal']
            # Fonte 2: lotação DIRETAMENTE (589 tutores vs 298 do portfólio)
            # Fix principal: tutores que estão no GIOCONDA mas não no portfólio
            _lotacao_safe = lotacao if 'lotacao' in dir() and lotacao else {}
            if _lotacao_safe:
                for lot_nome, lot_info in lotacao.items():
                    lot_ch = lot_info.get('ch_semanal', 0) if isinstance(lot_info, dict) else 0
                    if lot_ch:
                        if lot_nome not in _ch_map:
                            _ch_map[lot_nome] = lot_ch
                        lot_fl = _nome_fl(lot_nome)
                        if lot_fl not in _ch_map_fl:
                            _ch_map_fl[lot_fl] = lot_ch
            print(f"[{ts()}] CH map Gerenciamento: {len(_ch_map)} entradas")
            # Injetar ch_semanal em cada oferta
            enr = 0
            enr_subsequencia = 0
            # Pré-computar lista de (nome_normalizado, nome_fl, ch) para lookup rápido
            _lot_list = [(k, _nome_fl(k), v) for k, v in _ch_map.items()]

            # PATCH 98: as antigas "Match 2"/"Match 3" comparavam por SUBSTRING
            # de token (ex: "ana" IN "juliana") — isso podia casar CH de uma
            # pessoa com o nome de outra completamente diferente, só porque um
            # pedaço curto do nome aparecia dentro do outro (achado: Ana Keila
            # Everton Araujo, CH real 8h, aparecendo com 4h — provável cruzamento
            # errado por esse motivo). Troca pra correspondência por
            # SUBSEQUÊNCIA de palavras inteiras (mesma lógica já validada em
            # vários outros pontos do sistema) — exige que cada token bata
            # como PALAVRA COMPLETA, na ordem certa, não como pedaço de outra.
            def _eh_subsequencia_ch2(curtos, longos):
                i = 0
                for tok in longos:
                    if i < len(curtos) and tok == curtos[i]:
                        i += 1
                return i == len(curtos)
            def _nomes_batem_ch2(nome_a, nome_b):
                if nome_a == nome_b:
                    return True
                ta, tb = nome_a.split(), nome_b.split()
                if not ta or not tb:
                    return False
                if len(ta) >= 2 and len(tb) >= 2 and ta[0] == tb[0] and ta[-1] == tb[-1]:
                    return True
                curtos, longos = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
                if len(curtos) < 2:
                    return False
                return _eh_subsequencia_ch2(curtos, longos)

            for oferta in dados.get('ger_ofertas', []):
                tutor = oferta.get('tutor', '')
                if not tutor or oferta.get('ch_semanal'): continue
                tn = _norm_nome(tutor); tfl = _nome_fl(tutor)

                # Match 1: exato ou FL
                ch = _ch_map.get(tn) or _ch_map.get(tfl) or _ch_map_fl.get(tn) or _ch_map_fl.get(tfl)

                # Match 2 (PATCH 98): subsequência de palavras completas — nome
                # com uma parte a mais/a menos, ou uma palavra do meio diferente.
                if not ch:
                    for lot_n, lot_fl, lot_ch in _lot_list:
                        if _nomes_batem_ch2(tn, lot_n):
                            ch = lot_ch; enr_subsequencia += 1; break

                if ch:
                    oferta['ch_semanal'] = ch; enr += 1
            print(f"[{ts()}] CH enriquecida: {enr}/{len(dados.get('ger_ofertas',[]))} ofertas ({enr_subsequencia} via correspondência por subsequência)")
        except Exception as e:
            print(f"[{ts()}] AVISO: Erro ao processar gerenciamento: {e}")
            import traceback; traceback.print_exc()
            dados['tem_gerenciamento'] = False
    else:
        dados['tem_gerenciamento'] = False
    # ── ALUNOS HUB: matrículas distintas ──────────────────────────────────────
    if p5:
        try:
            alunos_hub = carregar_alunos_hub(p5)
            if alunos_hub:
                dados['alunos_hub'] = alunos_hub
                # Sobrescrever total_alunos_matriculados nos ger_kpis
                if 'ger_kpis' in dados:
                    dados['ger_kpis']['total_alunos_matriculados'] = alunos_hub['total_distintos']
                    dados['ger_kpis']['alunos_mat_fonte'] = 'hub_csv'
                    # Atualizar também DB.kpis.total_alunos com o valor correto do hub
                    if 'kpis' in dados:
                        dados['kpis']['total_alunos'] = alunos_hub['total_distintos']
                    print(f"[{ts()}] KPI alunos substituído: {alunos_hub['total_distintos']:,} (matrículas distintas)")

                # PATCH 135: o hub CSV já calcula qual sub-área de
                # Multidisciplinar III (Fisioterapia/T.Ocupacional/Estética)
                # cada tutor leciona ('tutor_subcurso'), mas isso nunca era
                # aplicado de volta nos tutores — o filtro de curso do portal
                # de coordenadores só enxergava a categoria AMPLA
                # (BIO-FISIO-EST-TO), sem separar as 3 sub-áreas. Resultado
                # reportado pelo Leo: selecionar só "Estética" trazia também
                # tutores de Fisioterapia e T.Ocupacional junto (mesma
                # categoria ampla), inflando a contagem de tutores.
                _tutor_subcurso_map = alunos_hub.get('tutor_subcurso', {})
                if _tutor_subcurso_map:
                    import unicodedata as _ud7, re as _re7
                    def _norm_tutor_sub(s):
                        s = _re7.sub(r'\s*\(\d+\)\s*$', '', str(s or '')).strip()
                        s = _ud7.normalize('NFD', s.lower())
                        return ''.join(c for c in s if _ud7.category(c) != 'Mn')
                    # Converte pro MESMO nome de curso usado no seletor do
                    # portal de coordenadores (CURSO_PARA_CATEGORIA no JS) —
                    # não pro rótulo abreviado interno nem pro rótulo usado em
                    # ger_ofertas.subcurso ("Estética e Cosmética"), que é
                    # ligeiramente diferente do nome oficial do curso
                    # ("Estética e Imagem Pessoal") e não bateria no filtro.
                    _SUB_PARA_NOME_CURSO = {
                        'Fisio': 'Fisioterapia', 'T.Oc': 'Terapia Ocupacional', 'Est': 'Estética e Imagem Pessoal',
                    }
                    _enr_sub = 0
                    for _t in dados.get('tutores', []):
                        _nome_norm = _norm_tutor_sub(_t.get('n', ''))
                        _sub = _tutor_subcurso_map.get(_nome_norm)
                        if not _sub:
                            _partes = _nome_norm.split()
                            if len(_partes) >= 2:
                                _sub = _tutor_subcurso_map.get(_partes[0] + ' ' + _partes[-1])
                        if _sub:
                            _t['subcurso'] = _SUB_PARA_NOME_CURSO.get(_sub, _sub)
                            _enr_sub += 1
                    print(f"[{ts()}] Tutores enriquecidos com subcurso Multi III (Fisio/T.Oc/Est): {_enr_sub}")
                # BUG 3 FIX: enriquecer alunos por polo usando hub CSV (por_polo normalizado)
                # Cobre polos com alunos=0 porque TOTAL_ALUNOS está zerado na lotação 2026_2
                import unicodedata as _ud3, re as _re4
                def _norm_polo_hub_main(s):
                    s = _ud3.normalize('NFD', str(s or '').upper().strip())
                    s = ''.join(c for c in s if _ud3.category(c) != 'Mn')
                    s = _re4.sub(r'^LAP\s*[-–]\s*', '', s).strip()
                    return _re4.sub(r'\s+', ' ', s)
                _hub_por_polo = alunos_hub.get('por_polo', {})
                _enr_polo = 0
                for _ps in dados.get('polo_stats', []):
                    if _ps.get('a', _ps.get('alunos', 0)) == 0:
                        _pn = _norm_polo_hub_main(_ps.get('n', _ps.get('polo', _ps.get('POLO', ''))))
                        _al_hub = _hub_por_polo.get(_pn, 0)
                        if _al_hub:
                            _ps['a'] = int(_al_hub)
                            _ps['alunos'] = int(_al_hub)
                            _enr_polo += 1
                print(f"[{ts()}] Polos enriquecidos com alunos (hub CSV): {_enr_polo}")

                # PATCH 126: base de alunos do polo em cada vaga — pedido pelo
                # Leo pra dar contexto de prioridade (vaga aberta num polo com
                # muitos alunos pesa mais que num polo pequeno). Mesma fonte
                # autoritativa (matrículas distintas do hub), mesma normalização
                # de nome de polo já usada acima.
                _enr_vaga = 0
                for _v in dados.get('vagas', {}).get('vagas', []):
                    _pn_v = _norm_polo_hub_main(_v.get('polo', ''))
                    _al_v = _hub_por_polo.get(_pn_v, 0)
                    _v['alunos_polo'] = int(_al_v) if _al_v else 0
                    if _al_v: _enr_vaga += 1
                print(f"[{ts()}] Vagas enriquecidas com base de alunos do polo: {_enr_vaga}/{len(dados.get('vagas', {}).get('vagas', []))}")
        except Exception as e:
            print(f"[{ts()}] AVISO: erro ao ler alunos hub: {e}")
    else:
        print(f"[{ts()}] INFO: Relatorio_alunos_por_hub.csv não encontrado — usando contagem GIOCONDA")

    # PATCH 131: prioridade invertida — antes só usava o hub CSV (matrículas
    # DISTINTAS, deduplicado de verdade) se dados['alunos_por_curso'] tivesse
    # ficado vazio, ou seja, praticamente nunca, porque a Lotação quase
    # sempre tem a coluna "TOTAL ALUNOS" preenchida (mas sem garantia de
    # deduplicação — o Leo reportou "quantidade errada de alunos
    # matriculados"). O mesmo KPI principal de "Alunos" já troca pro hub CSV
    # como fonte de verdade quando disponível ("KPI alunos substituído",
    # ver mais abaixo) — aplicando o mesmo critério aqui, pra "Alunos por
    # Curso"/"Categorias" na Visão Geral bater com o resto do dashboard.
    if 'alunos_hub' in dir():
        _por_cat = alunos_hub.get('por_cat', {}) if alunos_hub else {}
        _CAT_NOME = {
            'ENF-INS (Multidisciplinar II)':          'Enfermagem e Instrumentação Cirúrgica',
            'BIO-FAR (Multidisciplinar I)':           'Biomedicina e Farmácia',
            'BIO-FISIO-EST-TO (Multidisciplinar III)':'Fisioterapia, T.Ocupacional e Estética',
            'NUTRI (Multidisciplinar IV)':            'Nutrição',
            'ENGMAKER':                               'Engenharias e Licenciaturas',
            'QUÍMICA E FÍSICA':                       'Química e Física',
        }
        if _por_cat:
            _total_hub_cat = sum(int(v) for v in _por_cat.values() if v > 0)
            _total_lotacao_cat = sum(x['alunos'] for x in dados.get('alunos_por_curso', []))
            dados['alunos_por_curso'] = [
                {'sigla': k, 'curso': _CAT_NOME.get(k, k), 'alunos': int(v)}
                for k, v in sorted(_por_cat.items(), key=lambda x: -x[1])
                if v > 0
            ]
            print(f"[{ts()}] Alunos por curso: fonte trocada de Lotação ({_total_lotacao_cat:,}) pra hub CSV/matrículas distintas ({_total_hub_cat:,}) — {len(dados['alunos_por_curso'])} categorias")

    html = gerar_html(dados)
    try:
        gerar_html_coordenadores(dados)
    except Exception as e:
        print(f"[{ts()}] AVISO: Erro ao gerar portal de coordenadores: {e}")
    if '--sem-browser' not in sys.argv:
        print(f"[{ts()}] Abrindo navegador...")
        webbrowser.open(Path(html).as_uri())
    if WATCH_MODE: modo_watch(p1, p2)
    else: print(f"[{ts()}] Concluído!")
